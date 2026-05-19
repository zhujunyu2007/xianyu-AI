from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import List, Tuple, Optional, Dict, Any, Callable, Awaitable
from pathlib import Path
from urllib.parse import unquote
from urllib import request as urllib_request, error as urllib_error
import hashlib
import secrets
import time
import json
import os
import re
import uuid
import base64
from datetime import datetime, timedelta
import uvicorn
import pandas as pd
import io
import asyncio
import concurrent.futures
import queue
from collections import defaultdict

import cookie_manager
from db_manager import db_manager
from config import RISK_CONTROL
from file_log_collector import setup_file_logging, get_file_log_collector
from ai_reply_engine import ai_reply_engine
from utils.qr_login import qr_login_manager
from utils.qr_login_lite import qrcode_login_lite
from utils.xianyu_utils import trans_cookies
from utils.image_utils import image_manager
from utils.time_utils import (
    LOCAL_TIMEZONE,
    get_local_now,
    local_date_to_utc_end_exclusive,
    local_date_to_utc_start,
    parse_db_timestamp,
    utc_timestamp_to_local_date_string,
    utc_timestamp_to_local_datetime,
)
from utils.notification_dispatcher import (
    build_face_verify_notification,
    SUPPORTED_NOTIFICATION_TEMPLATE_TYPES,
    dispatch_account_notifications_sync,
    render_notification_template,
    resolve_verification_type_label,
)
from chat_event_hub import chat_event_hub, publish_chat_message
from order_event_hub import order_event_hub, publish_order_update_event

from loguru import logger

# 刮刮乐远程控制路由
try:
    from api_captcha_remote import router as captcha_router
    CAPTCHA_ROUTER_AVAILABLE = True
except ImportError:
    logger.warning("⚠️ api_captcha_remote 未找到，刮刮乐远程控制功能不可用")
    CAPTCHA_ROUTER_AVAILABLE = False

# 关键字文件路径
KEYWORDS_FILE = Path(__file__).parent / "回复关键字.txt"

# 简单的用户认证配置
ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"  # 系统初始化时的默认密码
SESSION_TOKENS = {}  # 存储会话token: {token: {'user_id': int, 'username': str, 'timestamp': float}}
TOKEN_EXPIRE_TIME = 24 * 60 * 60  # token过期时间：24小时

# HTTP Bearer认证
security = HTTPBearer(auto_error=False)

# 扫码登录检查锁 - 防止并发处理同一个session
qr_check_locks = defaultdict(lambda: asyncio.Lock())
qr_check_processed = {}  # 记录已处理的session: {session_id: {'processed': bool, 'timestamp': float}}

# ========================= 防暴力破解配置 =========================
# IP 登录失败记录: {ip: {'attempts': int, 'first_attempt': float, 'last_attempt': float, 'blocked_until': float}}
login_ip_tracker = {}
# 用户名登录失败记录: {username: {'attempts': int, 'first_attempt': float, 'last_attempt': float, 'locked_until': float}}
login_user_tracker = {}
# 永久黑名单IP列表
ip_blacklist = set()

# 验证码存储: {captcha_id: {'code': str, 'created_at': float, 'ip': str}}
captcha_storage = {}
CAPTCHA_EXPIRE_SECONDS = 300  # 验证码5分钟过期
CAPTCHA_REQUIRE_AFTER_FAILURES = 2  # 失败2次后要求验证码

# 防暴力破解参数
BRUTE_FORCE_CONFIG = {
    'ip_max_attempts': 5,           # 单IP最大尝试次数
    'ip_window_seconds': 300,       # IP计数窗口时间（5分钟）
    'ip_block_seconds': 1800,       # IP封禁时间（30分钟）
    'user_max_attempts': 10,        # 单用户名最大尝试次数
    'user_window_seconds': 600,     # 用户名计数窗口时间（10分钟）
    'user_lock_seconds': 3600,      # 用户名锁定时间（1小时）
    'auto_blacklist_threshold': 20, # 自动加入永久黑名单的失败次数阈值
    'response_delay_base': 1,       # 基础响应延迟（秒）
    'response_delay_multiplier': 0.5,  # 每次失败增加的延迟（秒）
    'max_response_delay': 10,       # 最大响应延迟（秒）
    'captcha_require_failures': 2,  # 失败多少次后需要验证码
}

SENSITIVE_FIELD_PATTERNS = [
    re.compile(r'((?:api[_-]?key|secret|token|cookie|password|proxy_pass)\s*[=:]\s*)([^\s,;]+)', re.IGNORECASE),
    re.compile(r'([?&](?:api[_-]?key|secret|token|cookie|password|proxy_pass)=)([^&\s]+)', re.IGNORECASE),
]

ORDER_STATUS_ALIASES = {
    'success': 'completed',
    'finished': 'completed',
    'pending_delivery': 'pending_ship',
    'delivered': 'shipped',
    'closed': 'cancelled',
    'refunded': 'cancelled',
    'canceled': 'cancelled',
    '处理中': 'processing',
    '待付款': 'pending_payment',
    '待发货': 'pending_ship',
    '部分发货': 'partial_success',
    '部分待收尾': 'partial_pending_finalize',
    '已发货': 'shipped',
    '已完成': 'completed',
    '退款中': 'refunding',
    '退款撤销': 'refund_cancelled',
    '已关闭': 'cancelled',
}

SALES_ELIGIBLE_ORDER_STATUSES = {
    'pending_ship',
    'partial_success',
    'partial_pending_finalize',
    'shipped',
    'completed',
}

ORDER_SALES_TIME_SQL = "COALESCE(NULLIF(platform_paid_at, ''), NULLIF(platform_created_at, ''), created_at)"

ORDER_HISTORY_SYNC_JOB_RETENTION_SECONDS = 3600
order_history_sync_jobs: Dict[str, Dict[str, Any]] = {}
order_history_sync_tasks: Dict[str, asyncio.Task] = {}
ANNOUNCEMENT_CACHE_TTL_SECONDS = 300
announcement_cache: Dict[str, Any] = {
    'expires_at': 0.0,
    'current': None,
    'history': [],
    'last_success_current': None,
    'last_success_history': [],
    'has_remote_success': False,
}


def _get_announcement_remote_url() -> str:
    configured_url = str(os.getenv('DASHBOARD_ANNOUNCEMENT_URL') or '').strip()
    if configured_url:
        return configured_url

    owner = str(os.getenv('UPDATE_GITHUB_OWNER') or 'GuDong2003').strip() or 'GuDong2003'
    repo = str(os.getenv('UPDATE_GITHUB_REPO') or 'xianyu-auto-reply-fix').strip() or 'xianyu-auto-reply-fix'
    branch = str(os.getenv('DASHBOARD_ANNOUNCEMENT_BRANCH') or 'main').strip() or 'main'
    file_path = str(os.getenv('DASHBOARD_ANNOUNCEMENT_FILE') or 'announcement.json').strip().lstrip('/')
    return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}"


def _get_announcement_local_path() -> Path:
    file_path = str(os.getenv('DASHBOARD_ANNOUNCEMENT_FILE') or 'announcement.json').strip().lstrip('/')
    return Path(__file__).parent / file_path


def _parse_announcement_datetime(value: Any) -> Optional[datetime]:
    raw_value = str(value or '').strip()
    if not raw_value:
        return None

    normalized_value = raw_value.replace('Z', '+00:00') if raw_value.endswith('Z') else raw_value
    try:
        parsed = datetime.fromisoformat(normalized_value)
    except ValueError:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=LOCAL_TIMEZONE)
    return parsed.astimezone(LOCAL_TIMEZONE)


def _build_announcement_id(payload: Dict[str, Any]) -> str:
    raw_id = str(payload.get('id') or '').strip()
    if raw_id:
        return raw_id

    stable_source = json.dumps(
        {
            'level': str(payload.get('level') or '').strip(),
            'title': str(payload.get('title') or '').strip(),
            'message': str(payload.get('message') or '').strip(),
            'action_text': str(payload.get('action_text') or '').strip(),
            'action_type': str(payload.get('action_type') or '').strip(),
            'action_url': str(payload.get('action_url') or '').strip(),
            'dismissible': payload.get('dismissible', True),
            'start_at': str(payload.get('start_at') or '').strip(),
            'end_at': str(payload.get('end_at') or '').strip(),
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    return f"announcement-{hashlib.sha1(stable_source.encode('utf-8')).hexdigest()[:12]}"


def _coerce_announcement_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    normalized = str(value).strip().lower()
    if normalized in {'1', 'true', 'yes', 'y', 'on', 'enabled'}:
        return True
    if normalized in {'0', 'false', 'no', 'n', 'off', 'disabled', ''}:
        return False
    return default


def _empty_dashboard_announcement_snapshot() -> Dict[str, Any]:
    return {
        'current': None,
        'history': [],
    }


def _normalize_dashboard_announcement_entry(payload: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(payload, dict):
        return None

    enabled = _coerce_announcement_bool(payload.get('enabled'), default=False)
    start_at = _parse_announcement_datetime(payload.get('start_at'))
    end_at = _parse_announcement_datetime(payload.get('end_at'))
    title = str(payload.get('title') or '').strip()
    message = str(payload.get('message') or '').strip()
    summary = str(payload.get('summary') or payload.get('brief') or payload.get('short_message') or '').strip()
    if not title and not message and not summary:
        return None

    level = str(payload.get('level') or 'info').strip().lower()
    if level not in {'info', 'success', 'warning', 'danger'}:
        level = 'info'

    action_type = str(payload.get('action_type') or '').strip().lower()
    if action_type not in {'', 'url', 'changelog', 'update'}:
        action_type = ''

    action_url = str(payload.get('action_url') or '').strip()
    if action_type == 'url' and not action_url:
        action_type = ''

    action_text = str(payload.get('action_text') or '').strip()
    if action_type and not action_text:
        action_text = '查看详情' if action_type == 'url' else '立即查看'
    if not action_type:
        action_text = ''

    published_at = _parse_announcement_datetime(payload.get('published_at'))
    now = get_local_now()
    if not enabled:
        status = 'disabled'
    elif start_at and now < start_at:
        status = 'scheduled'
    elif end_at and now > end_at:
        status = 'expired'
    else:
        status = 'active'

    return {
        'id': _build_announcement_id(payload),
        'enabled': enabled,
        'status': status,
        'level': level,
        'title': title,
        'summary': summary,
        'message': message,
        'action_text': action_text,
        'action_type': action_type,
        'action_url': action_url,
        'dismissible': _coerce_announcement_bool(payload.get('dismissible'), default=True),
        'published_at': published_at.isoformat() if published_at else '',
        'start_at': start_at.isoformat() if start_at else '',
        'end_at': end_at.isoformat() if end_at else '',
    }


def _normalize_dashboard_announcement_snapshot(payload: Any) -> Optional[Dict[str, Any]]:
    announcements_payload = payload if isinstance(payload, list) else payload.get('announcements') if isinstance(payload, dict) else None
    if not isinstance(announcements_payload, list):
        return None

    history: List[Dict[str, Any]] = []
    for item in announcements_payload:
        normalized_item = _normalize_dashboard_announcement_entry(item)
        if normalized_item:
            history.append(normalized_item)

    history.sort(
        key=lambda item: item.get('published_at') or item.get('start_at') or item.get('end_at') or '',
        reverse=True,
    )

    current_id = ''
    for item in history:
        if item.get('status') == 'active':
            current_id = str(item.get('id') or '').strip()
            break

    normalized_history: List[Dict[str, Any]] = []
    current_announcement: Optional[Dict[str, Any]] = None
    for item in history:
        normalized_item = dict(item)
        normalized_item['is_current'] = bool(current_id and normalized_item.get('id') == current_id)
        normalized_history.append(normalized_item)
        if normalized_item['is_current'] and current_announcement is None:
            current_announcement = dict(normalized_item)

    return {
        'current': current_announcement,
        'history': normalized_history,
    }


def _try_load_dashboard_announcement_snapshot_from_remote() -> Tuple[bool, Optional[Dict[str, Any]]]:
    remote_url = _get_announcement_remote_url()
    try:
        request = urllib_request.Request(
            remote_url,
            headers={
                'User-Agent': 'XianyuDashboardAnnouncement/1.0',
                'Accept': 'application/json',
            }
        )
        with urllib_request.urlopen(request, timeout=8) as response:
            status_code = getattr(response, 'status', 200)
            if status_code != 200:
                logger.warning(f"获取远端公告失败: http_status={status_code}, url={remote_url}")
                return False, None
            raw_content = response.read().decode('utf-8')
    except urllib_error.HTTPError as exc:
        logger.warning(f"获取远端公告失败: http_status={exc.code}, url={remote_url}")
        return False, None
    except Exception as exc:
        logger.warning(f"获取远端公告异常: url={remote_url}, error={mask_sensitive_text(exc)}")
        return False, None

    try:
        payload = json.loads(raw_content)
    except json.JSONDecodeError as exc:
        logger.warning(f"解析远端公告失败: url={remote_url}, error={exc}")
        return False, None

    snapshot = _normalize_dashboard_announcement_snapshot(payload)
    if snapshot is None:
        logger.warning(f"远端公告格式无效: url={remote_url}")
        return False, None

    return True, snapshot


def _try_load_dashboard_announcement_snapshot_from_local() -> Optional[Dict[str, Any]]:
    local_path = _get_announcement_local_path()
    if not local_path.exists():
        return None

    try:
        payload = json.loads(local_path.read_text(encoding='utf-8'))
    except Exception as exc:
        logger.warning(f"读取本地公告文件失败: path={local_path}, error={mask_sensitive_text(exc)}")
        return None

    snapshot = _normalize_dashboard_announcement_snapshot(payload)
    if snapshot is None:
        logger.warning(f"本地公告格式无效: path={local_path}")
        return None

    return snapshot


def _get_dashboard_announcement_payload(force_refresh: bool = False) -> Dict[str, Any]:
    now_ts = time.time()
    if not force_refresh and announcement_cache.get('expires_at', 0) > now_ts:
        return {
            'current': announcement_cache.get('current'),
            'history': list(announcement_cache.get('history') or []),
        }

    loaded_remote, remote_snapshot = _try_load_dashboard_announcement_snapshot_from_remote()
    if loaded_remote and remote_snapshot is not None:
        announcement_cache.update({
            'expires_at': now_ts + ANNOUNCEMENT_CACHE_TTL_SECONDS,
            'current': remote_snapshot.get('current'),
            'history': list(remote_snapshot.get('history') or []),
            'last_success_current': remote_snapshot.get('current'),
            'last_success_history': list(remote_snapshot.get('history') or []),
            'has_remote_success': True,
        })
        return remote_snapshot

    if announcement_cache.get('has_remote_success'):
        snapshot = {
            'current': announcement_cache.get('last_success_current'),
            'history': list(announcement_cache.get('last_success_history') or []),
        }
    else:
        snapshot = _try_load_dashboard_announcement_snapshot_from_local() or _empty_dashboard_announcement_snapshot()

    announcement_cache.update({
        'expires_at': now_ts + ANNOUNCEMENT_CACHE_TTL_SECONDS,
        'current': snapshot.get('current'),
        'history': list(snapshot.get('history') or []),
    })
    return snapshot


def mask_sensitive_text(text: Any) -> str:
    raw_text = str(text or '')
    masked_text = raw_text

    def _mask_match(match):
        prefix = match.group(1)
        secret = match.group(2)
        if len(secret) <= 8:
            masked = '***'
        else:
            masked = f"{secret[:3]}***{secret[-2:]}"
        return f"{prefix}{masked}"

    for pattern in SENSITIVE_FIELD_PATTERNS:
        masked_text = pattern.sub(_mask_match, masked_text)

    return masked_text


def mask_cookie_value(cookie_value: str) -> str:
    cookie_value = str(cookie_value or '')
    if not cookie_value:
        return ''
    if len(cookie_value) <= 16:
        return '***'
    return f"{cookie_value[:8]}...{cookie_value[-8:]}"


def mask_secret_value(secret_value: str) -> str:
    secret_value = str(secret_value or '')
    if not secret_value:
        return ''
    if len(secret_value) <= 8:
        return '***'
    return f"{secret_value[:2]}***{secret_value[-2:]}"


def safe_client_error(message: str = '操作失败，请稍后重试') -> str:
    return message


def normalize_order_status_value(status: Any) -> str:
    normalized = str(status or '').strip().lower()
    if not normalized:
        return 'unknown'
    return ORDER_STATUS_ALIASES.get(normalized, normalized)


def is_sales_eligible_order_status(status: Any) -> bool:
    return normalize_order_status_value(status) in SALES_ELIGIBLE_ORDER_STATUSES


def parse_order_amount_value(raw_amount: Any) -> Optional[float]:
    if raw_amount is None:
        return None

    amount_text = str(raw_amount).strip()
    if not amount_text or amount_text.lower() in {'none', 'null', 'nan'}:
        return None

    normalized = re.sub(r'[^\d.-]', '', amount_text)
    if normalized in {'', '-', '.', '-.'}:
        return None

    try:
        return float(normalized)
    except (TypeError, ValueError):
        return None


def format_sse_event(event_name: str, data: Dict[str, Any]) -> str:
    return f"event: {event_name}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def cleanup_login_trackers():
    """清理过期的登录追踪记录"""
    current_time = time.time()
    
    # 清理IP追踪记录
    expired_ips = []
    for ip, data in login_ip_tracker.items():
        # 如果封禁已过期且超出窗口时间，则清理
        if data.get('blocked_until', 0) < current_time:
            if current_time - data.get('last_attempt', 0) > BRUTE_FORCE_CONFIG['ip_window_seconds'] * 2:
                expired_ips.append(ip)
    for ip in expired_ips:
        del login_ip_tracker[ip]
    
    # 清理用户名追踪记录
    expired_users = []
    for username, data in login_user_tracker.items():
        if data.get('locked_until', 0) < current_time:
            if current_time - data.get('last_attempt', 0) > BRUTE_FORCE_CONFIG['user_window_seconds'] * 2:
                expired_users.append(username)
    for username in expired_users:
        del login_user_tracker[username]


def check_ip_blocked(client_ip: str) -> tuple[bool, str, int]:
    """
    检查IP是否被封禁
    返回: (是否封禁, 原因, 剩余封禁秒数)
    """
    # 检查永久黑名单
    if client_ip in ip_blacklist:
        return True, "IP已被永久封禁", -1
    
    current_time = time.time()
    
    if client_ip in login_ip_tracker:
        data = login_ip_tracker[client_ip]
        
        # 检查是否在封禁期内
        if data.get('blocked_until', 0) > current_time:
            remaining = int(data['blocked_until'] - current_time)
            return True, f"IP登录失败次数过多，请{remaining}秒后再试", remaining
        
        # 检查窗口内的失败次数
        if current_time - data.get('first_attempt', 0) <= BRUTE_FORCE_CONFIG['ip_window_seconds']:
            if data.get('attempts', 0) >= BRUTE_FORCE_CONFIG['ip_max_attempts']:
                # 触发封禁
                block_duration = BRUTE_FORCE_CONFIG['ip_block_seconds']
                data['blocked_until'] = current_time + block_duration
                logger.warning(f"🚫 IP {client_ip} 登录失败{data['attempts']}次，封禁{block_duration}秒")
                return True, f"登录失败次数过多，请{block_duration}秒后再试", block_duration
    
    return False, "", 0


def check_user_locked(username: str) -> tuple[bool, str, int]:
    """
    检查用户名是否被锁定
    返回: (是否锁定, 原因, 剩余锁定秒数)
    """
    current_time = time.time()
    
    if username in login_user_tracker:
        data = login_user_tracker[username]
        
        # 检查是否在锁定期内
        if data.get('locked_until', 0) > current_time:
            remaining = int(data['locked_until'] - current_time)
            return True, f"账户已被临时锁定，请{remaining}秒后再试", remaining
        
        # 检查窗口内的失败次数
        if current_time - data.get('first_attempt', 0) <= BRUTE_FORCE_CONFIG['user_window_seconds']:
            if data.get('attempts', 0) >= BRUTE_FORCE_CONFIG['user_max_attempts']:
                # 触发锁定
                lock_duration = BRUTE_FORCE_CONFIG['user_lock_seconds']
                data['locked_until'] = current_time + lock_duration
                logger.warning(f"🔒 用户 {username} 登录失败{data['attempts']}次，锁定{lock_duration}秒")
                return True, f"账户登录失败次数过多，已被临时锁定，请{lock_duration}秒后再试", lock_duration
    
    return False, "", 0


def record_login_failure(client_ip: str, username: str):
    """记录登录失败"""
    current_time = time.time()
    
    # 更新IP记录
    if client_ip not in login_ip_tracker:
        login_ip_tracker[client_ip] = {
            'attempts': 0,
            'first_attempt': current_time,
            'last_attempt': current_time,
            'blocked_until': 0
        }
    
    ip_data = login_ip_tracker[client_ip]
    
    # 如果超出窗口时间，重置计数
    if current_time - ip_data['first_attempt'] > BRUTE_FORCE_CONFIG['ip_window_seconds']:
        ip_data['attempts'] = 0
        ip_data['first_attempt'] = current_time
    
    ip_data['attempts'] += 1
    ip_data['last_attempt'] = current_time
    
    # 检查是否需要加入永久黑名单
    if ip_data['attempts'] >= BRUTE_FORCE_CONFIG['auto_blacklist_threshold']:
        ip_blacklist.add(client_ip)
        logger.error(f"⛔ IP {client_ip} 登录失败{ip_data['attempts']}次，已加入永久黑名单！")
    
    # 更新用户名记录
    if username:
        if username not in login_user_tracker:
            login_user_tracker[username] = {
                'attempts': 0,
                'first_attempt': current_time,
                'last_attempt': current_time,
                'locked_until': 0
            }
        
        user_data = login_user_tracker[username]
        
        # 如果超出窗口时间，重置计数
        if current_time - user_data['first_attempt'] > BRUTE_FORCE_CONFIG['user_window_seconds']:
            user_data['attempts'] = 0
            user_data['first_attempt'] = current_time
        
        user_data['attempts'] += 1
        user_data['last_attempt'] = current_time


def record_login_success(client_ip: str, username: str):
    """记录登录成功，重置计数"""
    if client_ip in login_ip_tracker:
        login_ip_tracker[client_ip]['attempts'] = 0
    if username and username in login_user_tracker:
        login_user_tracker[username]['attempts'] = 0


def get_response_delay(client_ip: str) -> float:
    """计算响应延迟时间（失败次数越多，延迟越长）"""
    if client_ip not in login_ip_tracker:
        return 0
    
    attempts = login_ip_tracker[client_ip].get('attempts', 0)
    if attempts <= 1:
        return 0
    
    delay = BRUTE_FORCE_CONFIG['response_delay_base'] + \
            (attempts - 1) * BRUTE_FORCE_CONFIG['response_delay_multiplier']
    return min(delay, BRUTE_FORCE_CONFIG['max_response_delay'])


def is_captcha_required(client_ip: str) -> bool:
    """检查是否需要验证码"""
    if client_ip not in login_ip_tracker:
        return False
    attempts = login_ip_tracker[client_ip].get('attempts', 0)
    return attempts >= BRUTE_FORCE_CONFIG.get('captcha_require_failures', 2)


def generate_captcha_image(code: str) -> bytes:
    """生成验证码图片"""
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
    import random
    
    # 图片尺寸
    width, height = 150, 50
    
    # 创建图片
    image = Image.new('RGB', (width, height), color=(255, 255, 255))
    draw = ImageDraw.Draw(image)
    
    # 添加干扰线
    for _ in range(5):
        x1 = random.randint(0, width)
        y1 = random.randint(0, height)
        x2 = random.randint(0, width)
        y2 = random.randint(0, height)
        draw.line([(x1, y1), (x2, y2)], fill=(random.randint(100, 200), random.randint(100, 200), random.randint(100, 200)), width=1)
    
    # 添加干扰点
    for _ in range(50):
        x = random.randint(0, width)
        y = random.randint(0, height)
        draw.point((x, y), fill=(random.randint(0, 150), random.randint(0, 150), random.randint(0, 150)))
    
    # 尝试加载字体，如果失败则使用默认字体
    font = None
    font_paths = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/ARIALBD.TTF", 
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
    ]
    
    for font_path in font_paths:
        try:
            font = ImageFont.truetype(font_path, 32)
            break
        except:
            continue
    
    if font is None:
        # 使用默认字体
        font = ImageFont.load_default()
    
    # 绘制验证码字符
    colors = [
        (0, 0, 139),      # 深蓝
        (139, 0, 0),      # 深红
        (0, 100, 0),      # 深绿
        (139, 69, 19),    # 棕色
        (75, 0, 130),     # 靛蓝
    ]
    
    x_offset = 15
    for i, char in enumerate(code):
        # 随机颜色
        color = random.choice(colors)
        # 随机角度（-15到15度）
        angle = random.randint(-15, 15)
        
        # 创建单个字符的图片用于旋转
        char_image = Image.new('RGBA', (35, 45), (255, 255, 255, 0))
        char_draw = ImageDraw.Draw(char_image)
        char_draw.text((5, 5), char, font=font, fill=color)
        
        # 旋转
        char_image = char_image.rotate(angle, expand=False, fillcolor=(255, 255, 255, 0))
        
        # 粘贴到主图
        y_offset = random.randint(2, 10)
        image.paste(char_image, (x_offset, y_offset), char_image)
        x_offset += 28
    
    # 添加轻微模糊
    image = image.filter(ImageFilter.SMOOTH)
    
    # 转换为bytes
    buffer = io.BytesIO()
    image.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer.getvalue()


def generate_captcha_code(length: int = 4) -> str:
    """生成验证码字符串（排除容易混淆的字符）"""
    # 排除 0, O, 1, I, l 等容易混淆的字符
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(secrets.choice(chars) for _ in range(length))


def cleanup_expired_captchas():
    """清理过期的验证码"""
    current_time = time.time()
    expired = [cid for cid, data in captcha_storage.items() 
               if current_time - data['created_at'] > CAPTCHA_EXPIRE_SECONDS]
    for cid in expired:
        del captcha_storage[cid]


def verify_login_captcha(captcha_id: str, captcha_code: str, client_ip: str) -> tuple[bool, str]:
    """
    验证登录验证码
    返回: (是否验证成功, 错误消息)
    """
    if not captcha_id or not captcha_code:
        return False, "请输入验证码"
    
    if captcha_id not in captcha_storage:
        return False, "验证码已过期，请刷新"
    
    captcha_data = captcha_storage[captcha_id]
    
    # 检查是否过期
    if time.time() - captcha_data['created_at'] > CAPTCHA_EXPIRE_SECONDS:
        del captcha_storage[captcha_id]
        return False, "验证码已过期，请刷新"
    
    # 检查IP是否匹配（防止验证码被其他IP使用）
    if captcha_data.get('ip') and captcha_data['ip'] != client_ip:
        return False, "验证码无效，请刷新"
    
    # 验证码比较（忽略大小写）
    if captcha_code.upper() != captcha_data['code'].upper():
        return False, "验证码错误"
    
    # 验证成功后删除验证码（一次性使用）
    del captcha_storage[captcha_id]
    return True, ""


def get_ip_failure_count(client_ip: str) -> int:
    """获取IP的登录失败次数"""
    if client_ip not in login_ip_tracker:
        return 0
    return login_ip_tracker[client_ip].get('attempts', 0)


# 账号密码登录会话管理
password_login_sessions = {}  # {session_id: {'account_id': str, 'account': str, 'show_browser': bool, 'status': str, 'verification_url': str, 'qr_code_url': str, 'slider_instance': object, 'task': asyncio.Task, 'timestamp': float}}
password_login_locks = defaultdict(lambda: asyncio.Lock())
manual_cookie_import_sessions = {}  # {session_id: {'account_id': str, 'status': str, 'verification_url': str, 'screenshot_path': str, 'slider_instance': object, 'task': asyncio.Task, 'timestamp': float}}
manual_cookie_import_locks = defaultdict(lambda: asyncio.Lock())
PASSWORD_LOGIN_TERMINAL_STATUSES = {'success', 'failed', 'cancelled'}

# ── 轻量扫码登录(qr_login_lite)会话表 ───────────────────────────
# value: {state, qr_data_url, error_message, account_info, started_at, finished, user_id}
# state: pending | waiting | success | error | expired
qr_lite_sessions: Dict[str, Dict[str, Any]] = {}
QR_LITE_SESSION_TTL = 600  # 10 分钟未完结即清理

# 不再需要单独的密码初始化，由数据库初始化时处理


def cleanup_qr_check_records():
    """清理过期的扫码检查记录"""
    current_time = time.time()
    expired_sessions = []

    for session_id, record in qr_check_processed.items():
        # 清理超过1小时的记录
        if current_time - record['timestamp'] > 3600:
            expired_sessions.append(session_id)

    for session_id in expired_sessions:
        if session_id in qr_check_processed:
            del qr_check_processed[session_id]
        if session_id in qr_check_locks:
            del qr_check_locks[session_id]


def load_keywords() -> List[Tuple[str, str]]:
    """读取关键字→回复映射表

    文件格式支持：
        关键字<空格/制表符/冒号>回复内容
    忽略空行和以 # 开头的注释行
    """
    mapping: List[Tuple[str, str]] = []
    if not KEYWORDS_FILE.exists():
        return mapping

    with KEYWORDS_FILE.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # 尝试用\t、空格、冒号分隔
            if '\t' in line:
                key, reply = line.split('\t', 1)
            elif ' ' in line:
                key, reply = line.split(' ', 1)
            elif ':' in line:
                key, reply = line.split(':', 1)
            else:
                # 无法解析的行，跳过
                continue
            mapping.append((key.strip(), reply.strip()))
    return mapping


KEYWORDS_MAPPING = load_keywords()


# 认证相关模型
class LoginRequest(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    verification_code: Optional[str] = None
    captcha_id: Optional[str] = None      # 验证码ID
    captcha_code: Optional[str] = None    # 用户输入的验证码


class LoginResponse(BaseModel):
    success: bool
    token: Optional[str] = None
    message: str
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None
    captcha_required: Optional[bool] = None  # 是否需要验证码


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class RegisterRequest(BaseModel):
    username: str
    email: str
    password: str
    verification_code: str


class RegisterResponse(BaseModel):
    success: bool
    message: str


class SendCodeRequest(BaseModel):
    email: str
    session_id: Optional[str] = None
    type: Optional[str] = 'register'  # 'register' 或 'login'


class SendCodeResponse(BaseModel):
    success: bool
    message: str


class CaptchaRequest(BaseModel):
    session_id: str


class CaptchaResponse(BaseModel):
    success: bool
    captcha_image: str
    session_id: str
    message: str


class VerifyCaptchaRequest(BaseModel):
    session_id: str
    captcha_code: str


class VerifyCaptchaResponse(BaseModel):
    success: bool
    message: str


def generate_token() -> str:
    """生成随机token"""
    return secrets.token_urlsafe(32)


def verify_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Optional[Dict[str, Any]]:
    """验证token并返回用户信息"""
    if not credentials:
        return None

    token = credentials.credentials
    if token not in SESSION_TOKENS:
        return None

    token_data = SESSION_TOKENS[token]

    # 检查token是否过期
    if time.time() - token_data['timestamp'] > TOKEN_EXPIRE_TIME:
        del SESSION_TOKENS[token]
        return None

    return token_data


def verify_admin_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    """验证管理员token"""
    user_info = verify_token(credentials)
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")

    # 检查是否是管理员（优先使用is_admin字段，兼容旧的admin用户名判断）
    is_admin = user_info.get('is_admin', False) or user_info['username'] == ADMIN_USERNAME
    if not is_admin:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    return user_info


def require_auth(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    """需要认证的依赖，返回用户信息"""
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")
    return user_info


def get_current_user(user_info: Dict[str, Any] = Depends(require_auth)) -> Dict[str, Any]:
    """获取当前登录用户信息"""
    return user_info


def get_current_user_optional(user_info: Optional[Dict[str, Any]] = Depends(verify_token)) -> Optional[Dict[str, Any]]:
    """获取当前用户信息（可选，不强制要求登录）"""
    return user_info


def get_user_log_prefix(user_info: Dict[str, Any] = None) -> str:
    """获取用户日志前缀"""
    if user_info:
        return f"【{user_info['username']}#{user_info['user_id']}】"
    return "【系统】"


def require_admin(current_user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    """要求管理员权限"""
    # 优先使用is_admin字段，兼容旧的admin用户名判断
    is_admin = current_user.get('is_admin', False) or current_user['username'] == 'admin'
    if not is_admin:
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return current_user


def log_with_user(level: str, message: str, user_info: Dict[str, Any] = None):
    """带用户信息的日志记录"""
    prefix = get_user_log_prefix(user_info)
    full_message = f"{prefix} {message}"

    if level.lower() == 'info':
        logger.info(full_message)
    elif level.lower() == 'error':
        logger.error(full_message)
    elif level.lower() == 'warning':
        logger.warning(full_message)
    elif level.lower() == 'debug':
        logger.debug(full_message)
    else:
        logger.info(full_message)


def match_reply(cookie_id: str, message: str) -> Optional[str]:
    """根据 cookie_id 及消息内容匹配回复
    只有启用的账号才会匹配关键字回复
    """
    mgr = cookie_manager.manager
    if mgr is None:
        return None

    # 检查账号是否启用
    if not mgr.get_cookie_status(cookie_id):
        return None  # 禁用的账号不参与自动回复

    # 优先账号级关键字
    if mgr.get_keywords(cookie_id):
        for k, r in mgr.get_keywords(cookie_id):
            if k in message:
                return r

    # 全局关键字
    for k, r in KEYWORDS_MAPPING:
        if k in message:
            return r
    return None


class RequestModel(BaseModel):
    cookie_id: str
    msg_time: str
    user_url: str
    send_user_id: str
    send_user_name: str
    item_id: str
    send_message: str
    chat_id: str


class ResponseData(BaseModel):
    send_msg: str


class ResponseModel(BaseModel):
    code: int
    data: ResponseData


app = FastAPI(
    title="Xianyu Management API",
    version="1.0.0",
    description="闲鱼管理系统API",
    docs_url="/docs",
    redoc_url="/redoc"
)

# 注册刮刮乐远程控制路由
if CAPTCHA_ROUTER_AVAILABLE:
    app.include_router(captcha_router)
    logger.info("✅ 已注册刮刮乐远程控制路由: /api/captcha")
else:
    logger.warning("⚠️ 刮刮乐远程控制路由未注册")

# 初始化文件日志收集器
setup_file_logging()

# 添加一条测试日志
from loguru import logger
logger.info("Web服务器启动，文件日志收集器已初始化")


# 启动定时任务调度器
@app.on_event("startup")
async def start_scheduled_task_checker():
    """应用启动时开启定时任务检查协程"""
    asyncio.create_task(scheduled_task_checker())
    logger.info("定时任务调度器已启动")


# 添加请求日志中间件
@app.middleware("http")
async def log_requests(request, call_next):
    start_time = time.time()

    # 获取用户信息
    user_info = "未登录"
    try:
        # 从请求头中获取Authorization
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
            if token in SESSION_TOKENS:
                token_data = SESSION_TOKENS[token]
                # 检查token是否过期
                if time.time() - token_data['timestamp'] <= TOKEN_EXPIRE_TIME:
                    user_info = f"【{token_data['username']}#{token_data['user_id']}】"
    except Exception:
        pass

    logger.info(f"🌐 {user_info} API请求: {request.method} {request.url.path}")

    response = await call_next(request)

    process_time = time.time() - start_time
    logger.info(f"✅ {user_info} API响应: {request.method} {request.url.path} - {response.status_code} ({process_time:.3f}s)")

    return response

# 提供前端静态文件
import os
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir, exist_ok=True)

app.mount('/static', StaticFiles(directory=static_dir), name='static')

# 确保图片上传目录存在
uploads_dir = os.path.join(static_dir, 'uploads', 'images')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir, exist_ok=True)
    logger.info(f"创建图片上传目录: {uploads_dir}")

# 健康检查端点
@app.get('/health')
async def health_check():
    """健康检查端点，用于Docker健康检查和负载均衡器"""
    try:
        # 检查Cookie管理器状态
        manager_status = "ok" if cookie_manager.manager is not None else "error"

        # 检查数据库连接
        from db_manager import db_manager
        try:
            db_manager.get_all_cookies()
            db_status = "ok"
        except Exception:
            db_status = "error"

        # 获取系统状态
        import psutil
        cpu_percent = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()

        status = {
            "status": "healthy" if manager_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": time.time(),
            "services": {
                "cookie_manager": manager_status,
                "database": db_status
            },
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory_info.percent,
                "memory_available": memory_info.available
            }
        }

        if status["status"] == "unhealthy":
            raise HTTPException(status_code=503, detail=status)

        return status

    except HTTPException:
        raise
    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e)
        }


# 重定向根路径到登录页面
@app.get('/', response_class=HTMLResponse)
async def root():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# ========================= 验证码API =========================

@app.get('/captcha/generate')
async def generate_captcha(request: Request):
    """生成验证码图片"""
    # 获取客户端IP
    client_ip = request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or \
                request.headers.get('X-Real-IP', '') or \
                request.client.host if request.client else 'unknown'
    
    # 清理过期验证码
    cleanup_expired_captchas()
    
    # 生成验证码
    code = generate_captcha_code(4)
    captcha_id = secrets.token_urlsafe(16)
    
    # 存储验证码
    captcha_storage[captcha_id] = {
        'code': code,
        'created_at': time.time(),
        'ip': client_ip
    }
    
    # 生成图片
    image_bytes = generate_captcha_image(code)
    
    logger.debug(f"🔢 生成验证码: {captcha_id[:8]}... (IP: {client_ip})")
    
    # 返回图片和ID
    return StreamingResponse(
        io.BytesIO(image_bytes),
        media_type="image/png",
        headers={
            "X-Captcha-Id": captcha_id,
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )


@app.get('/captcha/check-required')
async def check_captcha_required(request: Request):
    """检查是否需要验证码"""
    client_ip = request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or \
                request.headers.get('X-Real-IP', '') or \
                request.client.host if request.client else 'unknown'
    
    required = is_captcha_required(client_ip)
    failure_count = get_ip_failure_count(client_ip)
    
    return {
        'required': required,
        'failure_count': failure_count,
        'threshold': BRUTE_FORCE_CONFIG.get('captcha_require_failures', 2)
    }


# ========================= 验证码API结束 =========================


# 登录页面路由
@app.get('/login.html', response_class=HTMLResponse)
async def login_page():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# 注册页面路由
@app.get('/register.html', response_class=HTMLResponse)
async def register_page():
    # 检查注册是否开启
    from db_manager import db_manager
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        return HTMLResponse('''
        <!DOCTYPE html>
        <html>
        <head>
            <title>注册已关闭</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                .message { color: #666; font-size: 18px; }
                .back-link { margin-top: 20px; }
                .back-link a { color: #007bff; text-decoration: none; }
            </style>
        </head>
        <body>
            <h2>🚫 注册功能已关闭</h2>
            <p class="message">系统管理员已关闭用户注册功能</p>
            <div class="back-link">
                <a href="/">← 返回首页</a>
            </div>
        </body>
        </html>
        ''', status_code=403)

    register_path = os.path.join(static_dir, 'register.html')
    if os.path.exists(register_path):
        with open(register_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Register page not found</h3>')


# 管理页面（不需要服务器端认证，由前端JavaScript处理）
@app.get('/admin', response_class=HTMLResponse)
async def admin_page():
    index_path = os.path.join(static_dir, 'index.html')
    if not os.path.exists(index_path):
        return HTMLResponse('<h3>No front-end found</h3>')
    
    # 获取静态文件的修改时间作为版本号，解决浏览器缓存问题
    def get_file_version(file_path, default='1.0.0'):
        """获取文件的版本号（基于修改时间）"""
        if os.path.exists(file_path):
            try:
                mtime = os.path.getmtime(file_path)
                return str(int(mtime))
            except Exception as e:
                logger.warning(f"获取文件 {file_path} 修改时间失败: {e}")
        return default
    
    app_js_path = os.path.join(static_dir, 'js', 'app.js')
    app_css_path = os.path.join(static_dir, 'css', 'app.css')
    
    js_version = get_file_version(app_js_path, '2.2.0')
    css_version = get_file_version(app_css_path, '1.0.0')
    
    try:
        with open(index_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
            # 替换 app.js 的版本号参数
            js_pattern = r'/static/js/app\.js\?v=[^"\'\s>]+'
            js_new_url = f'/static/js/app.js?v={js_version}'
            if re.search(js_pattern, html_content):
                html_content = re.sub(js_pattern, js_new_url, html_content)
                logger.debug(f"已替换 app.js 版本号: {js_version}")
            
            # 为 app.css 添加或更新版本号参数
            css_pattern = r'/static/css/app\.css(\?v=[^"\'\s>]+)?'
            css_new_url = f'/static/css/app.css?v={css_version}'
            html_content = re.sub(css_pattern, css_new_url, html_content)
            
            return HTMLResponse(html_content)
    except Exception as e:
        logger.error(f"读取或处理 index.html 失败: {e}")
        return HTMLResponse('<h3>Error loading page</h3>')
















# 登录接口
@app.post('/login')
async def login(login_request: LoginRequest, request: Request):
    from db_manager import db_manager
    
    # 获取客户端IP（考虑代理）
    client_ip = request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or \
                request.headers.get('X-Real-IP', '') or \
                request.client.host if request.client else 'unknown'
    
    # 定期清理过期记录
    cleanup_login_trackers()
    
    # 检查IP是否被封禁
    ip_blocked, ip_block_reason, ip_remaining = check_ip_blocked(client_ip)
    if ip_blocked:
        logger.warning(f"🚫 IP {client_ip} 尝试登录但已被封禁: {ip_block_reason}")
        return LoginResponse(
            success=False,
            message=ip_block_reason
        )
    
    # 获取登录标识（用户名或邮箱）
    login_identifier = login_request.username or login_request.email or ''
    
    # 检查用户名是否被锁定
    if login_identifier:
        user_locked, user_lock_reason, user_remaining = check_user_locked(login_identifier)
        if user_locked:
            logger.warning(f"🔒 用户 {login_identifier} 尝试登录但账户已锁定 (IP: {client_ip})")
            # 即使锁定也要记录IP的尝试
            record_login_failure(client_ip, login_identifier)
            return LoginResponse(
                success=False,
                message=user_lock_reason,
                captcha_required=True
            )
    
    # 检查是否需要验证码
    captcha_enabled_str = db_manager.get_system_setting('login_captcha_enabled')
    captcha_enabled = captcha_enabled_str == 'true' if captcha_enabled_str is not None else True

    if captcha_enabled:
        # 验证码已开启，需要验证
        captcha_valid, captcha_error = verify_login_captcha(
            login_request.captcha_id,
            login_request.captcha_code,
            client_ip
        )
        if not captcha_valid:
            logger.warning(f"🔢 IP {client_ip} 验证码验证失败: {captcha_error}")
            return LoginResponse(
                success=False,
                message=captcha_error,
                captcha_required=True
            )
        logger.info(f"🔢 IP {client_ip} 验证码验证成功")
    else:
        logger.info(f"🔢 IP {client_ip} 登录验证码已关闭，跳过验证")

    # 判断登录方式
    if login_request.username and login_request.password:
        # 用户名/密码登录
        logger.info(f"【{login_request.username}】尝试用户名登录 (IP: {client_ip})")

        # 统一使用用户表验证（包括admin用户）
        if db_manager.verify_user_password(login_request.username, login_request.password):
            user = db_manager.get_user_by_username(login_request.username)
            if user:
                # 登录成功，重置计数
                record_login_success(client_ip, login_request.username)

                # 获取is_admin状态
                user_is_admin = user.get('is_admin', False)

                # 生成token
                token = generate_token()
                SESSION_TOKENS[token] = {
                    'user_id': user['id'],
                    'username': user['username'],
                    'is_admin': user_is_admin,
                    'timestamp': time.time()
                }

                # 区分管理员和普通用户的日志
                if user_is_admin:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功（管理员）(IP: {client_ip})")
                else:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功 (IP: {client_ip})")

                return LoginResponse(
                    success=True,
                    token=token,
                    message="登录成功",
                    user_id=user['id'],
                    username=user['username'],
                    is_admin=user_is_admin
                )

        # 登录失败，记录失败次数
        record_login_failure(client_ip, login_request.username)
        
        # 计算响应延迟（防止快速暴力破解）
        delay = get_response_delay(client_ip)
        if delay > 0:
            logger.info(f"🐢 IP {client_ip} 登录失败，延迟响应 {delay:.1f} 秒")
            await asyncio.sleep(delay)
        
        logger.warning(f"【{login_request.username}】登录失败：用户名或密码错误 (IP: {client_ip})")
        # 检查下次是否需要验证码
        next_captcha_required = is_captcha_required(client_ip)
        return LoginResponse(
            success=False,
            message="用户名或密码错误",
            captcha_required=next_captcha_required
        )

    elif login_request.email and login_request.password:
        # 邮箱/密码登录
        logger.info(f"【{login_request.email}】尝试邮箱密码登录 (IP: {client_ip})")

        user = db_manager.get_user_by_email(login_request.email)
        if user and db_manager.verify_user_password(user['username'], login_request.password):
            # 登录成功，重置计数
            record_login_success(client_ip, login_request.email)

            # 获取is_admin状态
            user_is_admin = user.get('is_admin', False)

            # 生成token
            token = generate_token()
            SESSION_TOKENS[token] = {
                'user_id': user['id'],
                'username': user['username'],
                'is_admin': user_is_admin,
                'timestamp': time.time()
            }

            if user_is_admin:
                logger.info(f"【{user['username']}#{user['id']}】邮箱登录成功（管理员）(IP: {client_ip})")
            else:
                logger.info(f"【{user['username']}#{user['id']}】邮箱登录成功 (IP: {client_ip})")

            return LoginResponse(
                success=True,
                token=token,
                message="登录成功",
                user_id=user['id'],
                username=user['username'],
                is_admin=user_is_admin
            )

        # 登录失败，记录失败次数
        record_login_failure(client_ip, login_request.email)
        
        # 计算响应延迟
        delay = get_response_delay(client_ip)
        if delay > 0:
            await asyncio.sleep(delay)
        
        logger.warning(f"【{login_request.email}】邮箱登录失败：邮箱或密码错误 (IP: {client_ip})")
        next_captcha_required = is_captcha_required(client_ip)
        return LoginResponse(
            success=False,
            message="邮箱或密码错误",
            captcha_required=next_captcha_required
        )

    elif login_request.email and login_request.verification_code:
        # 邮箱/验证码登录
        logger.info(f"【{login_request.email}】尝试邮箱验证码登录 (IP: {client_ip})")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(login_request.email, login_request.verification_code, 'login'):
            # 验证码错误也记录失败
            record_login_failure(client_ip, login_request.email)
            delay = get_response_delay(client_ip)
            if delay > 0:
                await asyncio.sleep(delay)

            logger.warning(f"【{login_request.email}】验证码登录失败：验证码错误或已过期 (IP: {client_ip})")
            next_captcha_required = is_captcha_required(client_ip)
            return LoginResponse(
                success=False,
                message="验证码错误或已过期",
                captcha_required=next_captcha_required
            )

        # 获取用户信息
        user = db_manager.get_user_by_email(login_request.email)
        if not user:
            logger.warning(f"【{login_request.email}】验证码登录失败：用户不存在 (IP: {client_ip})")
            return LoginResponse(
                success=False,
                message="用户不存在"
            )

        # 登录成功，重置计数
        record_login_success(client_ip, login_request.email)

        # 获取is_admin状态
        user_is_admin = user.get('is_admin', False)

        # 生成token
        token = generate_token()
        SESSION_TOKENS[token] = {
            'user_id': user['id'],
            'username': user['username'],
            'is_admin': user_is_admin,
            'timestamp': time.time()
        }

        if user_is_admin:
            logger.info(f"【{user['username']}#{user['id']}】验证码登录成功（管理员）(IP: {client_ip})")
        else:
            logger.info(f"【{user['username']}#{user['id']}】验证码登录成功 (IP: {client_ip})")

        return LoginResponse(
            success=True,
            token=token,
            message="登录成功",
            user_id=user['id'],
            username=user['username'],
            is_admin=user_is_admin
        )

    else:
        return LoginResponse(
            success=False,
            message="请提供有效的登录信息"
        )


# 验证token接口
@app.get('/verify')
async def verify(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    if user_info:
        return {
            "authenticated": True,
            "user_id": user_info['user_id'],
            "username": user_info['username'],
            "is_admin": user_info.get('is_admin', False) or user_info['username'] == ADMIN_USERNAME
        }
    return {"authenticated": False}


# 登出接口
@app.post('/logout')
async def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if credentials and credentials.credentials in SESSION_TOKENS:
        del SESSION_TOKENS[credentials.credentials]
    return {"message": "已登出"}


# 销售额数据查询接口
@app.get('/api/sales')
async def get_sales_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_info: Optional[Dict[str, Any]] = Depends(verify_token)
):
    """
    获取销售额数据
    - start_date: 开始日期 (格式: YYYY-MM-DD)
    - end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    try:
        from db_manager import db_manager

        current_user_id = (user_info or {}).get('user_id')
        if current_user_id is None:
            raise HTTPException(status_code=401, detail='未登录或登录已过期')

        user_cookies = db_manager.get_all_cookies(current_user_id)
        cookie_ids = list(user_cookies.keys())
        if not cookie_ids:
            return {
                'success': True,
                'data': {
                    'sales': [],
                    'total': 0.0,
                    'count': 0
                },
                'message': '获取销售额数据成功'
            }
        
        # 构建查询
        placeholders = ','.join(['?'] * len(cookie_ids))
        query = (
            f"SELECT amount, {ORDER_SALES_TIME_SQL} AS effective_sales_at, order_status "
            f"FROM orders WHERE cookie_id IN ({placeholders})"
        )
        params = list(cookie_ids)
        
        if start_date:
            utc_start = local_date_to_utc_start(start_date)
            if not utc_start:
                raise HTTPException(status_code=400, detail='开始日期格式错误，应为 YYYY-MM-DD')
            query += f" AND {ORDER_SALES_TIME_SQL} >= ?"
            params.append(utc_start)
        if end_date:
            utc_end_exclusive = local_date_to_utc_end_exclusive(end_date)
            if not utc_end_exclusive:
                raise HTTPException(status_code=400, detail='结束日期格式错误，应为 YYYY-MM-DD')
            query += f" AND {ORDER_SALES_TIME_SQL} < ?"
            params.append(utc_end_exclusive)
        
        # 执行查询
        orders = db_manager.execute_query(query, params)
        
        # 处理数据
        sales_by_date = {}
        total_sales = 0.0
        valid_count = 0
        skipped_invalid_amount = 0
        skipped_ineligible_status = 0

        for order in orders:
            amount_str = order[0]
            effective_sales_at = order[1]
            order_status = order[2]

            if not is_sales_eligible_order_status(order_status):
                skipped_ineligible_status += 1
                continue

            amount = parse_order_amount_value(amount_str)
            if amount is None:
                skipped_invalid_amount += 1
                continue

            local_date = utc_timestamp_to_local_date_string(effective_sales_at)
            if not local_date:
                continue

            total_sales += amount
            valid_count += 1

            if local_date not in sales_by_date:
                sales_by_date[local_date] = 0
            sales_by_date[local_date] += amount

        logger.info(
            f"销售额数据统计完成: valid_count={valid_count}, skipped_invalid_amount={skipped_invalid_amount}, "
            f"skipped_ineligible_status={skipped_ineligible_status}"
        )
        
        # 转换为列表格式
        formatted_data = [
            {
                'date': date,
                'amount': round(amount, 2)
            }
            for date, amount in sorted(sales_by_date.items())
        ]
        
        return {
            'success': True,
            'data': {
                'sales': formatted_data,
                'total': round(total_sales, 2),
                'count': valid_count
            },
            'message': '获取销售额数据成功'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取销售额数据失败: {e}")
        return {
            'success': False,
            'data': None,
            'message': f'获取销售额数据失败: {str(e)}'
        }


# 周销售额和月销售额查询接口
@app.get('/api/sales/summary')
async def get_sales_summary(
    user_info: Optional[Dict[str, Any]] = Depends(verify_token)
):
    """
    获取当日、本周和本月销售额摘要
    """
    try:
        from db_manager import db_manager

        current_user_id = (user_info or {}).get('user_id')
        if current_user_id is None:
            raise HTTPException(status_code=401, detail='未登录或登录已过期')

        user_cookies = db_manager.get_all_cookies(current_user_id)
        cookie_ids = list(user_cookies.keys())
        if not cookie_ids:
            now = get_local_now()
            return {
                'success': True,
                'data': {
                    'today_sales': 0.0,
                    'week_sales': 0.0,
                    'month_sales': 0.0,
                    'update_time': now.strftime('%Y-%m-%d %H:%M:%S')
                },
                'message': '获取销售额摘要成功'
            }
        
        # 计算时间范围
        now = get_local_now()
        
        # 当日开始
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_start_str = today_start.strftime('%Y-%m-%d')
        
        # 本周开始（周一）
        week_start = today_start - timedelta(days=today_start.weekday())
        week_start_str = week_start.strftime('%Y-%m-%d')
        
        # 本月开始
        month_start = today_start.replace(day=1)
        month_start_str = month_start.strftime('%Y-%m-%d')
        
        # 单次查询获取所有数据，减少数据库访问
        placeholders = ','.join(['?'] * len(cookie_ids))
        month_start_utc = local_date_to_utc_start(month_start_str)
        query = (
            f"SELECT amount, {ORDER_SALES_TIME_SQL} AS effective_sales_at, order_status "
            f"FROM orders WHERE {ORDER_SALES_TIME_SQL} >= ? AND cookie_id IN ({placeholders})"
        )
        all_orders = db_manager.execute_query(query, [month_start_utc] + cookie_ids)

        # 计算销售额
        today_sales = 0.0
        week_sales = 0.0
        month_sales = 0.0
        skipped_invalid_amount = 0
        skipped_ineligible_status = 0

        for order in all_orders:
            amount_str = order[0]
            effective_sales_at = order[1]
            order_status = order[2]

            if not is_sales_eligible_order_status(order_status):
                skipped_ineligible_status += 1
                continue

            amount = parse_order_amount_value(amount_str)
            if amount is None:
                skipped_invalid_amount += 1
                continue

            local_effective_sales_at = utc_timestamp_to_local_datetime(effective_sales_at)
            if not local_effective_sales_at:
                continue

            if local_effective_sales_at >= month_start:
                month_sales += amount

            if local_effective_sales_at >= week_start:
                week_sales += amount

            if local_effective_sales_at >= today_start:
                today_sales += amount

        logger.info(
            f"销售额摘要统计完成: skipped_invalid_amount={skipped_invalid_amount}, "
            f"skipped_ineligible_status={skipped_ineligible_status}"
        )
        
        today_sales = round(today_sales, 2)
        week_sales = round(week_sales, 2)
        month_sales = round(month_sales, 2)
        
        return {
            'success': True,
            'data': {
                'today_sales': today_sales,
                'week_sales': week_sales,
                'month_sales': month_sales,
                'update_time': now.strftime('%Y-%m-%d %H:%M:%S')
            },
            'message': '获取销售额摘要成功'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取销售额摘要失败: {e}")
        return {
            'success': False,
            'data': None,
            'message': f'获取销售额摘要失败: {str(e)}'
        }


# ========================= 防暴力破解管理API =========================

@app.get('/admin/security/login-stats')
async def get_login_security_stats(admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    """获取登录安全统计信息（仅管理员）"""
    current_time = time.time()
    
    # 统计IP封禁信息
    blocked_ips = []
    for ip, data in login_ip_tracker.items():
        if data.get('blocked_until', 0) > current_time:
            blocked_ips.append({
                'ip': ip,
                'attempts': data.get('attempts', 0),
                'blocked_until': data.get('blocked_until', 0),
                'remaining_seconds': int(data['blocked_until'] - current_time)
            })
    
    # 统计用户锁定信息
    locked_users = []
    for username, data in login_user_tracker.items():
        if data.get('locked_until', 0) > current_time:
            locked_users.append({
                'username': username,
                'attempts': data.get('attempts', 0),
                'locked_until': data.get('locked_until', 0),
                'remaining_seconds': int(data['locked_until'] - current_time)
            })
    
    # 最近失败的IP
    recent_failed_ips = []
    for ip, data in login_ip_tracker.items():
        if data.get('attempts', 0) > 0:
            recent_failed_ips.append({
                'ip': ip,
                'attempts': data.get('attempts', 0),
                'last_attempt': data.get('last_attempt', 0)
            })
    recent_failed_ips.sort(key=lambda x: x['last_attempt'], reverse=True)
    
    return {
        'success': True,
        'data': {
            'blocked_ips': blocked_ips,
            'blocked_ip_count': len(blocked_ips),
            'locked_users': locked_users,
            'locked_user_count': len(locked_users),
            'blacklisted_ips': list(ip_blacklist),
            'blacklist_count': len(ip_blacklist),
            'recent_failed_ips': recent_failed_ips[:20],  # 最近20个
            'config': BRUTE_FORCE_CONFIG
        }
    }


@app.post('/admin/security/unblock-ip/{ip}')
async def unblock_ip(ip: str, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    """解除IP封禁（仅管理员）"""
    unblocked = False
    
    # 从临时封禁中移除
    if ip in login_ip_tracker:
        login_ip_tracker[ip]['blocked_until'] = 0
        login_ip_tracker[ip]['attempts'] = 0
        unblocked = True
        logger.info(f"🔓 管理员 {admin_user['username']} 解除了IP {ip} 的临时封禁")
    
    # 从永久黑名单中移除
    if ip in ip_blacklist:
        ip_blacklist.discard(ip)
        unblocked = True
        logger.info(f"🔓 管理员 {admin_user['username']} 将IP {ip} 从永久黑名单中移除")
    
    if unblocked:
        return {'success': True, 'message': f'IP {ip} 已解除封禁'}
    else:
        return {'success': False, 'message': f'IP {ip} 未在封禁列表中'}


@app.post('/admin/security/unlock-user/{username}')
async def unlock_user(username: str, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    """解除用户锁定（仅管理员）"""
    if username in login_user_tracker:
        login_user_tracker[username]['locked_until'] = 0
        login_user_tracker[username]['attempts'] = 0
        logger.info(f"🔓 管理员 {admin_user['username']} 解除了用户 {username} 的锁定")
        return {'success': True, 'message': f'用户 {username} 已解除锁定'}
    else:
        return {'success': False, 'message': f'用户 {username} 未在锁定列表中'}


@app.post('/admin/security/blacklist-ip/{ip}')
async def add_ip_to_blacklist(ip: str, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    """将IP加入永久黑名单（仅管理员）"""
    ip_blacklist.add(ip)
    logger.warning(f"⛔ 管理员 {admin_user['username']} 将IP {ip} 加入永久黑名单")
    return {'success': True, 'message': f'IP {ip} 已加入永久黑名单'}


@app.post('/admin/security/update-config')
async def update_brute_force_config(
    config: Dict[str, Any],
    admin_user: Dict[str, Any] = Depends(verify_admin_token)
):
    """更新防暴力破解配置（仅管理员）"""
    valid_keys = set(BRUTE_FORCE_CONFIG.keys())
    updated = []
    
    for key, value in config.items():
        if key in valid_keys and isinstance(value, (int, float)):
            BRUTE_FORCE_CONFIG[key] = value
            updated.append(key)
    
    if updated:
        logger.info(f"⚙️ 管理员 {admin_user['username']} 更新了防暴力破解配置: {updated}")
        return {'success': True, 'message': f'已更新配置: {updated}', 'config': BRUTE_FORCE_CONFIG}
    else:
        return {'success': False, 'message': '没有有效的配置项被更新'}


# ========================= 防暴力破解管理API结束 =========================


# 修改管理员密码接口
@app.post('/change-admin-password')
async def change_admin_password(request: ChangePasswordRequest, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    from db_manager import db_manager

    try:
        # 验证当前密码（使用用户表验证）
        if not db_manager.verify_user_password('admin', request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 更新密码（使用用户表更新）
        success = db_manager.update_user_password('admin', request.new_password)

        if success:
            logger.info(f"【admin#{admin_user['user_id']}】管理员密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改管理员密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 生成图形验证码接口
@app.post('/generate-captcha')
async def generate_captcha(request: CaptchaRequest):
    from db_manager import db_manager

    try:
        # 生成图形验证码
        captcha_text, captcha_image = db_manager.generate_captcha()

        if not captcha_image:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码生成失败"
            )

        # 保存验证码到数据库
        if db_manager.save_captcha(request.session_id, captcha_text):
            return CaptchaResponse(
                success=True,
                captcha_image=captcha_image,
                session_id=request.session_id,
                message="图形验证码生成成功"
            )
        else:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码保存失败"
            )

    except Exception as e:
        logger.error(f"生成图形验证码失败: {e}")
        return CaptchaResponse(
            success=False,
            captcha_image="",
            session_id=request.session_id,
            message="图形验证码生成失败"
        )


# 验证图形验证码接口
@app.post('/verify-captcha')
async def verify_captcha(request: VerifyCaptchaRequest):
    from db_manager import db_manager

    try:
        if db_manager.verify_captcha(request.session_id, request.captcha_code):
            return VerifyCaptchaResponse(
                success=True,
                message="图形验证码验证成功"
            )
        else:
            return VerifyCaptchaResponse(
                success=False,
                message="图形验证码错误或已过期"
            )

    except Exception as e:
        logger.error(f"验证图形验证码失败: {e}")
        return VerifyCaptchaResponse(
            success=False,
            message="图形验证码验证失败"
        )


# 发送验证码接口（需要先验证图形验证码）
@app.post('/send-verification-code')
async def send_verification_code(request: SendCodeRequest):
    from db_manager import db_manager

    try:
        # 检查是否已验证图形验证码
        # 通过检查数据库中是否存在已验证的图形验证码记录
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            current_time = time.time()

            # 查找最近5分钟内该session_id的验证记录
            # 由于验证成功后验证码会被删除，我们需要另一种方式来跟踪验证状态
            # 这里我们检查该session_id是否在最近验证过（通过检查是否有已删除的记录）

            # 为了简化，我们要求前端在验证图形验证码成功后立即发送邮件验证码
            # 或者我们可以在验证成功后设置一个临时标记
            pass

        # 根据验证码类型进行不同的检查
        if request.type == 'register':
            # 注册验证码：检查邮箱是否已注册
            existing_user = db_manager.get_user_by_email(request.email)
            if existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱已被注册"
                )
        elif request.type == 'login':
            # 登录验证码：检查邮箱是否存在
            existing_user = db_manager.get_user_by_email(request.email)
            if not existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱未注册"
                )

        # 生成验证码
        code = db_manager.generate_verification_code()

        # 保存验证码到数据库
        if not db_manager.save_verification_code(request.email, code, request.type):
            return SendCodeResponse(
                success=False,
                message="验证码保存失败，请稍后重试"
            )

        # 发送验证码邮件
        if await db_manager.send_verification_email(request.email, code):
            return SendCodeResponse(
                success=True,
                message="验证码已发送到您的邮箱，请查收"
            )
        else:
            return SendCodeResponse(
                success=False,
                message="验证码发送失败，请检查邮箱地址或稍后重试"
            )

    except Exception as e:
        logger.error(f"发送验证码失败: {e}")
        return SendCodeResponse(
            success=False,
            message="发送验证码失败，请稍后重试"
        )


# 用户注册接口
@app.post('/register')
async def register(request: RegisterRequest):
    from db_manager import db_manager

    # 检查注册是否开启
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        logger.warning(f"【{request.username}】注册失败: 注册功能已关闭")
        return RegisterResponse(
            success=False,
            message="注册功能已关闭，请联系管理员"
        )

    try:
        logger.info(f"【{request.username}】尝试注册，邮箱: {request.email}")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(request.email, request.verification_code):
            logger.warning(f"【{request.username}】注册失败: 验证码错误或已过期")
            return RegisterResponse(
                success=False,
                message="验证码错误或已过期"
            )

        # 检查用户名是否已存在
        existing_user = db_manager.get_user_by_username(request.username)
        if existing_user:
            logger.warning(f"【{request.username}】注册失败: 用户名已存在")
            return RegisterResponse(
                success=False,
                message="用户名已存在"
            )

        # 检查邮箱是否已注册
        existing_email = db_manager.get_user_by_email(request.email)
        if existing_email:
            logger.warning(f"【{request.username}】注册失败: 邮箱已被注册")
            return RegisterResponse(
                success=False,
                message="该邮箱已被注册"
            )

        # 创建用户
        if db_manager.create_user(request.username, request.email, request.password):
            logger.info(f"【{request.username}】注册成功")
            return RegisterResponse(
                success=True,
                message="注册成功，请登录"
            )
        else:
            logger.error(f"【{request.username}】注册失败: 数据库操作失败")
            return RegisterResponse(
                success=False,
                message="注册失败，请稍后重试"
            )

    except Exception as e:
        logger.error(f"【{request.username}】注册异常: {e}")
        return RegisterResponse(
            success=False,
            message="注册失败，请稍后重试"
        )


# ------------------------- 发送消息接口 -------------------------

# 固定的API秘钥（生产环境中应该从配置文件或环境变量读取）
# 注意：现在从系统设置中读取QQ回复消息秘钥
API_SECRET_KEY = "xianyu_api_secret_2024"  # 保留作为后备

class SendMessageRequest(BaseModel):
    api_key: str
    cookie_id: str
    chat_id: str
    to_user_id: str
    message: str


class SendMessageResponse(BaseModel):
    success: bool
    message: str


def verify_api_key(api_key: str) -> bool:
    """验证API秘钥"""
    try:
        # 从系统设置中获取QQ回复消息秘钥
        from db_manager import db_manager
        qq_secret_key = db_manager.get_system_setting('qq_reply_secret_key')

        # 如果系统设置中没有配置，使用默认值
        if not qq_secret_key:
            qq_secret_key = API_SECRET_KEY

        return api_key == qq_secret_key
    except Exception as e:
        logger.error(f"验证API秘钥时发生异常: {e}")
        # 异常情况下使用默认秘钥验证
        return api_key == API_SECRET_KEY


@app.post('/send-message', response_model=SendMessageResponse)
async def send_message_api(request: SendMessageRequest):
    """发送消息API接口（使用秘钥验证）"""
    try:
        # 清理所有参数中的换行符
        def clean_param(param_str):
            """清理参数中的换行符"""
            if isinstance(param_str, str):
                return param_str.replace('\\n', '').replace('\n', '')
            return param_str

        # 清理所有参数
        cleaned_api_key = clean_param(request.api_key)
        cleaned_cookie_id = clean_param(request.cookie_id)
        cleaned_chat_id = clean_param(request.chat_id)
        cleaned_to_user_id = clean_param(request.to_user_id)
        cleaned_message = clean_param(request.message)

        # 验证API秘钥不能为空
        if not cleaned_api_key:
            logger.warning("API秘钥为空")
            return SendMessageResponse(
                success=False,
                message="API秘钥不能为空"
            )

        # 特殊测试秘钥处理
        if cleaned_api_key == "zhinina_test_key":
            logger.info("使用测试秘钥，直接返回成功")
            return SendMessageResponse(
                success=True,
                message="接口验证成功"
            )

        # 验证API秘钥
        if not verify_api_key(cleaned_api_key):
            logger.warning(f"API秘钥验证失败: {mask_sensitive_text(cleaned_api_key)}")
            return SendMessageResponse(
                success=False,
                message="API秘钥验证失败"
            )

        # 验证必需参数不能为空
        required_params = {
            'cookie_id': cleaned_cookie_id,
            'chat_id': cleaned_chat_id,
            'to_user_id': cleaned_to_user_id,
            'message': cleaned_message
        }

        for param_name, param_value in required_params.items():
            if not param_value:
                logger.warning(f"必需参数 {param_name} 为空")
                return SendMessageResponse(
                    success=False,
                    message=f"参数 {param_name} 不能为空"
                )

        # 直接获取XianyuLive实例，跳过cookie_manager检查
        from XianyuAutoAsync import XianyuLive, ConnectionState
        live_instance = XianyuLive.get_instance(cleaned_cookie_id)

        if not live_instance:
            logger.warning(f"账号实例不存在或未连接: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="账号实例不存在或未连接，请检查账号状态"
            )

        # 检查WebSocket连接状态（使用connection_state作为主要判断依据）
        # connection_state 是项目维护的连接状态，比 ws.closed 更可靠
        if live_instance.connection_state != ConnectionState.CONNECTED:
            logger.warning(f"账号WebSocket连接状态异常: {cleaned_cookie_id}, 状态: {live_instance.connection_state}")
            return SendMessageResponse(
                success=False,
                message=f"账号WebSocket连接状态异常({live_instance.connection_state.value})，请等待重连"
            )
        
        # 额外检查ws对象是否存在
        if not live_instance.ws:
            logger.warning(f"账号WebSocket对象不存在: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="账号WebSocket连接未就绪，请等待重连"
            )

        # 发送消息时需要回到账号实例所属事件循环，避免跨 loop 直接操作 ws
        await _run_live_instance_on_manager_loop(
            cleaned_cookie_id,
            lambda: live_instance.send_msg(
                live_instance.ws,
                cleaned_chat_id,
                cleaned_to_user_id,
                cleaned_message
            ),
            timeout=15,
        )

        logger.info(f"API成功发送消息: {cleaned_cookie_id} -> {cleaned_to_user_id}, 内容: {cleaned_message[:50]}{'...' if len(cleaned_message) > 50 else ''}")

        return SendMessageResponse(
            success=True,
            message="消息发送成功"
        )

    except HTTPException as e:
        # 使用清理后的参数记录日志
        cookie_id_for_log = clean_param(request.cookie_id) if 'clean_param' in locals() else request.cookie_id
        to_user_id_for_log = clean_param(request.to_user_id) if 'clean_param' in locals() else request.to_user_id
        logger.warning(f"API发送消息被拒绝: {cookie_id_for_log} -> {to_user_id_for_log}, 原因: {mask_sensitive_text(e.detail)}")
        return SendMessageResponse(
            success=False,
            message=str(e.detail or "发送消息失败，请稍后重试")
        )
    except Exception as e:
        # 使用清理后的参数记录日志
        cookie_id_for_log = clean_param(request.cookie_id) if 'clean_param' in locals() else request.cookie_id
        to_user_id_for_log = clean_param(request.to_user_id) if 'clean_param' in locals() else request.to_user_id
        logger.error(f"API发送消息异常: {cookie_id_for_log} -> {to_user_id_for_log}, 错误: {mask_sensitive_text(e)}")
        return SendMessageResponse(
            success=False,
            message="发送消息失败，请稍后重试"
        )


@app.post("/xianyu/reply", response_model=ResponseModel)
async def xianyu_reply(req: RequestModel):
    msg_template = match_reply(req.cookie_id, req.send_message)
    is_default_reply = False

    if not msg_template:
        # 从数据库获取默认回复
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)

        if default_reply_settings and default_reply_settings.get('enabled', False):
            # 检查是否开启了"只回复一次"功能
            if default_reply_settings.get('reply_once', False):
                # 检查是否已经回复过这个chat_id
                if db_manager.has_default_reply_record(req.cookie_id, req.chat_id):
                    raise HTTPException(status_code=404, detail="该对话已使用默认回复，不再重复回复")

            msg_template = default_reply_settings.get('reply_content', '')
            is_default_reply = True

        # 如果数据库中没有设置或为空，返回错误
        if not msg_template:
            raise HTTPException(status_code=404, detail="未找到匹配的回复规则且未设置默认回复")

    # 按占位符格式化
    try:
        send_msg = msg_template.format(
            send_user_id=req.send_user_id,
            send_user_name=req.send_user_name,
            send_message=req.send_message,
        )
    except Exception:
        # 如果格式化失败，返回原始内容
        send_msg = msg_template

    # 如果是默认回复且开启了"只回复一次"，记录回复记录
    if is_default_reply:
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)
        if default_reply_settings and default_reply_settings.get('reply_once', False):
            db_manager.add_default_reply_record(req.cookie_id, req.chat_id)

    return {"code": 200, "data": {"send_msg": send_msg}}

# ------------------------- 账号 / 关键字管理接口 -------------------------


class CookieIn(BaseModel):
    id: str
    value: str


class ManualCookieImportRequest(BaseModel):
    account_id: str
    cookie: str
    show_browser: bool = False


class CookieStatusIn(BaseModel):
    enabled: bool


class DefaultReplyIn(BaseModel):
    enabled: bool
    reply_content: Optional[str] = None
    reply_once: bool = False


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    value: str
    description: Optional[str] = None


NIGHT_MODE_SYSTEM_SETTING_KEYS = {
    'risk_control_night_mode_enabled',
    'risk_control_night_start_hour',
    'risk_control_night_end_hour',
}


def _validate_system_setting_value(key: str, value: str) -> str:
    if key == 'risk_control_night_mode_enabled':
        normalized = str(value).strip().lower()
        if normalized in {'true', '1', 'yes', 'on'}:
            return 'true'
        if normalized in {'false', '0', 'no', 'off'}:
            return 'false'
        raise HTTPException(status_code=400, detail='夜间降频开关只能为 true 或 false')

    if key in {'risk_control_night_start_hour', 'risk_control_night_end_hour'}:
        try:
            hour = int(str(value).strip())
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail='夜间时间必须是 0-23 的整数')
        if hour < 0 or hour > 23:
            raise HTTPException(status_code=400, detail='夜间时间必须是 0-23 的整数')
        return str(hour)

    return value


class SystemSettingCreateIn(BaseModel):
    key: str
    value: str
    description: Optional[str] = None


class ChatSendRequest(BaseModel):
    cookie_id: str
    chat_id: str
    to_user_id: str
    message: str


class SaveItemKeywordsRequest(BaseModel):
    keywords: list
    item_reply: Optional[str] = None


class CopyKeywordsRequest(BaseModel):
    source_item_id: str
    target_item_ids: List[str]


class ChatHydrationDebug(BaseModel):
    success: bool
    cookie_id: str
    chat_id: str
    stage: str
    message: str
    fetched: int = 0
    saved: int = 0
    normalized_count: int = 0
    skipped_count: int = 0
    sample_sender_id: Optional[str] = None
    sample_sender_name: Optional[str] = None
    sample_content: Optional[str] = None
    remote_history_status: Optional[str] = None
    remote_history_checked_at: Optional[str] = None
    runtime_status: Optional[Dict[str, Any]] = None


_chat_session_enrichment_cache: Dict[str, Dict[str, Any]] = {}
_CHAT_SESSION_ENRICHMENT_TTL_SECONDS = 180
_chat_history_probe_cache: Dict[str, Dict[str, Any]] = {}
_CHAT_HISTORY_PROBE_TTL_SECONDS = 6 * 60 * 60


def _build_chat_history_probe_key(cookie_id: str, chat_id: str) -> str:
    return f"{str(cookie_id or '').strip()}::{str(chat_id or '').strip()}"


def _get_cached_chat_history_probe(cookie_id: str, chat_id: str) -> Optional[Dict[str, Any]]:
    cache_key = _build_chat_history_probe_key(cookie_id, chat_id)
    cached = _chat_history_probe_cache.get(cache_key)
    if not cached:
        return None

    checked_at = float(cached.get('checked_at') or 0)
    if checked_at <= 0 or (time.time() - checked_at) > _CHAT_HISTORY_PROBE_TTL_SECONDS:
        _chat_history_probe_cache.pop(cache_key, None)
        return None

    return dict(cached)


def _set_cached_chat_history_probe(
    cookie_id: str,
    chat_id: str,
    *,
    status: str,
    fetched: int = 0,
    normalized_count: int = 0,
    saved: int = 0,
    note: Optional[str] = None,
) -> Dict[str, Any]:
    checked_at = time.time()
    payload = {
        'status': str(status or '').strip() or 'unknown',
        'fetched': max(0, int(fetched or 0)),
        'normalized_count': max(0, int(normalized_count or 0)),
        'saved': max(0, int(saved or 0)),
        'note': str(note or '').strip() or None,
        'checked_at': checked_at,
        'checked_at_display': datetime.fromtimestamp(checked_at, tz=LOCAL_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S'),
    }
    _chat_history_probe_cache[_build_chat_history_probe_key(cookie_id, chat_id)] = payload
    return dict(payload)


def _apply_chat_history_probe_to_session(cookie_id: str, session: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(session or {})
    chat_id = str(normalized.get('chat_id') or '').strip()
    if not chat_id:
        return normalized

    probe = _get_cached_chat_history_probe(cookie_id, chat_id)
    if probe:
        normalized['remote_history_status'] = probe.get('status')
        normalized['remote_history_checked_at'] = probe.get('checked_at_display')
        normalized['remote_history_note'] = probe.get('note')
        normalized['remote_history_fetched'] = probe.get('fetched', 0)

    return normalized


def _compact_chat_user_ext(user_ext: Any) -> Dict[str, Any]:
    if not isinstance(user_ext, dict):
        return {}
    allowed_keys = {
        'yuxiaopuDomain', 'yuxiaopuLevelImage', 'fansTag', 'userMedal',
        'avatarPendant', 'chatBackground', 'guestChatBubble', 'ownerChatBubble'
    }
    return {key: value for key, value in user_ext.items() if key in allowed_keys and value}


def _parse_item_pre_info(raw_value: Any) -> Dict[str, Any]:
    parsed = _safe_json_loads(raw_value)
    if parsed:
        return parsed
    if not isinstance(raw_value, str) or not raw_value.strip():
        return {}
    try:
        return json.loads(raw_value.replace('\\"', '"'))
    except Exception:
        return {}


def _normalize_headinfo_buttons(buttons: Any) -> List[Dict[str, Any]]:
    normalized = []
    if not isinstance(buttons, list):
        return normalized
    for button in buttons:
        if not isinstance(button, dict):
            continue
        normalized.append({
            'name': button.get('name'),
            'style': button.get('style'),
            'trade_action': button.get('tradeAction'),
            'url': (((button.get('clickEvent') or {}).get('data') or {}).get('url')),
        })
    return normalized


def _build_chat_session_cache_key(cookie_id: str, session: Dict[str, Any]) -> str:
    return f"{cookie_id}:{session.get('chat_id') or ''}:{session.get('item_id') or ''}:{session.get('sender_id') or ''}"


def _get_cached_chat_session_enrichment(cache_key: str) -> Optional[Dict[str, Any]]:
    cached = _chat_session_enrichment_cache.get(cache_key)
    if not cached:
        return None
    if (time.time() - float(cached.get('cached_at') or 0)) > _CHAT_SESSION_ENRICHMENT_TTL_SECONDS:
        _chat_session_enrichment_cache.pop(cache_key, None)
        return None
    return dict(cached.get('value') or {})


def _set_cached_chat_session_enrichment(cache_key: str, value: Dict[str, Any]) -> None:
    _chat_session_enrichment_cache[cache_key] = {
        'cached_at': time.time(),
        'value': dict(value or {}),
    }


async def _enrich_single_chat_session(cookie_id: str, session: Dict[str, Any]) -> Dict[str, Any]:
    from XianyuAutoAsync import XianyuLive

    cache_key = _build_chat_session_cache_key(cookie_id, session)
    cached = _get_cached_chat_session_enrichment(cache_key)
    if cached is not None:
        return {**session, **cached}

    live_instance = XianyuLive.get_instance(cookie_id)
    if not live_instance:
        return session

    session_id = str(session.get('chat_id') or '').strip()
    if not session_id:
        return session

    item_id = str(session.get('item_id') or '').strip()
    sender_id = str(session.get('sender_id') or session.get('buyer_id') or '').strip()
    session_type = int(session.get('session_type') or 1)

    enriched: Dict[str, Any] = {}

    try:
        user_info_result = await live_instance.fetch_im_user_info(
            session_id=session_id,
            session_type=session_type,
            is_owner=False,
            message_id=session.get('message_id') or None,
        )
        user_info = user_info_result.get('userInfo', {}) if isinstance(user_info_result, dict) else {}
        if user_info:
            enriched.update({
                'avatar': user_info.get('logo'),
                'fish_nick': user_info.get('fishNick') or user_info.get('nick') or session.get('buyer_name') or session.get('sender_name'),
                'user_ext': _compact_chat_user_ext(user_info.get('ext')),
                'buyer_name_resolved': user_info.get('fishNick') or user_info.get('nick') or session.get('buyer_name'),
                'sender_id': sender_id or session.get('sender_id'),
            })
    except Exception as e:
        logger.debug(f"会话用户信息增强失败: cookie_id={cookie_id}, session_id={session_id}, error={mask_sensitive_text(e)}")

    if item_id:
        try:
            headinfo = await live_instance.fetch_im_head_info(session_id=session_id, item_id=item_id, session_type=session_type)
            common_data = headinfo.get('commonData', {}) if isinstance(headinfo, dict) else {}
            item_pre_info = _parse_item_pre_info(common_data.get('itemPreInfo'))
            left_data = ((headinfo.get('left') or {}).get('data') or {}) if isinstance(headinfo, dict) else {}
            middle_data = ((headinfo.get('middle') or {}).get('data') or {}) if isinstance(headinfo, dict) else {}
            right_data = ((headinfo.get('right') or {}).get('data') or {}) if isinstance(headinfo, dict) else {}
            ut_args = headinfo.get('utArgs', {}) if isinstance(headinfo, dict) else {}
            enriched.update({
                'headinfo_template': headinfo.get('template') if isinstance(headinfo, dict) else None,
                'item_title': item_pre_info.get('title') or session.get('item_title'),
                'item_price': item_pre_info.get('soldPrice') or middle_data.get('price'),
                'item_pic': left_data.get('picUrl'),
                'item_jump_url': left_data.get('jumpUrl'),
                'item_subtitle': middle_data.get('subTitle'),
                'item_tips': middle_data.get('tips'),
                'action_buttons': _normalize_headinfo_buttons(right_data.get('btnList')),
                'order_id': headinfo.get('orderId') if isinstance(headinfo, dict) else None,
                'order_detail_url': headinfo.get('orderDetailUrl') if isinstance(headinfo, dict) else None,
                'order_status_name': (ut_args.get('orderStatusName') if isinstance(ut_args, dict) else None),
            })
        except Exception as e:
            logger.debug(f"会话头信息增强失败: cookie_id={cookie_id}, session_id={session_id}, item_id={item_id}, error={mask_sensitive_text(e)}")

    try:
        blacklist_info = await live_instance.fetch_im_blacklist_status(session_id=session_id)
        if blacklist_info:
            enriched['blacklist_status'] = {
                'is_in_black': bool(blacklist_info.get('isInBlack')),
                'show_blacklist': bool(blacklist_info.get('showBlackList')),
            }
    except Exception as e:
        logger.debug(f"会话黑名单增强失败: cookie_id={cookie_id}, session_id={session_id}, error={mask_sensitive_text(e)}")

    _set_cached_chat_session_enrichment(cache_key, enriched)
    return {**session, **enriched}


async def _enrich_chat_sessions(cookie_id: str, sessions: List[Dict[str, Any]], limit: int = 30) -> List[Dict[str, Any]]:
    if not sessions:
        return []
    sessions = list(sessions)
    priority_sessions = sessions[:max(1, min(limit, len(sessions)))]
    remaining_sessions = sessions[len(priority_sessions):]
    enriched_priority = []
    for session in priority_sessions:
        enriched_priority.append(await _enrich_single_chat_session(cookie_id, session))
    return enriched_priority + remaining_sessions


def _safe_json_loads(raw_value: Any) -> Dict[str, Any]:
    if isinstance(raw_value, dict):
        return raw_value
    if not isinstance(raw_value, str) or not raw_value.strip():
        return {}
    try:
        parsed = json.loads(raw_value)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _extract_history_message_payload(message: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(message, dict):
        return {}
    try:
        message_1 = message.get('1', {}) if isinstance(message, dict) else {}
        message_6 = message_1.get('6', {}) if isinstance(message_1, dict) else {}
        message_6_3 = message_6.get('3', {}) if isinstance(message_6, dict) else {}
        return _safe_json_loads(message_6_3.get('5', '') or '{}')
    except Exception:
        return {}


def _extract_rich_message_fields(message: Dict[str, Any]) -> Dict[str, Any]:
    payload = _extract_history_message_payload(message)
    result = {
        'display_type': None,
        'content': '',
        'image_url': None,
        'media_url': None,
        'link_url': None,
        'extra_json': None,
    }

    if not payload:
        return result

    dx_card = payload.get('dxCard', {}) if isinstance(payload, dict) else {}
    dx_item = dx_card.get('item', {}) if isinstance(dx_card, dict) else {}
    main = dx_item.get('main', {}) if isinstance(dx_item, dict) else {}
    ex_content = main.get('exContent', {}) if isinstance(main, dict) else {}
    title = str(ex_content.get('title') or main.get('title') or payload.get('title') or '').strip()
    content = str(ex_content.get('content') or payload.get('text') or '').strip()
    button_text = str((ex_content.get('button') or {}).get('text') or '').strip()

    image_url = (
        ((payload.get('image') or {}).get('pics') or [{}])[0].get('url')
        if isinstance(payload.get('image'), dict) and (payload.get('image').get('pics') or [])
        else None
    )
    video_url = (
        ((payload.get('video') or {}).get('playUrl'))
        or ((payload.get('video') or {}).get('url'))
        or ((main.get('video') or {}).get('playUrl') if isinstance(main.get('video'), dict) else None)
    )
    link_url = (
        str(payload.get('targetUrl') or '').strip()
        or str(payload.get('url') or '').strip()
        or str((ex_content.get('button') or {}).get('actionUrl') or '').strip()
    ) or None

    item_id = None
    item_title = None
    item_image = None
    if isinstance(dx_item, dict):
        item_id = dx_item.get('itemId') or dx_item.get('id')
        item_title = dx_item.get('title') or title
        item_image = dx_item.get('itemMainPic') or dx_item.get('pic')

    extra = {
        'payload': payload,
        'title': title or None,
        'button_text': button_text or None,
        'item_share': {
            'item_id': item_id,
            'title': item_title,
            'image_url': item_image,
            'seller_id': dx_item.get('itemSellerId') if isinstance(dx_item, dict) else None,
        } if item_id or item_title or item_image else None,
    }

    if video_url:
        result['display_type'] = 'video'
        result['content'] = title or content or '[视频]'
        result['media_url'] = str(video_url).strip()
        result['image_url'] = image_url or item_image
        result['link_url'] = link_url
    elif image_url:
        result['display_type'] = 'image'
        result['content'] = title or content or '[图片]'
        result['image_url'] = str(image_url).strip()
        result['link_url'] = link_url
    elif item_id or item_title:
        result['display_type'] = 'item_share'
        result['content'] = item_title or title or content or '[商品分享]'
        result['image_url'] = item_image
        result['link_url'] = link_url
    elif link_url:
        result['display_type'] = 'link'
        result['content'] = title or content or button_text or '[链接]'
        result['link_url'] = link_url
    elif title or content or button_text:
        result['display_type'] = 'card'
        result['content'] = ' / '.join([part for part in [title, content, button_text] if part])

    if result['display_type']:
        result['extra_json'] = json.dumps(extra, ensure_ascii=False)

    return result


def _extract_history_message_text(message: Dict[str, Any]) -> str:
    """从闲鱼历史消息结构中尽量提取可展示文本。"""
    if not isinstance(message, dict):
        return ''

    try:
        message_1 = message.get('1', {}) if isinstance(message, dict) else {}
        message_10 = message_1.get('10', {}) if isinstance(message_1, dict) else {}
        payload = _extract_history_message_payload(message)
        candidates = [
            message_10.get('reminderContent'),
            message_10.get('detailNotice'),
            message_10.get('reminderTitle'),
            message_10.get('reminderNotice'),
            (((payload.get('dxCard') or {}).get('item') or {}).get('main') or {}).get('title'),
            ((((payload.get('dxCard') or {}).get('item') or {}).get('main') or {}).get('exContent') or {}).get('title'),
            ((((payload.get('dxCard') or {}).get('item') or {}).get('main') or {}).get('exContent') or {}).get('content'),
            ((((payload.get('dxCard') or {}).get('item') or {}).get('main') or {}).get('exContent') or {}).get('button', {}).get('text'),
            (payload.get('text') if isinstance(payload, dict) else None),
        ]
        for candidate in candidates:
            text = str(candidate or '').strip()
            if text and text not in {'{}', '[]'}:
                return text
    except Exception:
        pass

    raw_text = str(message.get('raw') or '').strip()
    return raw_text[:120] if raw_text else ''


def _normalize_chat_history_message_record(
    raw: Dict[str, Any],
    cookie_id: str,
    chat_id: str,
    owner_user_id: Optional[str] = None,
    fallback_item_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """将闲鱼历史消息结构转换为本地 chat_messages 记录格式。"""
    if not isinstance(raw, dict):
        logger.debug(f"聊天历史记录格式异常，raw 不是 dict: cookie_id={cookie_id}, chat_id={chat_id}, type={type(raw).__name__}")
        return None

    message = raw.get('message')
    if not isinstance(message, dict):
        logger.debug(
            f"聊天历史记录缺少 message 结构: cookie_id={cookie_id}, chat_id={chat_id}, "
            f"keys={list(raw.keys())[:8]}"
        )
        return None

    sender_id = str(raw.get('send_user_id') or '').strip()
    message_extension = raw.get('message_extension') if isinstance(raw.get('message_extension'), dict) else {}
    sender_name = (
        str(raw.get('send_user_name') or '').strip()
        or str(message_extension.get('senderNick') or '').strip()
        or str(message_extension.get('reminderTitle') or '').strip()
        or sender_id
        or chat_id
    )
    content = ''
    content_type = 1
    image_url = None
    media_url = None
    link_url = None
    extra_json = None
    item_id = None
    created_at = None

    try:
        message_1 = message.get('1', {}) if isinstance(message, dict) else {}
        message_10 = message_1.get('10', {}) if isinstance(message_1, dict) else {}
        created_ts = raw.get('created_at', message_1.get('5'))
        created_at = _format_history_created_at(created_ts)
        content = _extract_history_message_text(message)
        message_6 = message_1.get('6', {}) if isinstance(message_1, dict) else {}
        message_6_3 = message_6.get('3', {}) if isinstance(message_6, dict) else {}
        content_type = int(message_6_3.get('4', 1) or 1)
        rich_fields = _extract_rich_message_fields(message)
        if rich_fields.get('display_type') == 'image':
            content_type = 2
        elif rich_fields.get('display_type') == 'video':
            content_type = 3
        elif rich_fields.get('display_type') == 'link':
            content_type = 4
        elif rich_fields.get('display_type') == 'item_share':
            content_type = 5
        elif rich_fields.get('display_type') == 'card':
            content_type = 6
        if rich_fields.get('content'):
            content = rich_fields.get('content')
        image_url = rich_fields.get('image_url') or image_url
        media_url = rich_fields.get('media_url')
        link_url = rich_fields.get('link_url')
        extra_json = rich_fields.get('extra_json')
        if content_type == 2:
            content_json_str = message_6_3.get('5', '')
            if content_json_str:
                content_obj = json.loads(content_json_str)
                pics = content_obj.get('image', {}).get('pics', [])
                if pics:
                    image_url = pics[0].get('url', '') or image_url
            if not content:
                content = '[图片]'
        elif content_type == 26 and not content:
            card_title = (
                (((_extract_history_message_payload(message).get('dxCard') or {}).get('item') or {}).get('main') or {}).get('exContent', {})
            )
            content = str(card_title.get('title') or message_10.get('detailNotice') or message_10.get('reminderContent') or '[交易卡片]').strip()
        reminder_url = str(message_10.get('reminderUrl') or '').strip()
        if reminder_url:
            parsed = urlparse(reminder_url)
            item_id = parse_qs(parsed.query or '').get('itemId', [None])[0]
            if not link_url:
                link_url = reminder_url
    except Exception as normalize_exc:
        logger.warning(
            f"聊天历史记录解析失败: cookie_id={cookie_id}, chat_id={chat_id}, "
            f"sender_id={sender_id or '-'}, error={mask_sensitive_text(normalize_exc)}"
        )

    if not content:
        fallback_content = (
            str(message_extension.get('detailNotice') or '').strip()
            or str(message_extension.get('reminderContent') or '').strip()
            or str(message_extension.get('reminderNotice') or '').strip()
        )
        if fallback_content:
            content = fallback_content

    if not item_id and fallback_item_id:
        item_id = fallback_item_id

    owner_id = str(owner_user_id or '').strip()
    direction = 1 if sender_id and owner_id and sender_id == owner_id else 2

    return {
        'cookie_id': cookie_id,
        'chat_id': chat_id,
        'sender_id': sender_id,
        'sender_name': sender_name,
        'content': content or ('[图片]' if content_type == 2 else '[系统消息]'),
        'content_type': content_type,
        'image_url': image_url,
        'item_id': item_id,
        'direction': direction,
        'reply_source': None,
        'media_url': media_url,
        'link_url': link_url,
        'extra_json': extra_json,
        'created_at': created_at,
    }


def _format_history_created_at(raw_value: Any) -> Optional[str]:
    if raw_value in (None, '', 0, '0'):
        return None

    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return None
        if any(sep in text for sep in ('-', '/')) and ':' in text:
            normalized = text.replace('T', ' ')
            return normalized[:19]
        raw_value = text

    try:
        value = int(float(raw_value))
    except (TypeError, ValueError):
        return None

    if value <= 0:
        return None

    if value < 10**11:
        value *= 1000

    try:
        return datetime.fromtimestamp(value / 1000, tz=LOCAL_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')
    except (OverflowError, OSError, ValueError):
        return None


def _build_chat_message_signature(record: Dict[str, Any]) -> tuple:
    return (
        str(record.get('chat_id') or ''),
        str(record.get('sender_id') or ''),
        str(record.get('content') or ''),
        int(record.get('content_type') or 0),
        str(record.get('image_url') or ''),
        str(record.get('media_url') or ''),
        str(record.get('link_url') or ''),
        str(record.get('item_id') or ''),
        int(record.get('direction') or 0),
        str(record.get('created_at') or ''),
    )


def _build_chat_sessions_from_recent_orders(cookie_id: str, limit: int = 50) -> List[Dict[str, Any]]:
    """当本地 chat_messages 为空时，基于最近订单构造可点击会话入口。"""
    sessions: List[Dict[str, Any]] = []
    seen_chat_ids = set()
    orders = db_manager.get_orders_by_cookie(cookie_id, limit=max(limit * 4, 100))

    for order in orders:
        sid = str(order.get('sid') or '').strip()
        if not sid:
            continue
        chat_id = sid.split('@')[0]
        if not chat_id or chat_id in seen_chat_ids:
            continue
        seen_chat_ids.add(chat_id)
        sessions.append({
            'chat_id': chat_id,
            'sender_id': order.get('buyer_id') or '',
            'buyer_id': order.get('buyer_id') or '',
            'sender_name': order.get('buyer_nick') or order.get('buyer_id') or chat_id,
            'buyer_name': order.get('buyer_nick') or '',
            'content': '',
            'content_type': 1,
            'item_id': order.get('item_id') or '',
            'direction': 2,
            'created_at': order.get('updated_at') or order.get('created_at') or '',
        })
        if len(sessions) >= limit:
            break

    sessions.sort(key=lambda item: item.get('created_at') or '', reverse=True)
    return sessions


def _merge_chat_sessions_with_order_fallback(
    local_sessions: List[Dict[str, Any]],
    fallback_sessions: List[Dict[str, Any]],
    limit: int = 100,
) -> List[Dict[str, Any]]:
    """合并本地会话和订单兜底会话，避免本地只有少量会话时隐藏其他历史入口。"""
    merged: List[Dict[str, Any]] = []
    seen_chat_ids = set()

    for session in local_sessions or []:
        chat_id = str(session.get('chat_id') or '').strip()
        if not chat_id or chat_id in seen_chat_ids:
            continue
        merged.append(session)
        seen_chat_ids.add(chat_id)

    for session in fallback_sessions or []:
        chat_id = str(session.get('chat_id') or '').strip()
        if not chat_id or chat_id in seen_chat_ids:
            continue
        merged.append(session)
        seen_chat_ids.add(chat_id)

    merged.sort(key=lambda item: str(item.get('created_at') or ''), reverse=True)
    return merged[:limit]


def _annotate_chat_sessions(cookie_id: str, sessions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    annotated = []
    for session in sessions or []:
        annotated.append(_apply_chat_history_probe_to_session(cookie_id, session))
    return annotated


def _get_user_cookies_map(current_user: Dict[str, Any]) -> Dict[str, str]:
    user_id = current_user['user_id']
    return db_manager.get_all_cookies(user_id)


def _ensure_cookie_access(cid: str, current_user: Dict[str, Any]) -> str:
    cleaned_cid = str(cid or '').strip()
    if not cleaned_cid:
        raise HTTPException(status_code=400, detail="缺少Cookie ID")

    user_cookies = _get_user_cookies_map(current_user)
    if cleaned_cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")
    return cleaned_cid


def _normalize_runtime_timestamp(value: Any) -> Optional[float]:
    try:
        timestamp = float(value)
    except (TypeError, ValueError):
        return None
    return timestamp if timestamp > 0 else None


def _format_runtime_timestamp(value: Any) -> Optional[str]:
    timestamp = _normalize_runtime_timestamp(value)
    if timestamp is None:
        return None

    return datetime.fromtimestamp(timestamp, tz=LOCAL_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')


def _get_runtime_age_seconds(value: Any) -> Optional[int]:
    timestamp = _normalize_runtime_timestamp(value)
    if timestamp is None:
        return None
    return max(0, int(time.time() - timestamp))


def _is_runtime_timestamp_recent(value: Any, window_seconds: Any) -> bool:
    timestamp = _normalize_runtime_timestamp(value)
    if timestamp is None:
        return False

    try:
        window = max(1, int(float(window_seconds)))
    except (TypeError, ValueError):
        return False

    return (time.time() - timestamp) <= window


def _build_live_runtime_status(cookie_id: str) -> Dict[str, Any]:
    cleaned_cid = str(cookie_id or '').strip()
    runtime_status = {
        'instance_exists': False,
        'running': False,
        'connection_state': 'not_running',
        'ws_ready': False,
        'session_ready': False,
        'has_current_token': False,
        'message_stream_ready': False,
        'message_stream_status': 'not_running',
        'message_stream_note': None,
        'token_refresh_status': None,
        'token_refresh_error_message': None,
        'token_last_refreshed_at': None,
        'token_last_refreshed_at_display': None,
        'token_age_seconds': None,
        'token_cached': False,
        'session_keepalive_status': None,
        'session_keepalive_display_status': None,
        'session_keepalive_display_note': None,
        'session_keepalive_error_message': None,
        'session_keepalive_at': None,
        'session_keepalive_at_display': None,
        'session_keepalive_age_seconds': None,
        'session_transport_ready': False,
        'last_heartbeat_response_at': None,
        'last_heartbeat_response_at_display': None,
        'last_heartbeat_age_seconds': None,
        'last_heartbeat_sent_at': None,
        'last_heartbeat_sent_at_display': None,
        'last_heartbeat_sent_age_seconds': None,
        'ws_transport_ready': False,
        'last_business_activity_at': None,
        'last_business_activity_at_display': None,
        'last_business_activity_age_seconds': None,
        'last_sync_package_at': None,
        'last_sync_package_at_display': None,
        'last_sync_package_age_seconds': None,
        'last_user_chat_at': None,
        'last_user_chat_at_display': None,
        'last_user_chat_age_seconds': None,
        'last_stream_watchdog_reconnect_at': None,
        'last_stream_watchdog_reconnect_at_display': None,
        'last_stream_watchdog_reconnect_age_seconds': None,
        'last_message_received_at': None,
        'last_message_received_at_display': None,
        'last_message_age_seconds': None,
        'last_successful_connection_at': None,
        'last_successful_connection_at_display': None,
        'state_last_changed_at': None,
        'state_last_changed_at_display': None,
        'cookie_refresh_enabled': None,
        'manual_refresh_active': False,
        'auth_recovery_owner': None,
    }
    if not cleaned_cid:
        return runtime_status

    live_instance = None
    try:
        if cookie_manager.manager:
            live_instance = getattr(cookie_manager.manager, 'live_instances', {}).get(cleaned_cid)
    except Exception:
        live_instance = None

    try:
        from XianyuAutoAsync import XianyuLive
    except Exception as e:
        if not live_instance:
            runtime_status['error'] = f"import_failed: {mask_sensitive_text(e)}"
            return runtime_status
    else:
        if not live_instance:
            live_instance = XianyuLive.get_instance(cleaned_cid)
        auth_recovery_state = XianyuLive.get_auth_recovery_lock_state(cleaned_cid)
        runtime_status['auth_recovery_owner'] = (auth_recovery_state or {}).get('owner')

    if not live_instance:
        return runtime_status

    connection_state = getattr(live_instance, 'connection_state', None)
    connection_state_value = getattr(connection_state, 'value', str(connection_state or 'unknown'))
    ws = getattr(live_instance, 'ws', None)
    session = getattr(live_instance, 'session', None)
    ws_transport_ready = bool(ws and not getattr(ws, 'closed', False))
    session_transport_ready = bool(session and not getattr(session, 'closed', True))
    token_cached = bool(getattr(live_instance, 'current_token', None))
    token_refresh_status = getattr(live_instance, 'last_token_refresh_status', None)
    session_keepalive_status = getattr(live_instance, 'last_session_keepalive_status', None)
    heartbeat_response_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_heartbeat_response', 0))
    heartbeat_sent_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_heartbeat_time', 0))
    token_refreshed_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_token_refresh_time', 0))
    session_keepalive_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_session_keepalive_time', 0))
    last_non_heartbeat_message_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_non_heartbeat_message_time', 0))
    last_sync_package_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_sync_package_time', 0))
    last_user_chat_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_user_chat_time', 0))
    last_stream_watchdog_reconnect_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_stream_watchdog_reconnect_time', 0))
    last_message_received_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_message_received_time', 0))
    last_successful_connection_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_successful_connection', 0))
    last_state_changed_at = _normalize_runtime_timestamp(getattr(live_instance, 'last_state_change_time', 0))

    heartbeat_interval = max(1, int(getattr(live_instance, 'heartbeat_interval', 15) or 15))
    heartbeat_timeout = max(1, int(getattr(live_instance, 'heartbeat_timeout', 30) or 30))
    token_refresh_interval = max(60, int(getattr(live_instance, 'token_refresh_interval', 72000) or 72000))
    token_retry_interval = max(30, int(getattr(live_instance, 'token_retry_interval', 180) or 180))
    session_keepalive_interval = max(60, int(getattr(live_instance, 'session_keepalive_interval', 600) or 600))
    session_keepalive_retry_interval = max(30, int(getattr(live_instance, 'session_keepalive_retry_interval', 180) or 180))
    stream_watchdog_grace_period = max(30, int(getattr(live_instance, 'stream_watchdog_grace_period', heartbeat_interval * 4) or heartbeat_interval * 4))
    message_stream_watchdog_timeout = max(60, int(getattr(live_instance, 'message_stream_watchdog_timeout', session_keepalive_interval * 3) or session_keepalive_interval * 3))

    ws_ready_window = max(heartbeat_timeout * 2, heartbeat_interval * 3, 45)
    recent_connection_window = max(heartbeat_interval + 5, 20)
    session_ready_window = max(session_keepalive_interval + session_keepalive_retry_interval + 30, 180)
    token_ready_window = max(token_refresh_interval + token_retry_interval, 300)
    now = time.time()

    recent_connection = _is_runtime_timestamp_recent(last_successful_connection_at, recent_connection_window)
    recent_heartbeat_ok = _is_runtime_timestamp_recent(heartbeat_response_at, ws_ready_window)
    recent_session_success = (
        session_keepalive_status == 'success'
        and _is_runtime_timestamp_recent(session_keepalive_at, session_ready_window)
    )
    recent_token_success = (
        token_refresh_status == 'success'
        and _is_runtime_timestamp_recent(token_refreshed_at, token_ready_window)
    )

    token_explicit_failure_statuses = {
        'captcha_max_retries_exceeded',
        'token_expired_recovery_failed',
        'token_refresh_failed',
        'token_refresh_exception',
        'token_init_failed',
    }
    session_display_status = session_keepalive_status
    session_display_note = None
    if (
        session_keepalive_status in {'auth_failed', 'api_failed', 'network_failed', 'response_parse_failed', 'exception'}
        and recent_token_success
        and session_transport_ready
    ):
        session_display_status = 'recovered'
        session_display_note = '轻保活最近一次失败，但已由后续 Token 恢复流程兜底恢复'

    ws_ready = (
        connection_state_value == 'connected'
        and ws_transport_ready
        and (recent_heartbeat_ok or recent_connection)
    )
    session_ready = (
        session_transport_ready
        and (
            recent_session_success
            or recent_token_success
        )
    )
    token_ready = (
        token_cached
        and token_refresh_status not in token_explicit_failure_statuses
        and (
            recent_token_success
            or (ws_ready and token_refresh_status in (None, 'success', 'started'))
            or (
                token_refresh_status is None
                and _is_runtime_timestamp_recent(token_refreshed_at, token_ready_window)
            )
        )
    )

    actual_business_activity_at = None
    if last_non_heartbeat_message_at is not None:
        if last_successful_connection_at is None or last_non_heartbeat_message_at > last_successful_connection_at:
            actual_business_activity_at = last_non_heartbeat_message_at

    connected_for_seconds = None
    if last_successful_connection_at is not None:
        connected_for_seconds = max(0, int(now - last_successful_connection_at))

    business_idle_reference = actual_business_activity_at or last_successful_connection_at
    business_idle_seconds = None
    if business_idle_reference is not None:
        business_idle_seconds = max(0, int(now - business_idle_reference))

    recent_watchdog_reconnect = _is_runtime_timestamp_recent(
        last_stream_watchdog_reconnect_at,
        message_stream_watchdog_timeout,
    )
    stream_stale_now = bool(
        ws_ready
        and recent_heartbeat_ok
        and connected_for_seconds is not None
        and connected_for_seconds >= stream_watchdog_grace_period
        and business_idle_seconds is not None
        and business_idle_seconds >= message_stream_watchdog_timeout
    )

    if connection_state_value in {'connecting', 'reconnecting'}:
        message_stream_status = 'recovering'
        message_stream_ready = False
    elif connection_state_value != 'connected' or not ws_transport_ready:
        message_stream_status = 'connection_unready'
        message_stream_ready = False
    elif stream_stale_now:
        message_stream_status = 'suspected_stale'
        message_stream_ready = False
    else:
        message_stream_ready = True
        if connected_for_seconds is not None and connected_for_seconds < stream_watchdog_grace_period and actual_business_activity_at is None:
            message_stream_status = 'warming_up'
        elif (
            recent_watchdog_reconnect
            and actual_business_activity_at is not None
            and last_stream_watchdog_reconnect_at is not None
            and actual_business_activity_at > last_stream_watchdog_reconnect_at
        ):
            message_stream_status = 'recovered'
        elif actual_business_activity_at is not None:
            message_stream_status = 'healthy'
        else:
            message_stream_status = 'watching'

    business_note = (
        f"最近非心跳业务包：{_format_runtime_timestamp(actual_business_activity_at)}"
        if actual_business_activity_at is not None else
        "当前连接尚未收到非心跳业务包"
    )
    sync_note = (
        f"最近同步包：{_format_runtime_timestamp(last_sync_package_at)}"
        if last_sync_package_at is not None else
        "当前连接尚未收到同步包"
    )
    user_chat_note = (
        f"最近真实买家消息：{_format_runtime_timestamp(last_user_chat_at)}"
        if last_user_chat_at is not None else
        "当前连接尚未收到真实买家消息"
    )
    message_stream_note_parts = [business_note]
    if message_stream_status == 'suspected_stale':
        message_stream_note_parts.extend([sync_note, user_chat_note])
    elif recent_watchdog_reconnect and last_stream_watchdog_reconnect_at is not None:
        message_stream_note_parts.append(
            f"最近一次假在线重连：{_format_runtime_timestamp(last_stream_watchdog_reconnect_at)}"
        )
        if actual_business_activity_at is None:
            message_stream_note_parts.append(sync_note)
    else:
        message_stream_note_parts.append(sync_note)
    message_stream_note = ' · '.join(message_stream_note_parts)

    runtime_status.update({
        'instance_exists': True,
        'running': True,
        'connection_state': connection_state_value,
        'ws_ready': ws_ready,
        'session_ready': session_ready,
        'has_current_token': token_ready,
        'message_stream_ready': message_stream_ready,
        'message_stream_status': message_stream_status,
        'message_stream_note': message_stream_note,
        'token_cached': token_cached,
        'token_refresh_status': token_refresh_status,
        'token_refresh_error_message': getattr(live_instance, 'last_token_refresh_error_message', None),
        'token_last_refreshed_at': token_refreshed_at,
        'token_last_refreshed_at_display': _format_runtime_timestamp(token_refreshed_at),
        'token_age_seconds': _get_runtime_age_seconds(token_refreshed_at),
        'session_keepalive_status': session_keepalive_status,
        'session_keepalive_display_status': session_display_status,
        'session_keepalive_display_note': session_display_note,
        'session_keepalive_error_message': getattr(live_instance, 'last_session_keepalive_error_message', None),
        'session_keepalive_at': session_keepalive_at,
        'session_keepalive_at_display': _format_runtime_timestamp(session_keepalive_at),
        'session_keepalive_age_seconds': _get_runtime_age_seconds(session_keepalive_at),
        'session_transport_ready': session_transport_ready,
        'last_heartbeat_response_at': heartbeat_response_at,
        'last_heartbeat_response_at_display': _format_runtime_timestamp(heartbeat_response_at),
        'last_heartbeat_age_seconds': _get_runtime_age_seconds(heartbeat_response_at),
        'last_heartbeat_sent_at': heartbeat_sent_at,
        'last_heartbeat_sent_at_display': _format_runtime_timestamp(heartbeat_sent_at),
        'last_heartbeat_sent_age_seconds': _get_runtime_age_seconds(heartbeat_sent_at),
        'ws_transport_ready': ws_transport_ready,
        'last_business_activity_at': actual_business_activity_at,
        'last_business_activity_at_display': _format_runtime_timestamp(actual_business_activity_at),
        'last_business_activity_age_seconds': _get_runtime_age_seconds(actual_business_activity_at),
        'last_sync_package_at': last_sync_package_at,
        'last_sync_package_at_display': _format_runtime_timestamp(last_sync_package_at),
        'last_sync_package_age_seconds': _get_runtime_age_seconds(last_sync_package_at),
        'last_user_chat_at': last_user_chat_at,
        'last_user_chat_at_display': _format_runtime_timestamp(last_user_chat_at),
        'last_user_chat_age_seconds': _get_runtime_age_seconds(last_user_chat_at),
        'last_stream_watchdog_reconnect_at': last_stream_watchdog_reconnect_at,
        'last_stream_watchdog_reconnect_at_display': _format_runtime_timestamp(last_stream_watchdog_reconnect_at),
        'last_stream_watchdog_reconnect_age_seconds': _get_runtime_age_seconds(last_stream_watchdog_reconnect_at),
        'last_message_received_at': last_message_received_at,
        'last_message_received_at_display': _format_runtime_timestamp(last_message_received_at),
        'last_message_age_seconds': _get_runtime_age_seconds(last_message_received_at),
        'last_successful_connection_at': last_successful_connection_at,
        'last_successful_connection_at_display': _format_runtime_timestamp(last_successful_connection_at),
        'state_last_changed_at': last_state_changed_at,
        'state_last_changed_at_display': _format_runtime_timestamp(last_state_changed_at),
        'cookie_refresh_enabled': getattr(live_instance, 'cookie_refresh_enabled', None),
        'manual_refresh_active': bool(XianyuLive.is_manual_refresh_active(cleaned_cid, allow_handoff_recovery=True)),
    })
    return runtime_status


async def _run_live_instance_on_manager_loop(
    cookie_id: str,
    coroutine_factory: Callable[[], Awaitable[Any]],
    *,
    timeout: Optional[float] = None,
) -> Any:
    """将运行中账号实例的协程调度回 CookieManager 所属事件循环执行。"""
    manager = getattr(cookie_manager, 'manager', None)
    target_loop = getattr(manager, 'loop', None)
    if not target_loop:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    if hasattr(target_loop, 'is_closed') and target_loop.is_closed():
        raise HTTPException(status_code=500, detail="账号事件循环已关闭")

    try:
        current_loop = asyncio.get_running_loop()
    except RuntimeError:
        current_loop = None

    if current_loop is target_loop:
        return await coroutine_factory()

    if not target_loop.is_running():
        raise HTTPException(status_code=500, detail="账号事件循环未运行")

    thread_future = asyncio.run_coroutine_threadsafe(coroutine_factory(), target_loop)
    wrapped_future = asyncio.wrap_future(thread_future)

    try:
        if timeout and timeout > 0:
            return await asyncio.wait_for(wrapped_future, timeout=timeout)
        return await wrapped_future
    except asyncio.TimeoutError:
        thread_future.cancel()
        raise HTTPException(status_code=504, detail="账号处理超时，请稍后重试")





@app.get("/cookies")
def list_cookies(current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)
    return list(user_cookies.keys())


@app.get("/cookies/details")
def get_cookies_details(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有Cookie的详细信息（包括值和状态）"""
    if cookie_manager.manager is None:
        return []

    user_cookies = _get_user_cookies_map(current_user)

    result = []
    for cookie_id, cookie_value in user_cookies.items():
        cookie_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
        auto_confirm = db_manager.get_auto_confirm(cookie_id)
        auto_comment = db_manager.get_auto_comment(cookie_id)
        # 获取备注信息
        cookie_details = db_manager.get_cookie_details(cookie_id)
        remark = cookie_details.get('remark', '') if cookie_details else ''
        status_note = cookie_details.get('status_note', '') if cookie_details else ''
        username = cookie_details.get('username', '') if cookie_details else ''
        has_password = bool(cookie_details.get('password')) if cookie_details else False

        result.append({
            'id': cookie_id,
            'value': mask_cookie_value(cookie_value),
            'has_cookie_value': bool(cookie_value),
            'enabled': cookie_enabled,
            'auto_confirm': auto_confirm,
            'auto_comment': auto_comment,
            'remark': remark,
            'status_note': status_note,
            'username': username,
            'has_password': has_password,
            'pause_duration': cookie_details.get('pause_duration', 10) if cookie_details else 10,
            'runtime_status': _build_live_runtime_status(cookie_id),
        })
    return result


@app.get("/api/announcement")
def get_dashboard_announcement(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取仪表盘公告，优先读取 GitHub 公告文件，本地文件兜底。"""
    try:
        _ = current_user['user_id']
        snapshot = _get_dashboard_announcement_payload()
        return {
            'success': True,
            'announcement': snapshot.get('current'),
            'current': snapshot.get('current'),
            'history': snapshot.get('history') or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取仪表盘公告失败: {mask_sensitive_text(e)}")
        return {
            'success': False,
            'announcement': None,
            'current': None,
            'history': [],
            'message': safe_client_error("获取公告失败，请稍后重试"),
        }


@app.post("/cookies")
def add_cookie(item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 添加cookie时绑定到当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager

        log_with_user('info', f"尝试添加Cookie: {item.id}, 当前用户ID: {user_id}, 用户名: {current_user.get('username', 'unknown')}", current_user)

        # 检查cookie是否已存在且属于其他用户
        existing_cookies = db_manager.get_all_cookies()
        if item.id in existing_cookies:
            # 检查是否属于当前用户
            user_cookies = db_manager.get_all_cookies(user_id)
            if item.id not in user_cookies:
                log_with_user('warning', f"Cookie ID冲突: {item.id} 已被其他用户使用", current_user)
                raise HTTPException(status_code=400, detail="该Cookie ID已被其他用户使用")

        # 保存到数据库时指定用户ID
        db_manager.save_cookie(item.id, item.value, user_id)

        # 添加到CookieManager，同时指定用户ID
        cookie_manager.manager.add_cookie(item.id, item.value, user_id=user_id)
        log_with_user('info', f"Cookie添加成功: {item.id}", current_user)
        return {"msg": "success"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"添加Cookie失败: {item.id} - {mask_sensitive_text(e)}", current_user)
        raise HTTPException(status_code=400, detail=safe_client_error("添加Cookie失败，请检查输入后重试"))


@app.put('/cookies/{cid}')
def update_cookie(cid: str, item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None

        # 使用 update_cookie_account_info 更新（只更新cookie值，不覆盖其他字段）
        success = db_manager.update_cookie_account_info(cid, cookie_value=item.value)
        
        if not success:
            raise HTTPException(status_code=400, detail="更新Cookie失败")
        
        # 只有当 cookie 值真的发生变化时才重启任务
        if item.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, item.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")
        
        return {'msg': 'updated'}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新Cookie失败: {cid} - {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("更新Cookie失败，请稍后重试"))


class CookieAccountInfo(BaseModel):
    """账号信息更新模型"""
    value: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    show_browser: Optional[bool] = None


@app.post("/cookie/{cid}/account-info")
def update_cookie_account_info(cid: str, info: CookieAccountInfo, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号信息（Cookie、用户名、密码、显示浏览器设置）"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None
        
        # 更新数据库
        success = db_manager.update_cookie_account_info(
            cid, 
            cookie_value=info.value,
            username=info.username,
            password=info.password,
            show_browser=info.show_browser
        )
        
        if not success:
            raise HTTPException(status_code=400, detail="更新账号信息失败")
        
        # 只有当 cookie 值真的发生变化时才重启任务
        if info.value is not None and info.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, info.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")
        
        return {'msg': 'updated', 'success': True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新账号信息失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("更新账号信息失败，请稍后重试"))


@app.get("/cookie/{cid}/details")
def get_cookie_account_details(cid: str, include_secrets: bool = False, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号详细信息（包括用户名、密码、显示浏览器设置）"""
    try:
        cid = _ensure_cookie_access(cid, current_user)

        # 获取详细信息
        details = db_manager.get_cookie_details(cid)
        
        if not details:
            raise HTTPException(status_code=404, detail="账号不存在")

        runtime_status = _build_live_runtime_status(cid)

        if not include_secrets:
            details = {
                **details,
                'value': mask_cookie_value(details.get('value')),
                'password': mask_secret_value(details.get('password')),
                'proxy_pass': mask_secret_value(details.get('proxy_pass')),
                'has_cookie_value': bool(details.get('value')),
                'has_password': bool(details.get('password')),
                'has_proxy_pass': bool(details.get('proxy_pass')),
                'runtime_status': runtime_status,
            }
        else:
            details = {
                **details,
                'runtime_status': runtime_status,
            }
        
        return details
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取账号详情失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("获取账号详情失败，请稍后重试"))


@app.get("/cookies/{cid}/runtime-status")
def get_cookie_runtime_status(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号运行态状态，便于排查保活/连接问题。"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        return {
            'cookie_id': cid,
            'runtime_status': _build_live_runtime_status(cid),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取账号运行态失败: {cid} - {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("获取账号运行态失败，请稍后重试"))


@app.get("/cookies/{cid}/conversations/{conversation_id}/history")
async def get_conversation_history(
    cid: str,
    conversation_id: str,
    page_size: int = 20,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取指定会话的历史消息。"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        normalized_conversation_id = str(conversation_id or '').strip().split('@')[0]
        if not normalized_conversation_id:
            raise HTTPException(status_code=400, detail="缺少会话ID")

        normalized_page_size = max(1, min(int(page_size or 20), 100))

        from XianyuAutoAsync import XianyuLive
        live_instance = XianyuLive.get_instance(cid)
        if not live_instance:
            raise HTTPException(status_code=400, detail="账号未启动，暂无法查询历史消息")

        log_with_user(
            'info',
            f"开始查询账号 {cid} 会话 {normalized_conversation_id} 的历史消息，page_size={normalized_page_size}",
            current_user
        )
        history_messages = await _run_live_instance_on_manager_loop(
            cid,
            lambda: live_instance.list_all_conversations(
                normalized_conversation_id,
                page_size=normalized_page_size,
            ),
            timeout=60,
        )
        return {
            'success': True,
            'cookie_id': cid,
            'conversation_id': normalized_conversation_id,
            'page_size': normalized_page_size,
            'count': len(history_messages),
            'messages': history_messages,
            'runtime_status': _build_live_runtime_status(cid),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史消息失败: {cid}/{conversation_id} - {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("获取历史消息失败，请稍后重试"))


@app.post("/cookies/{cid}/session-keepalive")
async def trigger_session_keepalive(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """手动触发一次轻量会话保活。"""
    try:
        cid = _ensure_cookie_access(cid, current_user)

        from XianyuAutoAsync import XianyuLive
        live_instance = XianyuLive.get_instance(cid)
        if not live_instance:
            try:
                live_instance = getattr(cookie_manager.manager, 'live_instances', {}).get(cid) if cookie_manager.manager else None
            except Exception:
                live_instance = None

        log_with_user('info', f"手动触发账号 {cid} 的轻量会话保活", current_user)
        used_temporary_instance = False

        if live_instance:
            keepalive_ok = await _run_live_instance_on_manager_loop(
                cid,
                lambda: live_instance.keep_session_alive(),
                timeout=40,
            )
        else:
            # 账号刚完成扫码/手动刷新、或旧误暂停导致主任务尚未恢复时，仍允许用数据库中的
            # 最新 Cookie 做一次 one-shot 轻保活；普通扫码登录不应因为“实例未注册”而无法验证会话。
            cookie_value = db_manager.get_cookie(cid)
            if not cookie_value:
                raise HTTPException(status_code=400, detail="账号Cookie不存在，暂无法执行轻量保活")

            async def _run_temporary_keepalive():
                temp_live = XianyuLive(cookie_value, cookie_id=cid, register_instance=False)
                try:
                    return await temp_live.keep_session_alive()
                finally:
                    try:
                        await temp_live.close_session()
                    except Exception as close_e:
                        logger.warning(f"临时轻量保活关闭会话失败: {cid} - {mask_sensitive_text(close_e)}")

            keepalive_ok = await _run_live_instance_on_manager_loop(
                cid,
                _run_temporary_keepalive,
                timeout=40,
            )
            used_temporary_instance = True

        runtime_status = _build_live_runtime_status(cid)
        return {
            'success': keepalive_ok,
            'cookie_id': cid,
            'message': '轻量会话保活成功' if keepalive_ok else '轻量会话保活失败',
            'runtime_status': runtime_status,
            'temporary_instance': used_temporary_instance,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"手动轻量保活失败: {cid} - {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("手动轻量保活失败，请稍后重试"))


# ========================= 代理配置相关接口 =========================

class ProxyConfig(BaseModel):
    """代理配置模型"""
    proxy_type: Optional[str] = 'none'  # none/http/https/socks5
    proxy_host: Optional[str] = ''
    proxy_port: Optional[int] = 0
    proxy_user: Optional[str] = ''
    proxy_pass: Optional[str] = ''


@app.get("/cookie/{cid}/proxy")
def get_cookie_proxy_config(cid: str, include_secret: bool = False, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的代理配置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取代理配置
        proxy_config = db_manager.get_cookie_proxy_config(cid)

        if not include_secret:
            proxy_config = {
                **proxy_config,
                'proxy_pass': mask_secret_value(proxy_config.get('proxy_pass')),
                'has_proxy_pass': bool(proxy_config.get('proxy_pass')),
            }
        
        return {
            'success': True,
            'data': proxy_config
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取代理配置失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("获取代理配置失败，请稍后重试"))


@app.post("/cookie/{cid}/proxy")
def update_cookie_proxy_config(cid: str, config: ProxyConfig, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的代理配置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证代理类型
        valid_proxy_types = ['none', 'http', 'https', 'socks5']
        if config.proxy_type not in valid_proxy_types:
            raise HTTPException(status_code=400, detail=f"无效的代理类型，支持的类型: {', '.join(valid_proxy_types)}")

        # 如果设置了代理类型（非none），验证必要字段
        if config.proxy_type != 'none':
            if not config.proxy_host:
                raise HTTPException(status_code=400, detail="代理地址不能为空")
            if not config.proxy_port or config.proxy_port <= 0:
                raise HTTPException(status_code=400, detail="代理端口无效")

        # 更新数据库
        success = db_manager.update_cookie_proxy_config(
            cid,
            proxy_type=config.proxy_type,
            proxy_host=config.proxy_host,
            proxy_port=config.proxy_port,
            proxy_user=config.proxy_user,
            proxy_pass=config.proxy_pass
        )
        
        if not success:
            raise HTTPException(status_code=400, detail="更新代理配置失败")
        
        # 重启账号任务以应用新的代理配置
        logger.info(f"代理配置已更新，重启账号任务: {cid}")
        cookie_value = user_cookies.get(cid)
        if cookie_value:
            cookie_manager.manager.update_cookie(cid, cookie_value, save_to_db=False)
        
        return {
            'success': True,
            'msg': '代理配置已更新，账号任务已重启'
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新代理配置失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=400, detail=safe_client_error("更新代理配置失败，请稍后重试"))


# ========================= 账号密码登录相关接口 =========================

def _new_risk_log_session_id(prefix: str = 'risk') -> str:
    return f"{prefix}_{secrets.token_hex(8)}"


def _build_risk_event_meta(base: Optional[Dict[str, Any]] = None, **extra_fields) -> Optional[Dict[str, Any]]:
    payload: Dict[str, Any] = {}
    if isinstance(base, dict):
        payload.update({key: value for key, value in base.items() if value is not None})
    payload.update({key: value for key, value in extra_fields.items() if value is not None})
    return payload or None


def _is_password_login_verification_timeout_message(message: str) -> bool:
    normalized = str(message or '').strip()
    if not normalized:
        return False

    if ('超时' in normalized or '失效' in normalized) and '重新发起验证' in normalized:
        return True

    timeout_markers = (
        '验证超时',
        '二维码已失效',
        '请重新扫码',
    )
    return any(marker in normalized for marker in timeout_markers)


def _derive_password_login_verification_failure_result_code(error_message: str) -> str:
    normalized = str(error_message or '').strip()
    if '二维码' in normalized:
        return 'qr_verify_timed_out' if _is_password_login_verification_timeout_message(normalized) else 'qr_verify_failed'
    if '人脸' in normalized:
        return 'face_verify_timed_out' if _is_password_login_verification_timeout_message(normalized) else 'face_verify_failed'
    if '短信' in normalized:
        return 'sms_verify_timed_out' if _is_password_login_verification_timeout_message(normalized) else 'sms_verify_failed'
    return 'verification_timed_out' if _is_password_login_verification_timeout_message(normalized) else 'verification_failed'


def _update_session_risk_log(
    session_id: str,
    status: str,
    processing_result: str = None,
    error_message: str = None,
    result_code: str = None,
    event_meta: Optional[Dict[str, Any]] = None,
):
    """更新登录会话关联的风控日志状态"""
    try:
        session = password_login_sessions.get(session_id)
        if not session:
            return
        log_id = session.get('risk_control_log_id')
        if not log_id:
            return

        risk_session_id = session.get('risk_session_id') or session_id
        duration_ms = None
        started_at = session.get('timestamp')
        if started_at:
            duration_ms = max(0, int((time.time() - float(started_at)) * 1000))

        if not result_code:
            refresh_mode = bool(session.get('refresh_mode'))
            if status == 'success':
                result_code = 'manual_cookie_refresh_success' if refresh_mode else 'password_login_success'
            elif status == 'failed':
                result_code = 'manual_cookie_refresh_failed' if refresh_mode else 'password_login_failed'

        merged_meta = _build_risk_event_meta(
            {
                'account_id': session.get('account_id'),
                'show_browser': session.get('show_browser'),
                'refresh_mode': bool(session.get('refresh_mode')),
            },
            **(event_meta or {}),
        )

        db_manager.update_risk_control_log(
            log_id=log_id,
            session_id=risk_session_id,
            processing_status=status,
            processing_result=processing_result,
            error_message=error_message,
            result_code=result_code,
            event_meta=merged_meta,
            duration_ms=duration_ms,
        )
    except Exception as e:
        logger.error(f"更新风控日志状态失败: {e}")


def _close_password_login_pending_verification_risk_logs(
    session_id: str,
    status: str,
    error_message: str = None,
    processing_result: str = None,
    result_code: str = None,
    event_meta: Optional[Dict[str, Any]] = None,
) -> int:
    """收口同一账密登录链路下遗留的 processing 验证风控日志。"""
    try:
        session = password_login_sessions.get(session_id)
        if not session:
            return 0

        risk_session_id = session.get('risk_session_id') or session_id
        if not risk_session_id:
            return 0

        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute(
                '''
                SELECT id
                FROM risk_control_logs
                WHERE session_id = ?
                  AND processing_status = 'processing'
                  AND event_type IN ('qr_verify', 'face_verify', 'sms_verify', 'unknown')
                ORDER BY id ASC
                ''',
                (risk_session_id,)
            )
            pending_rows = cursor.fetchall() or []

        if not pending_rows:
            return 0

        duration_ms = None
        started_at = session.get('timestamp')
        if started_at:
            duration_ms = max(0, int((time.time() - float(started_at)) * 1000))

        processing_status = 'success' if str(status or '').strip().lower() == 'success' else 'failed'
        if result_code:
            resolved_result_code = result_code
        elif processing_status == 'success':
            resolved_result_code = 'manual_cookie_refresh_verification_completed' if session.get('refresh_mode') else 'password_login_verification_completed'
        else:
            resolved_result_code = _derive_password_login_verification_failure_result_code(error_message)

        if processing_result is None:
            if processing_status == 'success':
                processing_result = '人工验证已完成，登录流程已成功收尾'
            else:
                processing_result = error_message or '验证流程已结束'

        merged_meta = _build_risk_event_meta(
            {
                'account_id': session.get('account_id'),
                'show_browser': session.get('show_browser'),
                'refresh_mode': bool(session.get('refresh_mode')),
            },
            **(event_meta or {}),
        )

        updated_count = 0
        for row in pending_rows:
            log_id = row[0] if isinstance(row, (tuple, list)) else row
            if not log_id:
                continue
            updated = db_manager.update_risk_control_log(
                log_id=log_id,
                processing_result=processing_result,
                processing_status=processing_status,
                error_message=error_message,
                session_id=risk_session_id,
                trigger_scene='manual_password_refresh' if session.get('refresh_mode') else 'password_login',
                result_code=resolved_result_code,
                event_meta=merged_meta,
                duration_ms=duration_ms,
            )
            if updated:
                updated_count += 1

        return updated_count
    except Exception as e:
        logger.error(f"收口待处理验证风控日志失败: {e}")
        return 0


def _set_password_login_session_status(session_id: str, status: str, **fields):
    session = password_login_sessions.get(session_id)
    if not session:
        return False

    current_status = str(session.get('status') or '').strip().lower()
    next_status = str(status or '').strip().lower()
    if current_status in PASSWORD_LOGIN_TERMINAL_STATUSES and next_status != current_status:
        logger.info(
            f"忽略密码登录会话终态回退: session_id={session_id}, current_status={current_status}, next_status={next_status}"
        )
        return False

    session['status'] = status
    session.update(fields)

    if next_status == 'success':
        session['error'] = None
        session['verification_url'] = None
        session['screenshot_path'] = None
        session['qr_code_url'] = None
        session['verification_type'] = None

    if next_status in PASSWORD_LOGIN_TERMINAL_STATUSES:
        session['completed_at'] = time.time()
    else:
        session['completed_at'] = None

    return True


def _finalize_password_login_session_failure(
    session_id: str,
    error_message: str,
    *,
    result_code: str = None,
    event_meta: Optional[Dict[str, Any]] = None,
) -> bool:
    session = password_login_sessions.get(session_id)
    if not session:
        return False

    extra_fields: Dict[str, Any] = {}
    if _is_password_login_verification_timeout_message(error_message):
        extra_fields.update(
            verification_url=None,
            screenshot_path=None,
            qr_code_url=None,
            verification_type=None,
        )

    _set_password_login_session_status(
        session_id,
        'failed',
        error=error_message,
        **extra_fields,
    )
    _update_session_risk_log(
        session_id,
        'failed',
        error_message=(error_message or '')[:200],
        result_code=result_code,
        event_meta=event_meta,
    )
    _close_password_login_pending_verification_risk_logs(
        session_id,
        'failed',
        error_message=error_message,
        event_meta=event_meta,
    )
    return True


def _get_latest_password_login_session_for_account(
    account_id: str,
    user_id: Optional[int] = None,
) -> Optional[Dict[str, Any]]:
    target_account_id = str(account_id)
    matched_sessions = []

    for session in password_login_sessions.values():
        if str(session.get('account_id')) != target_account_id:
            continue
        if user_id is not None and session.get('user_id') != user_id:
            continue
        matched_sessions.append(session)

    if not matched_sessions:
        return None

    return max(
        matched_sessions,
        key=lambda item: (
            float(item.get('timestamp') or 0),
            float(item.get('completed_at') or 0),
        ),
    )


def _is_timed_out_verification_risk_log(log: Optional[Dict[str, Any]]) -> bool:
    if not isinstance(log, dict):
        return False

    result_code = str(log.get('result_code') or '').strip().lower()
    if result_code == 'verification_timed_out' or result_code.endswith('_timed_out'):
        return True

    for field in ('error_message', 'processing_result', 'event_description'):
        if _is_password_login_verification_timeout_message(log.get(field)):
            return True

    return False


def _get_latest_verification_risk_log_for_account(account_id: str) -> Optional[Dict[str, Any]]:
    verification_event_types = {'qr_verify', 'face_verify', 'sms_verify', 'unknown'}
    logs = db_manager.get_risk_control_logs(cookie_id=str(account_id), limit=20)
    for log in logs:
        if str(log.get('event_type') or '').strip() in verification_event_types:
            return log
    return None


def _build_face_verification_screenshot_info(account_id: str, file_path: str) -> Dict[str, Any]:
    from datetime import datetime

    normalized_path = str(file_path or '').replace('\\', '/')
    filename = os.path.basename(normalized_path)
    stat = os.stat(normalized_path)
    return {
        'filename': filename,
        'account_id': account_id,
        'path': f'/static/uploads/images/{filename}',
        'size': stat.st_size,
        'created_time': stat.st_ctime,
        'created_time_str': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
    }


def _set_manual_cookie_import_session_status(session_id: str, status: str, **fields):
    session = manual_cookie_import_sessions.get(session_id)
    if not session:
        return

    session['status'] = status
    session.update(fields)

    if status in {'success', 'failed'}:
        session['completed_at'] = time.time()
    else:
        session['completed_at'] = None


def _empty_slider_session_stats() -> Dict[str, Any]:
    return {
        'has_data': False,
        'total_sessions': 0,
        'total_attempts': 0,
        'success_count': 0,
        'failure_count': 0,
        'processing_count': 0,
        'completed_sessions': 0,
        'success_rate': 0.0,
        'recent_success': None,
        'recent_failure': None,
        'accounts_with_sessions': 0,
        'accounts_with_failures': 0,
        'stats_mode': 'session',
        'summary_text': '暂无滑块验证记录',
        'selected_range': 'all',
        'range_label': '所有',
    }

async def _execute_password_login(session_id: str, account_id: str, account: str, password: str, show_browser: bool, user_id: int, current_user: Dict[str, Any]):
    """后台执行账号密码登录任务"""
    manual_refresh_acquired = False
    manual_refresh_owner = f"password_login:{session_id}"
    auth_recovery_owner = f"manual_password_login:{session_id}"
    auth_recovery_acquired = False
    login_thread_started = False
    manual_refresh_preflight_timeout = 45.0
    request_loop = asyncio.get_running_loop()
    try:
        log_with_user('info', f"开始执行账号密码登录任务: {session_id}, 账号: {account_id}", current_user)

        from XianyuAutoAsync import XianyuLive

        is_refresh_mode = password_login_sessions.get(session_id, {}).get('refresh_mode', False)
        auth_session_state = XianyuLive.begin_auth_recovery_session(
            account_id,
            auth_recovery_owner,
            mode='manual_cookie_refresh' if is_refresh_mode else 'manual_password_login',
            source=manual_refresh_owner,
            force_replace=False,
        )
        auth_recovery_acquired = auth_session_state.get('started', False)
        if auth_session_state.get('already_active'):
            active_owner = auth_session_state.get('active_owner', 'unknown')
            _set_password_login_session_status(
                session_id,
                'failed',
                error=f'该账号已有认证恢复流程进行中，请先完成当前验证或稍后再试（owner={active_owner}）'
            )
            _update_session_risk_log(session_id, 'failed', error_message=f'认证恢复流程进行中: {active_owner}')
            log_with_user('warning', f"账号已有认证恢复流程在执行，拒绝重复触发: {account_id}, owner={active_owner}", current_user)
            return

        if is_refresh_mode:
            manual_refresh_state = XianyuLive.begin_manual_refresh(account_id, source=manual_refresh_owner)
            manual_refresh_acquired = manual_refresh_state.get('started', False)
            if manual_refresh_state.get('already_active'):
                _set_password_login_session_status(
                    session_id,
                    'failed',
                    error='该账号正在执行手动刷新，请稍候再试'
                )
                _update_session_risk_log(session_id, 'failed', error_message='账号正在执行手动刷新')
                log_with_user('warning', f"账号已存在手动刷新任务，拒绝重复触发: {account_id}", current_user)
                return
        
        # 导入 XianyuSliderStealth
        from utils.xianyu_slider_stealth import XianyuSliderStealth
        import base64
        import io
        
        # 创建 XianyuSliderStealth 实例
        existing_cookie_info = db_manager.get_cookie_details(account_id) or {}
        proxy_config = db_manager.get_cookie_proxy_config(account_id)
        slider_instance = XianyuSliderStealth(
            user_id=account_id,
            enable_learning=True,
            headless=not show_browser,
            initial_cookies=existing_cookie_info.get('value', ''),
            proxy=proxy_config,
        )
        slider_instance.risk_session_id = password_login_sessions.get(session_id, {}).get('risk_session_id') or session_id
        slider_instance.risk_trigger_scene = 'manual_password_refresh' if is_refresh_mode else 'password_login'
        
        # 更新会话信息
        password_login_sessions[session_id]['slider_instance'] = slider_instance
        
        # 定义通知回调函数，用于检测到验证时返回验证链接或截图（同步函数）
        def notification_callback(
            message: str,
            screenshot_path: str = None,
            verification_url: str = None,
            screenshot_path_new: str = None,
            verification_type: str = None,
        ):
            """账号验证通知回调（同步）
            
            Args:
                message: 通知消息
                screenshot_path: 旧版截图路径（兼容参数）
                verification_url: 验证链接
                screenshot_path_new: 新版截图路径（新参数，优先使用）
                verification_type: 验证类型
            """
            try:
                # 优先使用新的截图路径参数
                actual_screenshot_path = screenshot_path_new if screenshot_path_new else screenshot_path
                verification_type_label = resolve_verification_type_label(
                    verification_type,
                    message,
                    verification_url,
                )

                if _is_password_login_verification_timeout_message(message):
                    _finalize_password_login_session_failure(session_id, message)
                    log_with_user('warning', f"密码登录会话检测到失效验证页，直接标记失败: {session_id}", current_user)
                    return
                
                # 优先使用截图路径，如果没有截图则使用验证链接
                if actual_screenshot_path and os.path.exists(actual_screenshot_path):
                    # 更新会话状态，保存截图路径
                    _set_password_login_session_status(
                        session_id,
                        'verification_required',
                        screenshot_path=actual_screenshot_path,
                        verification_url=None,
                        qr_code_url=None,
                        verification_type=verification_type_label,
                    )
                    log_with_user('info', f"账号验证截图已保存: {session_id}, 路径: {actual_screenshot_path}", current_user)
                    
                    # 发送通知到用户配置的渠道
                    def send_face_verification_notification():
                        """在后台线程中发送账号验证通知"""
                        try:
                            log_with_user('info', f"开始尝试发送账号验证通知: {account_id}", current_user)
                            notification_message = build_face_verify_notification(
                                account_id=account_id,
                                time_text=time.strftime('%Y-%m-%d %H:%M:%S'),
                                verification_type=verification_type_label,
                                verification_url=verification_url or '',
                                error_message=message,
                                has_screenshot=True,
                            )
                            notification_sent = dispatch_account_notifications_sync(
                                account_id,
                                notification_message,
                                title='闲鱼账号需要验证',
                                notification_type='face_verify',
                                attachment_path=actual_screenshot_path,
                            )
                            if notification_sent:
                                log_with_user('info', f"✅ 已发送账号验证通知: {account_id}", current_user)
                            else:
                                log_with_user('warning', f"账号验证通知未发送成功: {account_id}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"发送账号验证通知时出错: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                    
                    # 在后台线程中发送通知，避免阻塞登录流程
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"已启动账号验证通知发送线程: {account_id}", current_user)
                elif verification_url:
                    # 如果没有截图，使用验证链接（兼容旧版本）
                    _set_password_login_session_status(
                        session_id,
                        'verification_required',
                        verification_url=verification_url,
                        screenshot_path=None,
                        qr_code_url=None,
                        verification_type=verification_type_label,
                    )
                    log_with_user('info', f"账号验证链接已保存: {session_id}, URL: {verification_url}", current_user)
                    
                    # 发送通知到用户配置的渠道
                    def send_face_verification_notification():
                        """在后台线程中发送账号验证通知"""
                        try:
                            log_with_user('info', f"开始尝试发送账号验证通知: {account_id}", current_user)
                            notification_message = build_face_verify_notification(
                                account_id=account_id,
                                time_text=time.strftime('%Y-%m-%d %H:%M:%S'),
                                verification_type=verification_type_label,
                                verification_url=verification_url or '无',
                                error_message=message,
                                has_screenshot=False,
                            )
                            notification_sent = dispatch_account_notifications_sync(
                                account_id,
                                notification_message,
                                title='闲鱼账号需要验证',
                                notification_type='face_verify',
                            )
                            if notification_sent:
                                log_with_user('info', f"✅ 已发送账号验证通知: {account_id}", current_user)
                            else:
                                log_with_user('warning', f"账号验证通知未发送成功: {account_id}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"发送账号验证通知时出错: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                    
                    # 在后台线程中发送通知，避免阻塞登录流程
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"已启动账号验证通知发送线程: {account_id}", current_user)
            except Exception as e:
                log_with_user('error', f"处理账号验证通知失败: {str(e)}", current_user)
        
        # 调用登录方法（同步方法，需要在后台线程中执行）
        import threading

        def run_login():
            import asyncio  # 在函数开头导入，避免后续局部import导致UnboundLocalError
            from db_manager import db_manager  # 在函数开头导入，避免作用域问题
            from XianyuAutoAsync import XianyuLive
            try:
                cookies_dict = slider_instance.login_with_password_playwright(
                    account=account,
                    password=password,
                    show_browser=show_browser,
                    notification_callback=notification_callback,
                    force_clean_context=is_refresh_mode
                )
                
                if cookies_dict is None:
                    failure_message = slider_instance.last_login_error or '登录失败，请检查账号密码是否正确'
                    _finalize_password_login_session_failure(session_id, failure_message)
                    log_with_user('error', f"账号密码登录失败: {account_id}, 错误: {failure_message}", current_user)
                    return
                
                log_with_user('info', f"账号密码登录成功，获取到 {len(cookies_dict)} 个Cookie字段: {account_id}", current_user)
                
                # 检查是否已存在相同账号ID的Cookie
                existing_cookies = db_manager.get_all_cookies(user_id)
                is_new_account = account_id not in existing_cookies
                existing_cookie_value = existing_cookies.get(account_id, '') if not is_new_account else ''
                existing_cookie_dict = trans_cookies(existing_cookie_value) if existing_cookie_value else {}

                merge_result = XianyuLive.protected_merge_cookie_dicts(existing_cookie_dict, cookies_dict)
                if merge_result['incoming_missing_protected_fields']:
                    log_with_user(
                        'warning',
                        f"密码登录返回的Cookie快照缺少关键字段，将进行保护性合并: {', '.join(merge_result['incoming_missing_protected_fields'])}",
                        current_user
                    )
                if merge_result['preserved_protected_fields']:
                    log_with_user(
                        'warning',
                        f"密码登录保护性保留旧关键字段: {', '.join(merge_result['preserved_protected_fields'])}",
                        current_user
                    )
                if merge_result['account_switched']:
                    log_with_user('warning', f"检测到unb变化，按账号切换处理: {account_id}", current_user)

                merged_cookies_dict = merge_result['merged_cookies_dict']
                log_with_user(
                    'info',
                    f"manual_login_protected_merge incoming_count={merge_result.get('incoming_count', len(cookies_dict))} "
                    f"existing_count={merge_result.get('existing_count', len(existing_cookie_dict))} "
                    f"merged_count={merge_result.get('merged_count', len(merged_cookies_dict))} "
                    f"protected_preserved_fields={merge_result.get('preserved_protected_fields') or []} "
                    f"would_remove_fields={merge_result.get('would_remove_fields') or []} "
                    f"account_switched={merge_result.get('account_switched', False)}",
                    current_user
                )
                cookies_str = '; '.join([f"{k}={v}" for k, v in merged_cookies_dict.items()])

                if merge_result['missing_required_fields']:
                    missing_fields_text = ', '.join(merge_result['missing_required_fields'])
                    error_message = f"登录成功但Cookie核心字段仍缺失，未覆盖旧Cookie: {missing_fields_text}"
                    log_with_user('error', f"{error_message}: {account_id}", current_user)
                    _finalize_password_login_session_failure(
                        session_id,
                        error_message,
                        result_code='password_login_cookie_incomplete',
                        event_meta={
                            'missing_required_fields': merge_result['missing_required_fields'],
                            'incoming_missing_protected_fields': merge_result['incoming_missing_protected_fields'],
                            'preserved_protected_fields': merge_result['preserved_protected_fields'],
                        },
                    )
                    return

                if is_refresh_mode:
                    try:
                        log_with_user('info', f"刷新模式开始执行Token预检，确认新实例可直接恢复: {account_id}", current_user)
                        XianyuLive.mark_manual_refresh_handoff(account_id, source=manual_refresh_owner)
                        temp_xianyu = XianyuLive(
                            cookies_str=cookies_str,
                            cookie_id=account_id,
                            user_id=user_id,
                            register_instance=False,
                        )
                        preflight_future = asyncio.run_coroutine_threadsafe(
                            temp_xianyu.preflight_token_after_manual_refresh(),
                            request_loop,
                        )
                        try:
                            preflight_future.result(timeout=manual_refresh_preflight_timeout)
                        except concurrent.futures.TimeoutError as timeout_err:
                            preflight_future.cancel()
                            raise TimeoutError(
                                f"手动刷新后的Token预检在 {manual_refresh_preflight_timeout:.0f} 秒内未完成"
                            ) from timeout_err
                        cookies_str = temp_xianyu.cookies_str
                        merged_cookies_dict = trans_cookies(cookies_str)
                        log_with_user('info', f"刷新模式Token预检通过，将使用预检后的Cookie继续交接: {account_id}", current_user)
                    except Exception as preflight_err:
                        error_message = f"刷新模式认证预检失败，任务未切换: {str(preflight_err)}"
                        log_with_user('error', f"{error_message}: {account_id}", current_user)
                        _finalize_password_login_session_failure(
                            session_id,
                            error_message,
                            result_code='manual_refresh_preflight_failed',
                            event_meta={'account_id': account_id},
                        )
                        return
                
                # 保存账号密码和Cookie到数据库
                # 使用 update_cookie_account_info 来保存，它会自动处理新账号和现有账号的情况
                # 注意：刷新模式下不更新 show_browser，避免临时调试选项被永久保存
                update_success = db_manager.update_cookie_account_info(
                    account_id,
                    cookie_value=cookies_str,
                    username=account,
                    password=password,
                    show_browser=show_browser if not is_refresh_mode else None,  # 刷新模式不更新此字段
                    user_id=user_id  # 新账号时需要提供user_id
                )
                
                if update_success:
                    if is_new_account:
                        log_with_user('info', f"新账号Cookie和账号密码已保存: {account_id}", current_user)
                    else:
                        log_with_user('info', f"现有账号Cookie和账号密码已更新: {account_id}", current_user)
                else:
                    log_with_user('error', f"保存账号信息失败: {account_id}", current_user)
                
                # 统一走 CookieManager，确保任务登记、实例切换和运行态一致
                if cookie_manager.manager:
                    try:
                        if is_new_account:
                            cookie_manager.manager.add_cookie(account_id, cookies_str, user_id=user_id)
                            log_with_user('info', f"已将新账号加入cookie_manager并启动任务: {account_id}", current_user)
                        else:
                            cookie_manager.manager.update_cookie(account_id, cookies_str, save_to_db=False)
                            log_with_user('info', f"已更新cookie_manager并重启任务: {account_id}", current_user)
                    except Exception as manager_err:
                        action_desc = '启动新账号任务' if is_new_account else '切换账号任务'
                        log_with_user('warning', f"{action_desc}失败: {account_id}, 错误: {str(manager_err)}", current_user)
                
                if is_refresh_mode:
                    log_with_user('info', f"刷新模式已完成Token预检，直接切换到通过预检的新Cookie: {account_id}", current_user)
                else:
                    # 登录成功后，调用_refresh_cookies_via_browser刷新Cookie
                    try:
                        log_with_user('info', f"开始调用_refresh_cookies_via_browser刷新Cookie: {account_id}", current_user)
                        
                        # 创建临时的XianyuLive实例来刷新Cookie
                        temp_xianyu = XianyuLive(
                            cookies_str=cookies_str,
                            cookie_id=account_id,
                            user_id=user_id,
                            register_instance=False,
                        )
                        
                        # 重置扫码登录Cookie刷新标志，确保账号密码登录后能立即刷新
                        try:
                            temp_xianyu.reset_qr_cookie_refresh_flag()
                            log_with_user('info', f"已重置扫码登录Cookie刷新标志: {account_id}", current_user)
                        except Exception as reset_err:
                            log_with_user('debug', f"重置扫码登录Cookie刷新标志失败（不影响刷新）: {str(reset_err)}", current_user)
                        
                        # 在后台异步执行刷新（不阻塞主流程）
                        async def refresh_cookies_task():
                            try:
                                refresh_success = await temp_xianyu._refresh_cookies_via_browser(triggered_by_refresh_token=False)
                                if refresh_success:
                                    log_with_user('info', f"Cookie刷新成功: {account_id}", current_user)
                                    # 刷新成功后，从数据库获取更新后的Cookie
                                    updated_cookie_info = db_manager.get_cookie_details(account_id)
                                    if updated_cookie_info:
                                        refreshed_cookies = updated_cookie_info.get('value', '')
                                        if refreshed_cookies:
                                            # 更新cookie_manager中的Cookie
                                            if cookie_manager.manager:
                                                cookie_manager.manager.update_cookie(account_id, refreshed_cookies, save_to_db=False)
                                            log_with_user('info', f"已更新刷新后的Cookie到cookie_manager: {account_id}", current_user)
                                else:
                                    log_with_user('warning', f"Cookie刷新失败或跳过: {account_id}", current_user)
                            except Exception as refresh_e:
                                log_with_user('error', f"刷新Cookie时出错: {account_id}, 错误: {str(refresh_e)}", current_user)
                                import traceback
                                logger.error(traceback.format_exc())
                        
                        # 在后台线程中运行异步任务
                        # 由于run_login是在线程中运行的，需要创建新的事件循环
                        def run_async_refresh():
                            try:
                                import asyncio
                                # 创建新的事件循环
                                new_loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(new_loop)
                                try:
                                    new_loop.run_until_complete(refresh_cookies_task())
                                finally:
                                    new_loop.close()
                            except Exception as e:
                                log_with_user('error', f"运行异步刷新任务失败: {account_id}, 错误: {str(e)}", current_user)
                        
                        # 在后台线程中执行刷新任务
                        refresh_thread = threading.Thread(target=run_async_refresh, daemon=True)
                        refresh_thread.start()
                        
                    except Exception as refresh_err:
                        log_with_user('warning', f"调用_refresh_cookies_via_browser失败: {account_id}, 错误: {str(refresh_err)}", current_user)
                        # 刷新失败不影响登录成功
                
                # 更新会话状态
                _set_password_login_session_status(
                    session_id,
                    'success',
                    account_id=account_id,
                    is_new_account=is_new_account,
                    cookie_count=len(merged_cookies_dict)
                )
                _close_password_login_pending_verification_risk_logs(
                    session_id,
                    'success',
                    processing_result='人工验证已完成，登录流程已成功收尾',
                )
                # 更新风控日志状态
                _update_session_risk_log(
                    session_id,
                    'success',
                    processing_result='Cookie刷新成功，认证预检通过' if is_refresh_mode else 'Cookie刷新成功'
                )

                # 发送登录成功通知（使用模板系统）
                try:
                    # 根据模式选择不同模板
                    notify_refresh_mode = password_login_sessions[session_id].get('refresh_mode')
                    template_type = 'cookie_refresh_success' if notify_refresh_mode else 'password_login_success'

                    notification_message = render_notification_template(
                        template_type,
                        account_id=account_id,
                        time=time.strftime('%Y-%m-%d %H:%M:%S'),
                        cookie_count=str(len(merged_cookies_dict))
                    )

                    login_type = "刷新Cookie" if notify_refresh_mode else "密码登录"
                    notification_sent = dispatch_account_notifications_sync(
                        account_id,
                        notification_message,
                        title=f"{login_type}成功",
                        notification_type=template_type,
                    )
                    if notification_sent:
                        log_with_user('info', f"已发送{login_type}成功通知: {account_id}", current_user)
                    else:
                        log_with_user('warning', f"{login_type}成功通知未发送成功: {account_id}", current_user)
                except Exception as notify_err:
                    log_with_user('warning', f"发送登录成功通知失败: {account_id}, 错误: {str(notify_err)}", current_user)

                if is_refresh_mode and session_id in password_login_sessions:
                    screenshot_path = password_login_sessions[session_id].get('screenshot_path')
                    verification_url = password_login_sessions[session_id].get('verification_url')
                    verification_type = password_login_sessions[session_id].get('verification_type')
                    if screenshot_path or verification_url:
                        _set_password_login_session_status(
                            session_id,
                            'success',
                            screenshot_path=screenshot_path,
                            verification_url=verification_url,
                            verification_type=verification_type,
                        )
                
            except Exception as e:
                error_msg = str(e)
                _finalize_password_login_session_failure(session_id, error_msg)
                log_with_user('error', f"账号密码登录失败: {account_id}, 错误: {error_msg}", current_user)
                logger.info(f"会话 {session_id} 状态已更新为 failed，错误消息: {error_msg}")  # 添加日志确认状态更新
                import traceback
                logger.error(traceback.format_exc())
            finally:
                # 清理实例（释放并发槽位）
                try:
                    from utils.xianyu_slider_stealth import concurrency_manager
                    if concurrency_manager.unregister_instance(account_id, slider_instance):
                        log_with_user('debug', f"已释放并发槽位: {account_id}", current_user)
                except Exception as cleanup_e:
                    log_with_user('warning', f"清理实例时出错: {str(cleanup_e)}", current_user)

                if manual_refresh_acquired:
                    try:
                        from XianyuAutoAsync import XianyuLive
                        XianyuLive.end_manual_refresh(account_id, source=manual_refresh_owner)
                        log_with_user('info', f"已结束手动刷新保护: {account_id}", current_user)
                    except Exception as manual_cleanup_e:
                        log_with_user('warning', f"结束手动刷新保护失败: {account_id}, 错误: {str(manual_cleanup_e)}", current_user)

                if auth_recovery_acquired:
                    try:
                        from XianyuAutoAsync import XianyuLive
                        XianyuLive.end_auth_recovery_session(account_id, auth_recovery_owner)
                        log_with_user('info', f"已结束认证恢复单飞锁: {account_id}", current_user)
                    except Exception as auth_cleanup_e:
                        log_with_user('warning', f"结束认证恢复单飞锁失败: {account_id}, 错误: {str(auth_cleanup_e)}", current_user)
        
        # 在后台线程中执行登录
        login_thread = threading.Thread(target=run_login, daemon=True)
        login_thread.start()
        login_thread_started = True
        
    except Exception as e:
        _finalize_password_login_session_failure(session_id, str(e))
        log_with_user('error', f"执行账号密码登录任务异常: {str(e)}", current_user)
        if manual_refresh_acquired and not login_thread_started:
            try:
                from XianyuAutoAsync import XianyuLive
                XianyuLive.end_manual_refresh(account_id, source=manual_refresh_owner)
            except Exception:
                pass
        if auth_recovery_acquired and not login_thread_started:
            try:
                from XianyuAutoAsync import XianyuLive
                XianyuLive.end_auth_recovery_session(account_id, auth_recovery_owner)
            except Exception:
                pass
        import traceback
        logger.error(traceback.format_exc())


async def _execute_manual_cookie_import(
    session_id: str,
    account_id: str,
    cookie_value: str,
    show_browser: bool,
    user_id: int,
    current_user: Dict[str, Any],
):
    try:
        from utils.xianyu_slider_stealth import (
            XianyuSliderStealth,
            probe_cookie_verification_from_cookie,
        )
        from XianyuAutoAsync import XianyuLive

        existing_cookie_info = db_manager.get_cookie_details(account_id) or {}
        proxy_config = {
            'proxy_type': existing_cookie_info.get('proxy_type', 'none'),
            'proxy_host': existing_cookie_info.get('proxy_host', ''),
            'proxy_port': existing_cookie_info.get('proxy_port', 0),
            'proxy_user': existing_cookie_info.get('proxy_user', ''),
            'proxy_pass': existing_cookie_info.get('proxy_pass', ''),
        }
        slider_instance = XianyuSliderStealth(
            user_id=account_id,
            enable_learning=True,
            headless=not show_browser,
            initial_cookies=cookie_value,
            proxy=proxy_config,
        )
        manual_cookie_import_sessions[session_id]['slider_instance'] = slider_instance

        def merge_cookie_dicts_for_import(incoming_cookie_dict: Optional[Dict[str, Any]], source_label: str) -> Dict[str, Any]:
            existing_cookie_dict = trans_cookies(cookie_value)
            merge_result = XianyuLive.protected_merge_cookie_dicts(
                existing_cookie_dict,
                incoming_cookie_dict or {},
            )
            if merge_result['incoming_missing_protected_fields']:
                log_with_user(
                    'warning',
                    (
                        f"导入 Cookie {source_label}快照缺少关键字段，执行保护性合并: "
                        f"{', '.join(merge_result['incoming_missing_protected_fields'])}"
                    ),
                    current_user,
                )
            if merge_result['preserved_protected_fields']:
                log_with_user(
                    'warning',
                    f"导入 Cookie 保护性保留旧字段: {', '.join(merge_result['preserved_protected_fields'])}",
                    current_user,
                )
            return merge_result['merged_cookies_dict']

        def persist_manual_cookie_import_success(merged_cookies_dict: Dict[str, Any], source_label: str):
            if not merged_cookies_dict:
                raise ValueError(f"手动导入 Cookie {source_label}后未获取到有效 Cookie")

            cookies_str = '; '.join([f"{k}={v}" for k, v in merged_cookies_dict.items()])
            existing_same_user_cookie = db_manager.get_all_cookies(user_id)
            is_new_account = account_id not in existing_same_user_cookie
            if is_new_account:
                db_manager.save_cookie(account_id, cookies_str, user_id)
                if cookie_manager.manager:
                    cookie_manager.manager.add_cookie(account_id, cookies_str, user_id=user_id)
            else:
                db_manager.update_cookie_account_info(account_id, cookie_value=cookies_str)
                if cookie_manager.manager:
                    if account_id in getattr(cookie_manager.manager, 'cookies', {}):
                        cookie_manager.manager.update_cookie(account_id, cookies_str, save_to_db=False)
                    else:
                        cookie_manager.manager.add_cookie(account_id, cookies_str, user_id=user_id)

            _set_manual_cookie_import_session_status(
                session_id,
                'success',
                account_id=account_id,
                is_new_account=is_new_account,
                cookie_count=len(merged_cookies_dict),
            )
            log_with_user(
                'info',
                (
                    f"手动导入 Cookie {source_label}成功并已保存: "
                    f"{account_id}, cookie_count={len(merged_cookies_dict)}"
                ),
                current_user,
            )

        def notification_callback(
            message: str,
            screenshot_path: str = None,
            verification_url: str = None,
            screenshot_path_new: str = None,
            verification_type: str = None,
        ):
            """手动导入 Cookie 的验证通知回调。"""
            try:
                import threading

                actual_screenshot_path = screenshot_path_new if screenshot_path_new else screenshot_path
                if actual_screenshot_path and not os.path.exists(actual_screenshot_path):
                    actual_screenshot_path = None

                verification_type_label = resolve_verification_type_label(
                    verification_type,
                    message,
                    verification_url,
                )
                _set_manual_cookie_import_session_status(
                    session_id,
                    'verification_required',
                    verification_url=verification_url or None,
                    screenshot_path=actual_screenshot_path,
                    verification_type=verification_type_label,
                )

                if actual_screenshot_path:
                    log_with_user(
                        'info',
                        f"手动导入 Cookie 验证截图已保存: {session_id}, 路径: {actual_screenshot_path}",
                        current_user,
                    )
                elif verification_url:
                    log_with_user(
                        'info',
                        f"手动导入 Cookie 验证链接已保存: {session_id}, URL: {verification_url}",
                        current_user,
                    )
                else:
                    log_with_user(
                        'warning',
                        f"手动导入 Cookie 检测到{verification_type_label}，但未获取到可用的截图或验证链接: {session_id}",
                        current_user,
                    )

                def send_verification_notification():
                    try:
                        notification_message = build_face_verify_notification(
                            account_id=account_id,
                            time_text=time.strftime('%Y-%m-%d %H:%M:%S'),
                            verification_type=verification_type_label,
                            verification_url=verification_url or '',
                            error_message=message,
                            has_screenshot=bool(actual_screenshot_path),
                        )
                        notification_sent = dispatch_account_notifications_sync(
                            account_id,
                            notification_message,
                            title='闲鱼账号需要验证',
                            notification_type='face_verification',
                            attachment_path=actual_screenshot_path,
                        )
                        if notification_sent:
                            log_with_user('info', f"已发送手动导入 Cookie 验证通知: {account_id}", current_user)
                        else:
                            log_with_user('warning', f"手动导入 Cookie 验证通知未发送成功: {account_id}", current_user)
                    except Exception as notify_err:
                        log_with_user(
                            'warning',
                            f"发送手动导入 Cookie 验证通知失败: {account_id}, 错误: {str(notify_err)}",
                            current_user,
                        )

                notification_thread = threading.Thread(target=send_verification_notification, daemon=True)
                notification_thread.start()
            except Exception as callback_err:
                log_with_user(
                    'warning',
                    f"处理手动导入 Cookie 验证回调失败: {account_id}, 错误: {str(callback_err)}",
                    current_user,
                )

        def run_import():
            try:
                probe_result = probe_cookie_verification_from_cookie(cookie_value, proxy_config)
                if probe_result.get('status') == 'cookie_valid':
                    merged_cookies_dict = merge_cookie_dicts_for_import(
                        probe_result.get('session_cookies'),
                        '预检直通',
                    )
                    log_with_user(
                        'info',
                        f"手动导入 Cookie 预检已确认当前 Cookie 直接有效，跳过浏览器验证: {account_id}",
                        current_user,
                    )
                    persist_manual_cookie_import_success(merged_cookies_dict, '预检直通')
                    return

                target_url = probe_result.get('verification_url')
                if not target_url:
                    raise RuntimeError(
                        f"未拿到最新 verification_url: {probe_result.get('payload') or probe_result}"
                    )
                log_with_user('info', f"手动导入 Cookie 已解析 verification_url: {account_id}", current_user)

                success, cookies_dict = slider_instance.run(
                    target_url,
                    notification_callback=notification_callback,
                    notification_scene='手动导入 Cookie',
                )
                if not success or not cookies_dict:
                    failure_message = slider_instance._get_slider_failure_message('滑块验证失败，请稍后重试')
                    _set_manual_cookie_import_session_status(session_id, 'failed', error=failure_message)
                    log_with_user('error', f"手动导入 Cookie 验证失败: {account_id}, 错误: {failure_message}", current_user)
                    return

                merged_cookies_dict = merge_cookie_dicts_for_import(cookies_dict, '浏览器验证')
                persist_manual_cookie_import_success(merged_cookies_dict, '浏览器验证')
            except Exception as exc:
                error_message = str(exc)
                _set_manual_cookie_import_session_status(session_id, 'failed', error=error_message)
                log_with_user('error', f"手动导入 Cookie 执行异常: {account_id}, 错误: {error_message}", current_user)
                import traceback
                logger.error(traceback.format_exc())
            finally:
                try:
                    from utils.xianyu_slider_stealth import concurrency_manager
                    concurrency_manager.unregister_instance(account_id, slider_instance)
                except Exception:
                    pass

        import threading
        login_thread = threading.Thread(target=run_import, daemon=True)
        login_thread.start()
    except Exception as exc:
        _set_manual_cookie_import_session_status(session_id, 'failed', error=str(exc))
        log_with_user('error', f"执行手动导入 Cookie 任务异常: {str(exc)}", current_user)
        import traceback
        logger.error(traceback.format_exc())


@app.post("/manual-cookie-import")
async def manual_cookie_import(
    request: ManualCookieImportRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """手动导入 Cookie，并按单次调试链路执行真实浏览器滑块验证。"""
    try:
        account_id = str(request.account_id or '').strip()
        cookie_value = str(request.cookie or '').replace('\ufeff', '').strip()
        show_browser = bool(request.show_browser)
        user_id = current_user['user_id']

        if not account_id or not cookie_value:
            return {'success': False, 'message': '账号ID和Cookie不能为空'}

        existing_cookies = db_manager.get_all_cookies()
        if account_id in existing_cookies:
            user_cookies = db_manager.get_all_cookies(user_id)
            if account_id not in user_cookies:
                return {'success': False, 'message': '该账号ID已被其他用户使用'}

        session_id = secrets.token_urlsafe(16)
        manual_cookie_import_sessions[session_id] = {
            'account_id': account_id,
            'show_browser': show_browser,
            'status': 'processing',
            'verification_url': None,
            'screenshot_path': None,
            'verification_type': None,
            'slider_instance': None,
            'task': None,
            'timestamp': time.time(),
            'completed_at': None,
            'user_id': user_id,
        }

        task = asyncio.create_task(_execute_manual_cookie_import(
            session_id,
            account_id,
            cookie_value,
            show_browser,
            user_id,
            current_user,
        ))
        manual_cookie_import_sessions[session_id]['task'] = task

        return {
            'success': True,
            'session_id': session_id,
            'status': 'processing',
            'message': 'Cookie导入验证任务已启动，请等待...',
        }
    except Exception as exc:
        log_with_user('error', f"手动导入 Cookie 异常: {str(exc)}", current_user)
        import traceback
        logger.error(traceback.format_exc())
        return {'success': False, 'message': f'手动导入 Cookie 失败: {str(exc)}'}


@app.get("/manual-cookie-import/check/{session_id}")
async def check_manual_cookie_import_status(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """检查手动导入 Cookie 的执行状态。"""
    try:
        current_time = time.time()
        expired_sessions = [
            sid for sid, session in manual_cookie_import_sessions.items()
            if (
                session.get('completed_at') and current_time - session['completed_at'] > 300
            ) or current_time - session['timestamp'] > 3600
        ]
        for sid in expired_sessions:
            if sid in manual_cookie_import_sessions:
                del manual_cookie_import_sessions[sid]

        if session_id not in manual_cookie_import_sessions:
            return {'status': 'not_found', 'message': '会话不存在或已过期'}

        session = manual_cookie_import_sessions[session_id]
        if session['user_id'] != current_user['user_id']:
            return {'status': 'forbidden', 'message': '无权限访问该会话'}

        status = session['status']
        if status == 'verification_required':
            screenshot_path = session.get('screenshot_path')
            verification_url = session.get('verification_url')
            verification_type = session.get('verification_type') or '身份验证'
            return {
                'status': 'verification_required',
                'verification_url': verification_url,
                'screenshot_path': screenshot_path,
                'verification_type': verification_type,
                'message': f'需要{verification_type}，请查看验证截图' if screenshot_path else f'需要{verification_type}，请点击验证链接',
            }
        if status == 'success':
            return {
                'status': 'success',
                'message': f'账号 {session["account_id"]} Cookie 导入并验证成功',
                'account_id': session['account_id'],
                'is_new_account': session.get('is_new_account', False),
                'cookie_count': session.get('cookie_count', 0),
            }
        if status == 'failed':
            error_msg = session.get('error', 'Cookie 导入验证失败')
            return {
                'status': 'failed',
                'message': error_msg,
                'error': error_msg,
            }
        return {
            'status': 'processing',
            'message': 'Cookie 导入验证处理中，请稍候...',
        }
    except Exception as exc:
        log_with_user('error', f"检查手动导入 Cookie 状态异常: {str(exc)}", current_user)
        return {'status': 'error', 'message': str(exc)}


@app.post("/password-login")
async def password_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """账号密码登录接口（异步，支持人脸认证）"""
    try:
        account_id = request.get('account_id')
        account = request.get('account')
        password = request.get('password')
        # 检查前端是否明确指定了 show_browser 参数
        show_browser_specified = 'show_browser' in request
        show_browser = request.get('show_browser', False)
        refresh_mode = request.get('refresh_mode', False)  # 刷新模式：从数据库读取账密
        risk_log_id = None

        user_id = current_user['user_id']

        # 刷新模式：从数据库读取已保存的账号密码
        if refresh_mode and account_id:
            from XianyuAutoAsync import XianyuLive
            cookie_info = db_manager.get_cookie_details(account_id)
            if not cookie_info:
                return {'success': False, 'message': f'未找到账号: {account_id}'}

            # 验证账号归属
            if cookie_info.get('user_id') != user_id:
                return {'success': False, 'message': '无权操作此账号'}

            account = cookie_info.get('username')
            password = cookie_info.get('password')

            if not account or not password:
                return {'success': False, 'message': '该账号未配置用户名和密码，无法刷新Cookie'}

            # 获取 show_browser 设置（只有当前端没有明确指定时，才使用数据库配置）
            if not show_browser_specified:
                show_browser = cookie_info.get('show_browser', False)

            log_with_user('info', f"刷新Cookie模式: {account_id}, 用户名: {account}, show_browser: {show_browser}", current_user)

            if XianyuLive.is_manual_refresh_active(account_id):
                return {'success': False, 'message': f'账号 {account_id} 正在执行手动刷新，请稍候再试'}

        if not account_id or not account or not password:
            return {'success': False, 'message': '账号ID、登录账号和密码不能为空'}

        log_with_user('info', f"开始账号密码登录: {account_id}, 账号: {account}", current_user)
        
        # 生成会话ID
        session_id = secrets.token_urlsafe(16)
        risk_session_id = _new_risk_log_session_id('pwd')

        # 记录手动刷新Cookie到风控日志
        risk_log_id = None
        if refresh_mode:
            try:
                risk_log_id = db_manager.add_risk_control_log(
                    cookie_id=account_id,
                    event_type='cookie_refresh',
                    session_id=risk_session_id,
                    trigger_scene='manual_password_refresh',
                    result_code='manual_cookie_refresh_started',
                    event_description='手动触发账密Cookie刷新',
                    processing_status='processing',
                    event_meta=_build_risk_event_meta({
                        'account_id': account_id,
                        'show_browser': bool(show_browser),
                        'refresh_mode': True,
                    })
                )
            except Exception as log_e:
                risk_log_id = None
                logger.error(f"记录风控日志失败: {log_e}")
        
        user_id = current_user['user_id']
        
        # 创建登录会话
        password_login_sessions[session_id] = {
            'account_id': account_id,
            'account': account,
            'show_browser': show_browser,
            'refresh_mode': refresh_mode,  # 保存刷新模式标志
            'risk_control_log_id': risk_log_id if refresh_mode else None,  # 风控日志ID
            'risk_session_id': risk_session_id,
            'status': 'processing',
            'verification_url': None,
            'screenshot_path': None,
            'qr_code_url': None,
            'verification_type': None,
            'slider_instance': None,
            'task': None,
            'timestamp': time.time(),
            'completed_at': None,
            'user_id': user_id
        }
        
        # 启动后台登录任务
        task = asyncio.create_task(_execute_password_login(
            session_id, account_id, account, password, show_browser, user_id, current_user
        ))
        password_login_sessions[session_id]['task'] = task
        
        return {
            'success': True,
            'session_id': session_id,
            'status': 'processing',
            'message': '登录任务已启动，请等待...'
        }
        
    except Exception as e:
        log_with_user('error', f"账号密码登录异常: {str(e)}", current_user)
        import traceback
        logger.error(traceback.format_exc())
        return {'success': False, 'message': f'登录失败: {str(e)}'}


@app.get("/password-login/check/{session_id}")
async def check_password_login_status(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """检查账号密码登录状态"""
    try:
        # 清理过期会话（超过1小时）
        current_time = time.time()
        expired_sessions = [
            sid for sid, session in password_login_sessions.items()
            if (
                session.get('completed_at') and current_time - session['completed_at'] > 300
            ) or current_time - session['timestamp'] > 3600
        ]
        for sid in expired_sessions:
            expired_session = password_login_sessions.get(sid)
            if expired_session:
                expired_screenshot_path = expired_session.get('screenshot_path')
                if expired_screenshot_path:
                    try:
                        from utils.image_utils import image_manager
                        if image_manager.delete_image(expired_screenshot_path):
                            log_with_user('info', f"密码登录会话过期，已删除验证截图: {expired_screenshot_path}", current_user)
                        else:
                            log_with_user('warning', f"密码登录会话过期，但删除验证截图失败: {expired_screenshot_path}", current_user)
                    except Exception as cleanup_err:
                        log_with_user('error', f"清理过期密码登录截图时出错: {str(cleanup_err)}", current_user)
            if sid in password_login_sessions:
                del password_login_sessions[sid]
        
        if session_id not in password_login_sessions:
            return {'status': 'not_found', 'message': '会话不存在或已过期'}
        
        session = password_login_sessions[session_id]
        
        # 检查用户权限
        if session['user_id'] != current_user['user_id']:
            return {'status': 'forbidden', 'message': '无权限访问该会话'}
        
        status = session['status']
        
        if status == 'verification_required':
            # 需要身份验证
            screenshot_path = session.get('screenshot_path')
            verification_url = session.get('verification_url')
            verification_type = session.get('verification_type') or '身份验证'
            return {
                'status': 'verification_required',
                'verification_url': verification_url,
                'screenshot_path': screenshot_path,
                'qr_code_url': session.get('qr_code_url'),  # 保留兼容性
                'verification_type': verification_type,
                'message': f'需要{verification_type}，请查看验证截图' if screenshot_path else f'需要{verification_type}，请点击验证链接'
            }
        elif status == 'success':
            return {
                'status': 'success',
                'message': f'账号 {session["account_id"]} 登录成功',
                'account_id': session['account_id'],
                'is_new_account': session.get('is_new_account', False),
                'cookie_count': session.get('cookie_count', 0)
            }
        elif status == 'failed':
            error_msg = session.get('error', '登录失败')
            log_with_user('info', f"返回登录失败状态: {session_id}, 错误消息: {error_msg}", current_user)  # 添加日志
            return {
                'status': 'failed',
                'message': error_msg,
                'error': error_msg  # 也包含error字段，确保前端能获取到
            }
        elif status == 'cancelled':
            return {
                'status': 'cancelled',
                'message': session.get('error') or '登录已取消'
            }
        else:
            # 处理中
            return {
                'status': 'processing',
                'message': '登录处理中，请稍候...'
            }
        
    except Exception as e:
        log_with_user('error', f"检查账号密码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


@app.post("/password-login/cancel/{session_id}")
async def cancel_password_login(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """取消账号密码登录/刷新 Cookie 会话，避免前端反复弹出验证窗口。"""
    try:
        session = password_login_sessions.get(session_id)
        if not session:
            return {'success': False, 'status': 'not_found', 'message': '会话不存在或已过期'}

        if session['user_id'] != current_user['user_id']:
            return {'success': False, 'status': 'forbidden', 'message': '无权限访问该会话'}

        current_status = str(session.get('status') or '').strip().lower()
        if current_status in PASSWORD_LOGIN_TERMINAL_STATUSES:
            return {
                'success': True,
                'status': current_status,
                'message': session.get('error') or '会话已结束'
            }

        _set_password_login_session_status(session_id, 'cancelled', error='用户取消登录')
        _update_session_risk_log(session_id, 'failed', error_message='用户取消登录')
        _close_password_login_pending_verification_risk_logs(
            session_id,
            'failed',
            error_message='用户取消登录',
            result_code='password_login_cancelled',
        )

        slider_instance = session.get('slider_instance')
        if slider_instance:
            try:
                slider_instance.close_browser()
                log_with_user('info', f"已关闭密码登录浏览器实例: {session_id}", current_user)
            except Exception as close_err:
                log_with_user('warning', f"关闭密码登录浏览器实例失败: {session_id}, 错误: {close_err}", current_user)

        return {
            'success': True,
            'status': 'cancelled',
            'message': '登录已取消'
        }
    except Exception as exc:
        log_with_user('error', f"取消账号密码登录异常: {str(exc)}", current_user)
        import traceback
        logger.error(traceback.format_exc())
        return {'success': False, 'status': 'error', 'message': str(exc)}


# ========================= 人脸验证截图相关接口 =========================

@app.get("/face-verification/screenshot/{account_id}")
async def get_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """获取指定账号的人脸验证截图"""
    try:
        import glob
        
        # 检查账号是否属于当前用户
        user_id = current_user['user_id']
        username = current_user['username']
        
        # 如果是管理员，允许访问所有账号
        is_admin = username == 'admin'
        
        if not is_admin:
            cookie_info = db_manager.get_cookie_details(account_id)
            if not cookie_info:
                log_with_user('warning', f"账号 {account_id} 不存在", current_user)
                return {
                    'success': False,
                    'message': '账号不存在'
                }
            
            cookie_user_id = cookie_info.get('user_id')
            if cookie_user_id != user_id:
                log_with_user('warning', f"用户 {user_id} 尝试访问账号 {account_id}（归属用户: {cookie_user_id}）", current_user)
                return {
                    'success': False,
                    'message': '无权访问该账号'
                }

        session_scope_user_id = None if is_admin else user_id
        latest_password_login_session = _get_latest_password_login_session_for_account(
            account_id,
            user_id=session_scope_user_id,
        )
        if latest_password_login_session:
            session_status = str(latest_password_login_session.get('status') or '').strip().lower()
            session_screenshot_path = latest_password_login_session.get('screenshot_path')

            if session_status == 'verification_required' and session_screenshot_path and os.path.exists(session_screenshot_path):
                screenshot_info = _build_face_verification_screenshot_info(account_id, session_screenshot_path)
                log_with_user('info', f"优先返回账号 {account_id} 当前登录会话的验证截图", current_user)
                return {
                    'success': True,
                    'screenshot': screenshot_info
                }

            if session_status == 'failed':
                session_error_message = str(latest_password_login_session.get('error') or '').strip()
                if _is_password_login_verification_timeout_message(session_error_message):
                    log_with_user('info', f"账号 {account_id} 最近一次验证已超时，忽略历史截图", current_user)
                    return {
                        'success': False,
                        'message': session_error_message
                    }

        latest_verification_log = _get_latest_verification_risk_log_for_account(account_id)
        if latest_verification_log and str(latest_verification_log.get('processing_status') or '').strip().lower() == 'failed':
            if _is_timed_out_verification_risk_log(latest_verification_log):
                timeout_message = (
                    str(latest_verification_log.get('error_message') or '').strip()
                    or '当前验证页面已超时/失效，请重新发起验证'
                )
                log_with_user('info', f"账号 {account_id} 最新验证风控已超时，忽略历史截图", current_user)
                return {
                    'success': False,
                    'message': timeout_message
                }
        
        # 获取该账号的验证截图
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern_jpg = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        pattern_png = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.png')
        screenshot_files = glob.glob(pattern_jpg) + glob.glob(pattern_png)
        screenshot_files = [file_path for file_path in screenshot_files if os.path.exists(file_path)]
        
        log_with_user(
            'debug',
            f"查找截图: {pattern_jpg} / {pattern_png}, 找到 {len(screenshot_files)} 个有效文件",
            current_user,
        )
        
        if not screenshot_files:
            log_with_user('warning', f"账号 {account_id} 没有找到验证截图", current_user)
            return {
                'success': False,
                'message': '未找到验证截图'
            }
        
        # 获取最新的截图
        latest_file = max(screenshot_files, key=os.path.getmtime)
        screenshot_info = _build_face_verification_screenshot_info(account_id, latest_file)
        
        log_with_user('info', f"获取账号 {account_id} 的验证截图", current_user)
        
        return {
            'success': True,
            'screenshot': screenshot_info
        }
        
    except Exception as e:
        log_with_user('error', f"获取验证截图失败: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


@app.delete("/face-verification/screenshot/{account_id}")
async def delete_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除指定账号的人脸验证截图"""
    try:
        import glob
        
        # 检查账号是否属于当前用户
        user_id = current_user['user_id']
        cookie_info = db_manager.get_cookie_details(account_id)
        if not cookie_info or cookie_info.get('user_id') != user_id:
            return {
                'success': False,
                'message': '无权访问该账号'
            }
        
        # 删除该账号的所有验证截图
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        screenshot_files = glob.glob(pattern)
        
        deleted_count = 0
        for file_path in screenshot_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_count += 1
                    log_with_user('info', f"删除账号 {account_id} 的验证截图: {os.path.basename(file_path)}", current_user)
            except Exception as e:
                log_with_user('error', f"删除截图失败 {file_path}: {str(e)}", current_user)
        
        return {
            'success': True,
            'message': f'已删除 {deleted_count} 个验证截图',
            'deleted_count': deleted_count
        }
        
    except Exception as e:
        log_with_user('error', f"删除验证截图失败: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


# ========================= 扫码登录相关接口 =========================

@app.post("/qr-login/generate")
async def generate_qr_code(current_user: Dict[str, Any] = Depends(get_current_user)):
    """生成扫码登录二维码"""
    try:
        log_with_user('info', "请求生成扫码登录二维码", current_user)

        result = await qr_login_manager.generate_qr_code()

        if result['success']:
            log_with_user('info', f"扫码登录二维码生成成功: {result['session_id']}", current_user)
        else:
            log_with_user('warning', f"扫码登录二维码生成失败: {result.get('message', '未知错误')}", current_user)

        return result

    except Exception as e:
        log_with_user('error', f"生成扫码登录二维码异常: {str(e)}", current_user)
        return {'success': False, 'message': f'生成二维码失败: {str(e)}'}


@app.get("/qr-login/check/{session_id}")
async def check_qr_code_status(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """检查扫码登录状态"""
    try:
        # 清理过期记录
        cleanup_qr_check_records()

        # 检查是否已经处理过
        if session_id in qr_check_processed:
            record = qr_check_processed[session_id]
            if record['processed']:
                log_with_user('debug', f"扫码登录session {session_id} 已处理过，直接返回", current_user)
                if record.get('error'):
                    return {'status': 'error', 'message': record['error']}

                account_info = record.get('account_info')
                if account_info:
                    return {
                        'status': 'success',
                        'message': '扫码登录已完成',
                        'account_info': account_info,
                        'already_processed': True,
                    }

                return {'status': 'already_processed', 'message': '该会话已处理完成'}

        # 获取该session的锁
        session_lock = qr_check_locks[session_id]

        # 使用非阻塞方式尝试获取锁
        if session_lock.locked():
            log_with_user('debug', f"扫码登录session {session_id} 正在被其他请求处理，跳过", current_user)
            return {'status': 'processing', 'message': '正在处理中，请稍候...'}

        async with session_lock:
            # 再次检查是否已处理（双重检查）
            if session_id in qr_check_processed and qr_check_processed[session_id]['processed']:
                log_with_user('debug', f"扫码登录session {session_id} 在获取锁后发现已处理，直接返回", current_user)
                record = qr_check_processed[session_id]
                if record.get('error'):
                    return {'status': 'error', 'message': record['error']}

                account_info = record.get('account_info')
                if account_info:
                    return {
                        'status': 'success',
                        'message': '扫码登录已完成',
                        'account_info': account_info,
                        'already_processed': True,
                    }

                return {'status': 'already_processed', 'message': '该会话已处理完成'}

            # 清理过期会话
            qr_login_manager.cleanup_expired_sessions()

            # 获取会话状态
            status_info = qr_login_manager.get_session_status(session_id)
            log_with_user('info', f"获取会话状态1111111: {status_info}", current_user)
            if status_info['status'] == 'success':
                log_with_user('info', f"获取会话状态22222222: {status_info}", current_user)

                # 检查是否已经在后台处理中
                if session_id in qr_check_processed and qr_check_processed[session_id].get('processing'):
                    return {'status': 'confirmed', 'message': '已确认，正在获取Cookie...'}

                # 标记为处理中，立即返回"已确认"状态（不阻塞前端）
                qr_check_processed[session_id] = {
                    'processed': False,
                    'processing': True,
                    'timestamp': time.time()
                }

                # 获取 Cookie 信息
                cookies_info = qr_login_manager.get_session_cookies(session_id)
                log_with_user('info', f"获取会话Cookie: {cookies_info}", current_user)

                if cookies_info:
                    # 异步处理 Cookie（不阻塞当前请求）
                    async def _process_cookies_background():
                        try:
                            account_info = await process_qr_login_cookies(
                                cookies_info['cookies'],
                                cookies_info['unb'],
                                current_user
                            )
                            log_with_user('info', f"扫码登录处理完成: {session_id}, 账号: {account_info.get('account_id', 'unknown')}", current_user)
                            qr_check_processed[session_id] = {
                                'processed': True,
                                'processing': False,
                                'timestamp': time.time(),
                                'account_info': account_info
                            }
                        except Exception as bg_e:
                            log_with_user('error', f"后台处理扫码Cookie失败: {bg_e}", current_user)
                            qr_check_processed[session_id] = {
                                'processed': True,
                                'processing': False,
                                'timestamp': time.time(),
                                'error': str(bg_e)
                            }

                    asyncio.create_task(_process_cookies_background())

                # 立即返回"已确认"状态
                return {'status': 'confirmed', 'message': '已确认，正在获取Cookie...'}

            # 检查后台处理是否已完成
            if session_id in qr_check_processed:
                record = qr_check_processed[session_id]
                if record.get('processed') and not record.get('processing'):
                    if record.get('error'):
                        return {'status': 'error', 'message': record['error']}
                    status_info['status'] = 'success'
                    status_info['account_info'] = record.get('account_info', {})
                    return status_info
                elif record.get('processing'):
                    return {'status': 'confirmed', 'message': '已确认，正在获取Cookie...'}

            return status_info

    except Exception as e:
        log_with_user('error', f"检查扫码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


# ========================= 轻量扫码登录(qr_login_lite) =========================

def _cleanup_qr_lite_sessions():
    now = time.time()
    stale = [
        sid for sid, st in qr_lite_sessions.items()
        if st.get('finished') and now - st.get('finished_at', st.get('started_at', now)) > QR_LITE_SESSION_TTL
    ]
    for sid in stale:
        qr_lite_sessions.pop(sid, None)


def _render_qr_data_url(qr_url: str) -> str:
    """把 cv-cat 返回的二维码内容渲染成 data:image/png;base64,..."""
    import qrcode as _qrlib
    img = _qrlib.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode('ascii')


async def _run_qr_login_lite(session_id: str, current_user: Dict[str, Any]):
    state = qr_lite_sessions.get(session_id)
    if state is None:
        return

    def _on_qr_url(qr_url: str):
        try:
            state['qr_data_url'] = _render_qr_data_url(qr_url)
            state['state'] = 'waiting'
        except Exception as render_e:
            state['error_message'] = f'二维码渲染失败: {render_e}'
            state['state'] = 'error'

    def _on_status(raw: str):
        # cv-cat 内部 qrCodeStatus → 前端可识别的 state
        normalized = (raw or '').strip().upper()
        state['raw_qr_status'] = normalized
        if normalized == 'SCANNED':
            state['state'] = 'scanned'
        elif normalized == 'CONFIRMED':
            state['state'] = 'confirmed'
        elif normalized == 'NEW':
            # NEW 与初始 waiting 同义，避免回退覆盖更靠后的 confirmed
            if state.get('state') in (None, 'pending', 'waiting'):
                state['state'] = 'waiting'
        # EXPIRED / 异常字符串：让 qrcode_login_lite 抛 TimeoutError，由 finally 收口

    try:
        cookies, acct = await asyncio.to_thread(
            qrcode_login_lite,
            poll_interval=3.0,
            timeout=180.0,
            show_qrcode_in_terminal=False,
            on_qr_url=_on_qr_url,
            on_status=_on_status,
        )
        cookie_str = '; '.join(f"{k}={v}" for k, v in cookies.items())
        info = await process_qr_login_cookies(cookie_str, acct.get('unb', ''), current_user)
        merged = {**acct, **(info or {})}
        # process_qr_login_cookies 通常会回填 account_id/cookie_length 等
        state['account_info'] = merged
        state['state'] = 'success'
    except TimeoutError as exc:
        state['state'] = 'expired'
        state['error_message'] = str(exc) or '二维码已过期或扫码超时'
        log_with_user('warning', f"轻量扫码登录超时: {exc}", current_user)
    except Exception as exc:
        state['state'] = 'error'
        state['error_message'] = str(exc) or '轻量扫码登录失败'
        log_with_user('error', f"轻量扫码登录异常: {exc}", current_user)
    finally:
        state['finished'] = True
        state['finished_at'] = time.time()


@app.post("/qr-login-lite/generate")
async def generate_qr_code_lite(current_user: Dict[str, Any] = Depends(get_current_user)):
    """生成轻量扫码登录(纯 HTTP)二维码"""
    try:
        log_with_user('info', "请求生成轻量扫码登录二维码", current_user)
        _cleanup_qr_lite_sessions()

        session_id = uuid.uuid4().hex
        qr_lite_sessions[session_id] = {
            'state': 'pending',
            'qr_data_url': None,
            'error_message': None,
            'account_info': None,
            'started_at': time.time(),
            'finished': False,
            'user_id': current_user.get('user_id'),
        }

        asyncio.create_task(_run_qr_login_lite(session_id, current_user))

        # 等 build_initial_cookies + node tfstk + mini_login + generate.do 出二维码
        deadline = time.time() + 30
        while time.time() < deadline:
            st = qr_lite_sessions[session_id]
            if st.get('qr_data_url') or st.get('error_message') or st.get('finished'):
                break
            await asyncio.sleep(0.3)

        st = qr_lite_sessions[session_id]
        if st.get('error_message'):
            return {'success': False, 'message': st['error_message']}
        if not st.get('qr_data_url'):
            return {'success': False, 'message': '生成二维码超时（>30s），可能 node/网络异常'}

        log_with_user('info', f"轻量扫码登录二维码生成成功: {session_id}", current_user)
        return {
            'success': True,
            'session_id': session_id,
            'qr_code_url': st['qr_data_url'],
        }
    except Exception as e:
        log_with_user('error', f"生成轻量扫码登录二维码异常: {str(e)}", current_user)
        return {'success': False, 'message': f'生成二维码失败: {str(e)}'}


@app.get("/qr-login-lite/check/{session_id}")
async def check_qr_code_status_lite(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """检查轻量扫码登录状态"""
    try:
        st = qr_lite_sessions.get(session_id)
        if not st:
            return {'status': 'error', 'message': '会话不存在或已过期'}

        if st.get('user_id') and st['user_id'] != current_user.get('user_id'):
            return {'status': 'error', 'message': '无权访问该会话'}

        state = st.get('state', 'pending')
        if state == 'pending':
            return {'status': 'waiting', 'message': '正在生成二维码…'}
        if state == 'waiting':
            return {'status': 'waiting', 'message': '等待扫码…'}
        if state == 'scanned':
            return {'status': 'scanned', 'message': '已扫码，请在手机上确认…'}
        if state == 'confirmed':
            return {'status': 'confirmed', 'message': '已确认，正在获取Cookie…'}
        if state == 'success':
            return {
                'status': 'success',
                'message': '扫码登录已完成',
                'account_info': st.get('account_info') or {},
            }
        if state == 'expired':
            return {'status': 'expired', 'message': st.get('error_message') or '二维码已过期'}
        # error
        return {'status': 'error', 'message': st.get('error_message') or '扫码登录失败'}
    except Exception as e:
        log_with_user('error', f"检查轻量扫码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


async def process_qr_login_cookies(cookies: str, unb: str, current_user: Dict[str, Any]) -> Dict[str, Any]:
    """处理扫码登录获取的Cookie - 先获取真实cookie再保存到数据库"""
    try:
        user_id = current_user['user_id']

        # 检查是否已存在相同unb的账号
        existing_cookies = db_manager.get_all_cookies(user_id)
        existing_account_id = None
        previous_cookie_value = None

        for account_id, cookie_value in existing_cookies.items():
            try:
                # 解析现有Cookie中的unb
                existing_cookie_dict = trans_cookies(cookie_value)
                if existing_cookie_dict.get('unb') == unb:
                    existing_account_id = account_id
                    previous_cookie_value = cookie_value
                    break
            except:
                continue

        # 确定账号ID
        if existing_account_id:
            account_id = existing_account_id
            is_new_account = False
            log_with_user('info', f"扫码登录找到现有账号: {account_id}, UNB: {unb}", current_user)
        else:
            # 创建新账号，使用unb作为账号ID
            account_id = unb

            # 确保账号ID唯一
            counter = 1
            original_account_id = account_id
            while account_id in existing_cookies:
                account_id = f"{original_account_id}_{counter}"
                counter += 1

            is_new_account = True
            log_with_user('info', f"扫码登录准备创建新账号: {account_id}, UNB: {unb}", current_user)

        # 第一步：使用扫码cookie获取真实cookie
        log_with_user('info', f"开始使用扫码cookie获取真实cookie: {account_id}", current_user)

        # 记录扫码登录到风控日志
        risk_log_id = None
        risk_session_id = _new_risk_log_session_id('qr')
        risk_log_started_at = time.time()
        try:
            risk_log_id = db_manager.add_risk_control_log(
                cookie_id=account_id,
                event_type='cookie_refresh',
                session_id=risk_session_id,
                trigger_scene='qr_login',
                result_code='qr_cookie_refresh_started',
                event_description='扫码登录获取真实Cookie',
                processing_status='processing',
                event_meta=_build_risk_event_meta({
                    'account_id': account_id,
                    'is_new_account': is_new_account,
                })
            )
        except Exception as log_e:
            logger.error(f"记录风控日志失败: {log_e}")

        try:
            # 创建一个临时的XianyuLive实例来执行cookie刷新
            from XianyuAutoAsync import XianyuLive

            # 使用扫码登录的cookie创建临时实例
            temp_instance = XianyuLive(
                cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id,
                register_instance=False,
            )

            # 执行cookie刷新获取真实cookie
            refresh_success = await temp_instance.refresh_cookies_from_qr_login(
                qr_cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id
            )

            if refresh_success:
                log_with_user('info', f"扫码登录真实cookie获取成功: {account_id}", current_user)

                # 从数据库获取刚刚保存的真实cookie
                updated_cookie_info = db_manager.get_cookie_by_id(account_id)
                if updated_cookie_info:
                    real_cookies = updated_cookie_info['cookies_str']
                    log_with_user('info', f"已获取真实cookie，长度: {len(real_cookies)}", current_user)

                    qr_login_grace_minutes = max(5, int(RISK_CONTROL.get('qr_login_grace_minutes', 15) or 15))
                    qr_login_grace_until = int(time.time() + (qr_login_grace_minutes * 60))
                    task_restarted = False
                    warning_message = None
                    final_cookies = temp_instance.cookies_str or real_cookies

                    try:
                        if cookie_manager.manager:
                            if is_new_account:
                                cookie_manager.manager.add_cookie(account_id, final_cookies, user_id=user_id)
                                log_with_user('info', f"已将真实cookie添加到cookie_manager: {account_id}", current_user)
                            else:
                                # refresh_cookies_from_qr_login 已经保存到数据库了，这里不需要再保存
                                cookie_manager.manager.update_cookie(account_id, final_cookies, save_to_db=False)
                                log_with_user('info', f"已更新cookie_manager中的真实cookie: {account_id}", current_user)
                            task_restarted = True
                            db_manager.set_cookie_qr_login_grace_until(account_id, qr_login_grace_until)
                            XianyuLive.mark_qr_login_grace(account_id, stage='real_cookie_ready', grace_until=qr_login_grace_until)
                            # 扫码刚拿到全新可信 cookie，立即清掉旧的密码登录失败退避，
                            # 否则 init() 会被旧的 slider_failed/credentials 退避 skip，
                            # 表现为"扫码完成但 WS 起不来"（详见 22:43 / 22:08 那两次链路）。
                            XianyuLive.clear_password_login_failure_backoff(account_id)
                            log_with_user('info', f"扫码成功后已清除密码登录失败退避: {account_id}", current_user)
                            warning_message = f"真实Cookie已获取，账号任务已切换；为降低再次触发风控的概率，将进入 {qr_login_grace_minutes} 分钟稳定期，稳定期内不自动预热Token"
                            log_with_user('warning', f"{warning_message}: {account_id}", current_user)
                        else:
                            warning_message = "真实Cookie已获取，但任务管理器未初始化，未启动账号任务"
                            log_with_user('warning', f"{warning_message}: {account_id}", current_user)
                    except Exception as task_switch_e:
                        db_manager.set_cookie_qr_login_grace_until(account_id, 0)
                        XianyuLive.clear_qr_login_grace(account_id)
                        warning_message = f"真实Cookie已获取，但切换账号任务失败: {str(task_switch_e)}"
                        log_with_user('warning', f"{warning_message}: {account_id}", current_user)

                    if not task_restarted:
                        db_manager.set_cookie_qr_login_grace_until(account_id, 0)
                        XianyuLive.clear_qr_login_grace(account_id)
                        if not warning_message:
                            warning_message = "真实Cookie已获取，但任务管理器未初始化，未启动账号任务"
                            log_with_user('warning', f"{warning_message}: {account_id}", current_user)
                        if is_new_account:
                            db_manager.delete_cookie(account_id)
                            log_with_user('warning', f"扫码登录未完成切换，已删除临时创建的新账号记录: {account_id}", current_user)
                        elif previous_cookie_value:
                            db_manager.update_cookie_account_info(account_id, cookie_value=previous_cookie_value)
                            log_with_user('warning', f"扫码登录未完成切换，已回滚现有账号Cookie: {account_id}", current_user)
                        else:
                            log_with_user('warning', f"扫码登录未完成切换，但未找到可回滚的旧Cookie: {account_id}", current_user)

                    # 更新风控日志状态
                    if risk_log_id:
                        try:
                            if task_restarted:
                                processing_result = '扫码登录真实Cookie获取成功，账号任务已启动'
                                processing_result += f'；已进入 {qr_login_grace_minutes} 分钟稳定期，稳定期内不自动预热Token'
                                db_manager.update_risk_control_log(
                                    log_id=risk_log_id,
                                    processing_status='success',
                                    processing_result=processing_result,
                                    session_id=risk_session_id,
                                    trigger_scene='qr_login',
                                    result_code='qr_cookie_refresh_success',
                                    duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                                    event_meta=_build_risk_event_meta({
                                        'account_id': account_id,
                                        'is_new_account': is_new_account,
                                        'task_restarted': task_restarted,
                                        'token_prewarmed': False,
                                    })
                                )
                            else:
                                db_manager.update_risk_control_log(
                                    log_id=risk_log_id,
                                    processing_status='failed',
                                    error_message=(warning_message or '账号任务未启动')[:200],
                                    processing_result='扫码登录真实Cookie获取成功，但未切换到新任务',
                                    session_id=risk_session_id,
                                    trigger_scene='qr_login',
                                    result_code='qr_cookie_task_not_started',
                                    duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                                    event_meta=_build_risk_event_meta({
                                        'account_id': account_id,
                                        'is_new_account': is_new_account,
                                        'task_restarted': task_restarted,
                                        'token_prewarmed': False,
                                    })
                                )
                        except Exception:
                            pass

                    return {
                        'account_id': account_id,
                        'is_new_account': is_new_account,
                        'real_cookie_refreshed': task_restarted,  # 回滚时为 False，成功切换时为 True
                        'cookie_length': len(final_cookies),
                        'token_prewarmed': False,
                        'task_restarted': task_restarted,
                        'warning_message': warning_message
                    }
                else:
                    log_with_user('error', f"无法从数据库获取真实cookie: {account_id}", current_user)
                    if risk_log_id:
                        try:
                            db_manager.update_risk_control_log(
                                log_id=risk_log_id,
                                processing_status='failed',
                                error_message='无法从数据库获取真实cookie',
                                session_id=risk_session_id,
                                trigger_scene='qr_login',
                                result_code='qr_cookie_missing_after_refresh',
                                duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                                event_meta=_build_risk_event_meta({'account_id': account_id, 'is_new_account': is_new_account})
                            )
                        except Exception:
                            pass
                    # 降级处理：使用原始扫码cookie
                    return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "无法从数据库获取真实cookie")
            else:
                log_with_user('warning', f"扫码登录真实cookie获取失败: {account_id}", current_user)
                if risk_log_id:
                    try:
                        db_manager.update_risk_control_log(
                            log_id=risk_log_id,
                            processing_status='failed',
                            error_message='真实cookie获取失败',
                            session_id=risk_session_id,
                            trigger_scene='qr_login',
                            result_code='qr_cookie_refresh_failed',
                            duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                            event_meta=_build_risk_event_meta({'account_id': account_id, 'is_new_account': is_new_account})
                        )
                    except Exception:
                        pass
                # 降级处理：使用原始扫码cookie
                return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "真实cookie获取失败")

        except Exception as refresh_e:
            log_with_user('error', f"扫码登录真实cookie获取异常: {str(refresh_e)}", current_user)
            if risk_log_id:
                try:
                    db_manager.update_risk_control_log(
                        log_id=risk_log_id,
                        processing_status='failed',
                        error_message=str(refresh_e)[:200],
                        session_id=risk_session_id,
                        trigger_scene='qr_login',
                        result_code='qr_cookie_refresh_exception',
                        duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                        event_meta=_build_risk_event_meta({'account_id': account_id, 'is_new_account': is_new_account})
                    )
                except Exception:
                    pass
            # 降级处理：使用原始扫码cookie
            return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, f"获取真实cookie异常: {str(refresh_e)}")

    except Exception as e:
        log_with_user('error', f"处理扫码登录Cookie失败: {str(e)}", current_user)
        raise e


async def _fallback_save_qr_cookie(account_id: str, cookies: str, user_id: int, is_new_account: bool, current_user: Dict[str, Any], error_reason: str) -> Dict[str, Any]:
    """降级处理：当无法获取真实cookie时，保存原始扫码cookie"""
    try:
        log_with_user('warning', f"降级处理 - 保存原始扫码cookie: {account_id}, 原因: {error_reason}", current_user)

        # 保存原始扫码cookie到数据库
        if is_new_account:
            db_manager.save_cookie(account_id, cookies, user_id)
            log_with_user('info', f"降级处理 - 新账号原始cookie已保存: {account_id}", current_user)
        else:
            # 现有账号使用 update_cookie_account_info 避免覆盖其他字段
            db_manager.update_cookie_account_info(account_id, cookie_value=cookies)
            log_with_user('info', f"降级处理 - 现有账号原始cookie已更新: {account_id}", current_user)

        # 添加到或更新cookie_manager
        if cookie_manager.manager:
            if is_new_account:
                cookie_manager.manager.add_cookie(account_id, cookies)
                log_with_user('info', f"降级处理 - 已将原始cookie添加到cookie_manager: {account_id}", current_user)
            else:
                # update_cookie_account_info 已经保存到数据库了，这里不需要再保存
                cookie_manager.manager.update_cookie(account_id, cookies, save_to_db=False)
                log_with_user('info', f"降级处理 - 已更新cookie_manager中的原始cookie: {account_id}", current_user)

        return {
            'account_id': account_id,
            'is_new_account': is_new_account,
            'real_cookie_refreshed': False,
            'fallback_reason': error_reason,
            'cookie_length': len(cookies)
        }

    except Exception as fallback_e:
        log_with_user('error', f"降级处理失败: {str(fallback_e)}", current_user)
        raise fallback_e


@app.post("/qr-login/refresh-cookies")
async def refresh_cookies_from_qr_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """使用扫码登录获取的cookie访问指定界面获取真实cookie并存入数据库"""
    try:
        qr_cookies = request.get('qr_cookies')
        cookie_id = request.get('cookie_id')

        if not qr_cookies:
            return {'success': False, 'message': '缺少扫码登录cookie'}

        if not cookie_id:
            return {'success': False, 'message': '缺少cookie_id'}

        log_with_user('info', f"开始使用扫码cookie刷新真实cookie: {cookie_id}", current_user)

        # 记录扫码刷新Cookie到风控日志
        risk_log_id = None
        risk_session_id = _new_risk_log_session_id('qrrefresh')
        risk_log_started_at = time.time()
        try:
            risk_log_id = db_manager.add_risk_control_log(
                cookie_id=cookie_id,
                event_type='cookie_refresh',
                session_id=risk_session_id,
                trigger_scene='manual_qr_refresh',
                result_code='manual_qr_refresh_started',
                event_description='手动触发扫码Cookie刷新',
                processing_status='processing',
                event_meta=_build_risk_event_meta({'account_id': cookie_id})
            )
        except Exception as log_e:
            logger.error(f"记录风控日志失败: {log_e}")

        # 创建一个临时的XianyuLive实例来执行cookie刷新
        from XianyuAutoAsync import XianyuLive

        # 使用扫码登录的cookie创建临时实例
        temp_instance = XianyuLive(
            cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id'],
            register_instance=False,
        )

        # 执行cookie刷新
        success = await temp_instance.refresh_cookies_from_qr_login(
            qr_cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id']
        )

        if success:
            log_with_user('info', f"扫码cookie刷新成功: {cookie_id}", current_user)

            # 更新风控日志状态
            if risk_log_id:
                try:
                    db_manager.update_risk_control_log(
                        log_id=risk_log_id,
                        processing_status='success',
                        processing_result='扫码Cookie刷新成功',
                        session_id=risk_session_id,
                        trigger_scene='manual_qr_refresh',
                        result_code='manual_qr_refresh_success',
                        duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                        event_meta=_build_risk_event_meta({'account_id': cookie_id})
                    )
                except Exception:
                    pass

            # 如果cookie_manager存在，更新其中的cookie
            if cookie_manager.manager:
                # 从数据库获取更新后的cookie
                updated_cookie_info = db_manager.get_cookie_by_id(cookie_id)
                if updated_cookie_info:
                    # refresh_cookies_from_qr_login 已经保存到数据库了，这里不需要再保存
                    cookie_manager.manager.update_cookie(cookie_id, updated_cookie_info['cookies_str'], save_to_db=False)
                    log_with_user('info', f"已更新cookie_manager中的cookie: {cookie_id}", current_user)

            return {
                'success': True,
                'message': '真实cookie获取并保存成功',
                'cookie_id': cookie_id
            }
        else:
            log_with_user('error', f"扫码cookie刷新失败: {cookie_id}", current_user)
            # 更新风控日志状态
            if risk_log_id:
                try:
                    db_manager.update_risk_control_log(
                        log_id=risk_log_id,
                        processing_status='failed',
                        error_message='获取真实cookie失败',
                        session_id=risk_session_id,
                        trigger_scene='manual_qr_refresh',
                        result_code='manual_qr_refresh_failed',
                        duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                        event_meta=_build_risk_event_meta({'account_id': cookie_id})
                    )
                except Exception:
                    pass
            return {'success': False, 'message': '获取真实cookie失败'}

    except Exception as e:
        log_with_user('error', f"扫码cookie刷新异常: {str(e)}", current_user)
        # 更新风控日志状态
        if risk_log_id:
            try:
                db_manager.update_risk_control_log(
                    log_id=risk_log_id,
                    processing_status='failed',
                    error_message=str(e)[:200],
                    session_id=risk_session_id,
                    trigger_scene='manual_qr_refresh',
                    result_code='manual_qr_refresh_exception',
                    duration_ms=max(0, int((time.time() - risk_log_started_at) * 1000)),
                    event_meta=_build_risk_event_meta({'account_id': cookie_id})
                )
            except Exception:
                pass
        return {'success': False, 'message': f'刷新cookie失败: {str(e)}'}


@app.post("/qr-login/reset-cooldown/{cookie_id}")
async def reset_qr_cookie_refresh_cooldown(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """重置指定账号的扫码登录Cookie刷新冷却时间"""
    try:
        log_with_user('info', f"重置扫码登录Cookie刷新冷却时间: {cookie_id}", current_user)

        # 检查cookie是否存在
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': '账号不存在'}

        # 如果cookie_manager中有对应的实例，直接重置
        instance = cookie_manager.manager.get_xianyu_instance(cookie_id) if cookie_manager.manager else None
        if instance:
            remaining_time_before = instance.get_qr_cookie_refresh_remaining_time()
            instance.reset_qr_cookie_refresh_flag()

            log_with_user('info', f"已重置账号 {cookie_id} 的扫码登录冷却时间，原剩余时间: {remaining_time_before}秒", current_user)

            return {
                'success': True,
                'message': '扫码登录Cookie刷新冷却时间已重置',
                'cookie_id': cookie_id,
                'previous_remaining_time': remaining_time_before
            }
        else:
            # 如果没有活跃实例，返回成功（因为没有冷却时间需要重置）
            log_with_user('info', f"账号 {cookie_id} 没有活跃实例，无需重置冷却时间", current_user)
            return {
                'success': True,
                'message': '账号没有活跃实例，无需重置冷却时间',
                'cookie_id': cookie_id
            }

    except Exception as e:
        log_with_user('error', f"重置扫码登录冷却时间异常: {str(e)}", current_user)
        return {'success': False, 'message': f'重置冷却时间失败: {str(e)}'}


@app.get("/qr-login/cooldown-status/{cookie_id}")
async def get_qr_cookie_refresh_cooldown_status(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """获取指定账号的扫码登录Cookie刷新冷却状态"""
    try:
        # 检查cookie是否存在
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': '账号不存在'}

        # 如果cookie_manager中有对应的实例，获取冷却状态
        instance = cookie_manager.manager.get_xianyu_instance(cookie_id) if cookie_manager.manager else None
        if instance:
            remaining_time = instance.get_qr_cookie_refresh_remaining_time()
            cooldown_duration = instance.qr_cookie_refresh_cooldown
            last_refresh_time = instance.last_qr_cookie_refresh_time

            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': remaining_time,
                'cooldown_duration': cooldown_duration,
                'last_refresh_time': last_refresh_time,
                'is_in_cooldown': remaining_time > 0,
                'remaining_minutes': remaining_time // 60,
                'remaining_seconds': remaining_time % 60
            }
        else:
            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': 0,
                'cooldown_duration': 600,  # 默认10分钟
                'last_refresh_time': 0,
                'is_in_cooldown': False,
                'message': '账号没有活跃实例'
            }

    except Exception as e:
        log_with_user('error', f"获取扫码登录冷却状态异常: {str(e)}", current_user)
        return {'success': False, 'message': f'获取冷却状态失败: {str(e)}'}


@app.put('/cookies/{cid}/status')
def update_cookie_status(cid: str, status_data: CookieStatusIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的启用/禁用状态"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.update_cookie_status(cid, status_data.enabled)
        status_note = ''
        if status_data.enabled:
            db_manager.update_cookie_status_note(cid, '')
        else:
            cookie_details = db_manager.get_cookie_details(cid)
            status_note = cookie_details.get('status_note', '') if cookie_details else ''
        return {'msg': 'status updated', 'enabled': status_data.enabled, 'status_note': status_note}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------- 默认回复管理接口 -------------------------

@app.get('/default-replies/{cid}')
def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        result = db_manager.get_default_reply(cid)
        if result is None:
            # 如果没有设置，返回默认值
            return {'enabled': False, 'reply_content': '', 'reply_once': False}
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/default-replies/{cid}')
def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once)
        return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/default-replies')
def get_all_default_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的默认回复设置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_replies = db_manager.get_all_default_replies()
        # 过滤只属于当前用户的回复设置
        user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
        return user_replies
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/default-replies/{cid}')
def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_default_reply(cid)
        if success:
            return {'msg': 'default reply deleted'}
        else:
            raise HTTPException(status_code=400, detail='删除失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/default-replies/{cid}/clear-records')
def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.clear_default_reply_records(cid)
        return {'msg': 'default reply records cleared'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 通知渠道管理接口 -------------------------

@app.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get('/notification-channels/{channel_id}')
def get_notification_channel(channel_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel = db_manager.get_notification_channel(channel_id, user_id=user_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/notification-channels/{channel_id}')
def update_notification_channel(channel_id: int, channel_data: NotificationChannelUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled,
            user_id=user_id
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete('/notification-channels/{channel_id}')
def delete_notification_channel(channel_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        success = db_manager.delete_notification_channel(channel_id, user_id=user_id)
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 消息通知配置接口 -------------------------

@app.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的消息通知配置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # 过滤只属于当前用户的通知配置
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """设置账号的消息通知"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查通知渠道是否存在
        channel = db_manager.get_notification_channel(notification_data.channel_id, user_id=user_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')

        success = db_manager.set_message_notification(cid, notification_data.channel_id, notification_data.enabled)
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='设置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/account/{cid}')
def delete_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除账号的所有消息通知配置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_account_notifications(cid, user_id=user_id)
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='账号通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/{notification_id}')
def delete_message_notification(notification_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除消息通知配置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        success = db_manager.delete_message_notification(notification_id, user_id=user_id)
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 通知模板接口 -------------------------

@app.get('/notification-templates')
def get_notification_templates(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知模板"""
    from db_manager import db_manager
    try:
        templates = db_manager.get_all_notification_templates()
        return {'templates': templates}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class TestNotificationIn(BaseModel):
    template_type: str
    template: str


@app.post('/notification-templates/test')
async def test_notification_template(data: TestNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """发送测试通知"""
    import time as time_module
    import aiohttp
    from db_manager import db_manager

    try:
        if data.template_type not in SUPPORTED_NOTIFICATION_TEMPLATE_TYPES:
            raise HTTPException(status_code=400, detail='无效的模板类型')

        # 获取所有已启用的通知渠道
        channels = db_manager.get_notification_channels(current_user['user_id'])
        logger.info(f"获取到的通知渠道: {channels}")
        enabled_channels = [c for c in channels if c.get('enabled', False)]
        logger.info(f"已启用的通知渠道: {enabled_channels}")

        if not enabled_channels:
            raise HTTPException(status_code=400, detail='没有已启用的通知渠道，请先在「通知渠道」页面配置')

        # 准备测试数据
        test_data = {
            'message': {
                'account_id': '测试账号',
                'buyer_name': '测试买家',
                'buyer_id': '123456789',
                'item_id': '987654321',
                'chat_id': 'test_chat_001',
                'message': '这是一条测试消息',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S')
            },
            'token_refresh': {
                'account_id': '测试账号',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'error_message': '这是一条测试异常信息',
                'verification_url': 'https://example.com/verify'
            },
            'delivery': {
                'account_id': '测试账号',
                'buyer_name': '测试买家',
                'buyer_id': '234567890',
                'item_id': '876543210',
                'chat_id': 'test_chat_002',
                'result': '测试发货成功',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S')
            },
            'slider_success': {
                'account_id': '测试账号',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'status_text': 'cookies已自动更新到数据库'
            },
            'face_verify': {
                'account_id': '测试账号',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'verification_action': '请点击验证链接完成验证:',
                'verification_url': 'https://passport.goofish.com/mini_login.htm?example=test',
                'verification_type': '身份验证'
            },
            'password_login_success': {
                'account_id': '测试账号',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'cookie_count': '30'
            },
            'cookie_refresh_success': {
                'account_id': '测试账号',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'cookie_count': '30'
            },
            'account_paused': {
                'account_id': '测试账号',
                'status_note': '待二维码验证',
                'pause_reason': '二维码验证',
                'time': time_module.strftime('%Y-%m-%d %H:%M:%S'),
                'error_message': '检测到需要人工完成的二维码验证',
                'verification_url': 'https://passport.goofish.com/mini_login.htm?example=test',
                'action_hint': '请先完成验证，再恢复账号运行。'
            }
        }

        # 格式化模板
        template = data.template
        for key, value in test_data.get(data.template_type, {}).items():
            template = template.replace(f'{{{key}}}', str(value))

        # 发送测试通知到所有已启用的渠道
        success_channels = []
        failed_channels = []

        for channel in enabled_channels:
            channel_type = channel.get('type', '')
            channel_name = channel.get('name', channel_type)
            config_str = channel.get('config', '{}')
            logger.info(f"处理通知渠道: name={channel_name}, type={channel_type}, config={config_str}")

            try:
                import json
                config_data = json.loads(config_str) if isinstance(config_str, str) else config_str
                logger.info(f"解析后的配置: {config_data}")

                # 根据渠道类型发送通知
                if channel_type == 'feishu' or channel_type == 'lark':
                    webhook_url = config_data.get('webhook_url', '')
                    secret = config_data.get('secret', '')
                    logger.info(f"飞书渠道配置: webhook_url={webhook_url}, has_secret={bool(secret)}")
                    if webhook_url:
                        import hmac
                        import hashlib
                        import base64

                        # 生成签名（按照实际发送逻辑）
                        timestamp = str(int(time_module.time()))
                        sign = ""

                        if secret:
                            string_to_sign = f'{timestamp}\n{secret}'
                            hmac_code = hmac.new(
                                string_to_sign.encode('utf-8'),
                                ''.encode('utf-8'),
                                digestmod=hashlib.sha256
                            ).digest()
                            sign = base64.b64encode(hmac_code).decode('utf-8')
                            logger.info(f"飞书签名: timestamp={timestamp}")

                        # 构建请求数据
                        payload = {
                            "msg_type": "text",
                            "content": {
                                "text": f"【测试通知】\n\n{template}"
                            },
                            "timestamp": timestamp
                        }

                        if sign:
                            payload["sign"] = sign

                        logger.info(f"发送飞书通知: {payload}")
                        timeout = aiohttp.ClientTimeout(total=10)
                        async with aiohttp.ClientSession(timeout=timeout) as session:
                            async with session.post(webhook_url, json=payload) as resp:
                                resp_text = await resp.text()
                                logger.info(f"飞书响应: status={resp.status}, body={resp_text}")
                                if resp.status == 200:
                                    try:
                                        resp_json = json.loads(resp_text)
                                        if resp_json.get('code', 0) == 0:
                                            success_channels.append(channel_name)
                                        else:
                                            failed_channels.append(f"{channel_name} ({resp_json.get('msg', resp_text[:50])})")
                                    except:
                                        success_channels.append(channel_name)
                                else:
                                    failed_channels.append(f"{channel_name} (HTTP {resp.status}: {resp_text[:50]})")
                    else:
                        failed_channels.append(f"{channel_name} (未配置webhook_url)")

                elif channel_type == 'dingtalk' or channel_type == 'ding_talk':
                    webhook_url = config_data.get('webhook_url', '')
                    if webhook_url:
                        payload = {
                            "msgtype": "text",
                            "text": {
                                "content": f"【测试通知】\n\n{template}"
                            }
                        }
                        timeout = aiohttp.ClientTimeout(total=10)
                        async with aiohttp.ClientSession(timeout=timeout) as session:
                            async with session.post(webhook_url, json=payload) as resp:
                                resp_text = await resp.text()
                                logger.info(f"钉钉响应: status={resp.status}, body={resp_text}")
                                if resp.status == 200:
                                    success_channels.append(channel_name)
                                else:
                                    failed_channels.append(f"{channel_name} (HTTP {resp.status})")

                elif channel_type == 'bark':
                    server_url = config_data.get('server_url', 'https://api.day.app')
                    device_key = config_data.get('device_key', '')
                    if device_key:
                        import urllib.parse
                        encoded_template = urllib.parse.quote(template)
                        url = f"{server_url}/{device_key}/测试通知/{encoded_template}"
                        timeout = aiohttp.ClientTimeout(total=10)
                        async with aiohttp.ClientSession(timeout=timeout) as session:
                            async with session.get(url) as resp:
                                if resp.status == 200:
                                    success_channels.append(channel_name)
                                else:
                                    failed_channels.append(f"{channel_name} (HTTP {resp.status})")

                elif channel_type == 'telegram':
                    bot_token = config_data.get('bot_token', '')
                    chat_id = config_data.get('chat_id', '')
                    if bot_token and chat_id:
                        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                        payload = {
                            "chat_id": chat_id,
                            "text": f"【测试通知】\n\n{template}"
                        }
                        timeout = aiohttp.ClientTimeout(total=10)
                        async with aiohttp.ClientSession(timeout=timeout) as session:
                            async with session.post(url, json=payload) as resp:
                                if resp.status == 200:
                                    success_channels.append(channel_name)
                                else:
                                    failed_channels.append(f"{channel_name} (HTTP {resp.status})")

                elif channel_type == 'webhook':
                    webhook_url = config_data.get('webhook_url', '')
                    if webhook_url:
                        payload = {
                            "title": "测试通知",
                            "content": template,
                            "type": data.template_type
                        }
                        timeout = aiohttp.ClientTimeout(total=10)
                        async with aiohttp.ClientSession(timeout=timeout) as session:
                            async with session.post(webhook_url, json=payload) as resp:
                                if resp.status == 200:
                                    success_channels.append(channel_name)
                                else:
                                    failed_channels.append(f"{channel_name} (HTTP {resp.status})")

                elif channel_type == 'email':
                    failed_channels.append(f"{channel_name} (邮件测试暂不支持)")

                else:
                    failed_channels.append(f"{channel_name} (不支持的渠道类型)")

            except Exception as e:
                logger.error(f"渠道 {channel_name} 发送失败: {e}")
                import traceback
                logger.error(traceback.format_exc())
                failed_channels.append(f"{channel_name} ({str(e)})")

        # 返回结果
        if success_channels:
            return {
                'success': True,
                'message': f'测试通知发送成功: {", ".join(success_channels)}',
                'success_channels': success_channels,
                'failed_channels': failed_channels
            }
        else:
            raise HTTPException(
                status_code=400,
                detail=f'所有渠道发送失败: {", ".join(failed_channels)}'
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/notification-templates/{template_type}')
def get_notification_template(template_type: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定类型的通知模板"""
    from db_manager import db_manager
    try:
        if template_type not in SUPPORTED_NOTIFICATION_TEMPLATE_TYPES:
            raise HTTPException(status_code=400, detail='无效的模板类型')

        template = db_manager.get_notification_template(template_type)
        if template:
            return template
        else:
            # 返回默认模板
            default_template = db_manager.get_default_notification_template(template_type)
            return {
                'type': template_type,
                'template': default_template,
                'is_default': True
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class NotificationTemplateIn(BaseModel):
    template: str


@app.put('/notification-templates/{template_type}')
def update_notification_template(template_type: str, data: NotificationTemplateIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新通知模板"""
    from db_manager import db_manager
    try:
        if template_type not in SUPPORTED_NOTIFICATION_TEMPLATE_TYPES:
            raise HTTPException(status_code=400, detail='无效的模板类型')

        # 如果模板不存在，先插入默认值
        existing = db_manager.get_notification_template(template_type)
        if not existing:
            cursor = db_manager.conn.cursor()
            default_template = db_manager.get_default_notification_template(template_type)
            cursor.execute(
                'INSERT INTO notification_templates (type, template) VALUES (?, ?)',
                (template_type, default_template)
            )
            db_manager.conn.commit()

        success = db_manager.update_notification_template(template_type, data.template)
        if success:
            return {'msg': 'notification template updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/notification-templates/{template_type}/reset')
def reset_notification_template(template_type: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """重置通知模板为默认值"""
    from db_manager import db_manager
    try:
        if template_type not in SUPPORTED_NOTIFICATION_TEMPLATE_TYPES:
            raise HTTPException(status_code=400, detail='无效的模板类型')

        success = db_manager.reset_notification_template(template_type)
        if success:
            # 返回重置后的模板
            template = db_manager.get_notification_template(template_type)
            return {'msg': 'notification template reset', 'template': template}
        else:
            raise HTTPException(status_code=400, detail='重置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/notification-templates/{template_type}/default')
def get_default_notification_template(template_type: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取默认通知模板"""
    from db_manager import db_manager
    try:
        if template_type not in SUPPORTED_NOTIFICATION_TEMPLATE_TYPES:
            raise HTTPException(status_code=400, detail='无效的模板类型')

        default_template = db_manager.get_default_notification_template(template_type)
        if default_template:
            return {'type': template_type, 'template': default_template}
        else:
            raise HTTPException(status_code=404, detail='模板不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统设置接口 -------------------------

@app.get('/system-settings')
def get_system_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取系统设置（排除敏感信息）"""
    from db_manager import db_manager
    try:
        settings = db_manager.get_all_system_settings()
        # 移除敏感信息
        if 'admin_password_hash' in settings:
            del settings['admin_password_hash']
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@app.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新系统设置"""
    from db_manager import db_manager
    try:
        # 禁止直接修改密码哈希
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='请使用密码修改接口')

        value = _validate_system_setting_value(key, setting_data.value)

        if key in NIGHT_MODE_SYSTEM_SETTING_KEYS and not current_user.get('is_admin'):
            raise HTTPException(status_code=403, detail='仅管理员可修改夜间风控降频设置')

        success = db_manager.set_system_setting(key, value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 注册设置接口 -------------------------

@app.get('/registration-status')
def get_registration_status():
    """获取注册开关状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('registration_enabled')
        logger.info(f"从数据库获取的注册设置值: '{enabled_str}'")  # 调试信息

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
            message = '注册功能已开启'
        else:
            enabled_bool = enabled_str == 'true'
            message = '注册功能已开启' if enabled_bool else '注册功能已关闭'

        logger.info(f"解析后的注册状态: enabled={enabled_bool}, message='{message}'")  # 调试信息

        return {
            'enabled': enabled_bool,
            'message': message
        }
    except Exception as e:
        logger.error(f"获取注册状态失败: {e}")
        return {'enabled': True, 'message': '注册功能已开启'}  # 出错时默认开启


@app.get('/login-info-status')
def get_login_info_status():
    """获取默认登录信息显示状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('show_default_login_info')
        logger.debug(f"从数据库获取的登录信息显示设置值: '{enabled_str}'")

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录信息显示状态失败: {e}")
        # 出错时默认为开启
        return {"enabled": True}


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


class LoginInfoSettingUpdate(BaseModel):
    enabled: bool


@app.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新注册开关设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'registration_enabled',
            'true' if enabled else 'false',
            '是否开启用户注册'
        )
        if success:
            log_with_user('info', f"更新注册设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"注册功能已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新注册设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新注册设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/login-info-settings')
def update_login_info_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新默认登录信息显示设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'show_default_login_info',
            'true' if enabled else 'false',
            '是否显示默认登录信息'
        )
        if success:
            log_with_user('info', f"更新登录信息显示设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"默认登录信息显示已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新登录信息显示设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新登录信息显示设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/login-captcha-settings')
def get_login_captcha_settings(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取登录验证码设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('login_captcha_enabled')
        logger.debug(f"从数据库获取的登录验证码设置值: '{enabled_str}'")

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录验证码设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/login-captcha-settings')
def update_login_captcha_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新登录验证码设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'login_captcha_enabled',
            'true' if enabled else 'false',
            '是否开启登录验证码'
        )
        if success:
            log_with_user('info', f"更新登录验证码设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"登录验证码已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新登录验证码设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新登录验证码设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 公开接口：获取登录验证码是否启用（供登录页面使用）
@app.get('/api/login-captcha-enabled')
def get_login_captcha_enabled():
    """获取登录验证码是否启用（公开接口，供登录页面判断）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('login_captcha_enabled')
        enabled_bool = enabled_str == 'true' if enabled_str is not None else True
        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录验证码设置失败: {e}")
        return {"enabled": True}  # 出错时默认开启验证码




@app.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class AutoCommentUpdate(BaseModel):
    auto_comment: bool


class CommentTemplateCreate(BaseModel):
    name: str
    content: str
    is_active: Optional[bool] = False


class CommentTemplateUpdate(BaseModel):
    name: Optional[str] = None
    content: Optional[str] = None
    is_active: Optional[bool] = None


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@app.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_confirm设置
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动确认发货设置失败")

        # 通知CookieManager更新设置（如果账号正在运行）
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"自动确认发货已{'开启' if update_data.auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_confirm设置
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"自动确认发货当前{'开启' if auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 自动好评相关API ====================

@app.put("/cookies/{cid}/auto-comment")
def update_auto_comment(cid: str, update_data: AutoCommentUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动好评设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_comment设置
        success = db_manager.update_auto_comment(cid, update_data.auto_comment)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动好评设置失败")

        return {
            "msg": "success",
            "auto_comment": update_data.auto_comment,
            "message": f"自动好评已{'开启' if update_data.auto_comment else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/auto-comment")
def get_auto_comment(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动好评设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_comment设置
        auto_comment = db_manager.get_auto_comment(cid)
        return {
            "auto_comment": auto_comment,
            "message": f"自动好评当前{'开启' if auto_comment else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/comment-templates")
def get_comment_templates(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的好评模板列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        templates = db_manager.get_comment_templates(cid)
        return {
            "templates": templates,
            "message": "获取好评模板列表成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cookies/{cid}/comment-templates")
def add_comment_template(cid: str, template_data: CommentTemplateCreate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """添加好评模板"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        template_id = db_manager.add_comment_template(
            cid, 
            template_data.name, 
            template_data.content, 
            template_data.is_active
        )
        if template_id is None:
            raise HTTPException(status_code=500, detail="添加好评模板失败")

        return {
            "msg": "success",
            "template_id": template_id,
            "message": "添加好评模板成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/comment-templates/{template_id}")
def update_comment_template(cid: str, template_id: int, template_data: CommentTemplateUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新好评模板"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_comment_template(
            template_id,
            name=template_data.name,
            content=template_data.content,
            is_active=template_data.is_active
        )
        if not success:
            raise HTTPException(status_code=500, detail="更新好评模板失败")

        return {
            "msg": "success",
            "message": "更新好评模板成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/cookies/{cid}/comment-templates/{template_id}")
def delete_comment_template(cid: str, template_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除好评模板"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_comment_template(template_id)
        if not success:
            raise HTTPException(status_code=500, detail="删除好评模板失败")

        return {
            "msg": "success",
            "message": "删除好评模板成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/comment-templates/{template_id}/activate")
def activate_comment_template(cid: str, template_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """激活指定的好评模板"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.set_active_comment_template(cid, template_id)
        if not success:
            raise HTTPException(status_code=500, detail="激活好评模板失败")

        return {
            "msg": "success",
            "message": "激活好评模板成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新备注
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"更新账号备注: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "备注更新成功",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="备注更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取Cookie详细信息（包含备注）
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "获取备注成功"
            }
        else:
            raise HTTPException(status_code=404, detail="账号不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证暂停时间范围（0-60分钟，0表示不暂停）
        if not (0 <= update_data.pause_duration <= 60):
            raise HTTPException(status_code=400, detail="暂停时间必须在0-60分钟之间（0表示不暂停）")

        # 更新暂停时间
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"更新账号自动回复暂停时间: {cid} -> {update_data.pause_duration}分钟", current_user)
            return {
                "message": "暂停时间更新成功",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="暂停时间更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取暂停时间
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "获取暂停时间成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@app.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@app.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url'],
            "item_title": keyword_data.get('item_title', '')  # 添加商品名称
        })

    return result


@app.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@app.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        # 检查是否是图片关键词冲突
        if "已存在（图片关键词）" in error_msg:
            # 直接使用数据库管理器提供的友好错误信息
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            # 尝试从错误信息中提取具体的冲突关键词
            conflict_keyword = None
            conflict_type = None

            # 检查是否是数据库管理器抛出的详细错误
            if "关键词唯一约束冲突" in error_msg:
                # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                import re
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            # 构造用户友好的错误信息
            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@app.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的商品列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取该账号的所有商品
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, created_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY created_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or '未知商品',
                    'item_price': row[2] or '价格未知',
                    'created_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"获取商品列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@app.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")


@app.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        for index, row in df.iterrows():
            keyword = str(row['关键词']).strip()
            item_id = str(row['商品ID']).strip() if pd.notna(row['商品ID']) and str(row['商品ID']).strip() else None
            reply = str(row['关键词内容']).strip()

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")


@app.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")


@app.post("/keywords/{cid}/image-batch")
async def add_image_keyword_batch(
    cid: str,
    request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """批量添加图片关键词（使用已上传的图片URL）"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        body = await request.json()
        image_url = body.get('image_url', '').strip()
        keywords = body.get('keywords', [])
        item_ids = body.get('item_ids', [])

        if not image_url:
            raise HTTPException(status_code=400, detail="图片URL不能为空")

        if not keywords or len(keywords) == 0:
            raise HTTPException(status_code=400, detail="关键词列表不能为空")

        # 如果没有商品ID，则使用空字符串（通用关键词）
        if not item_ids or len(item_ids) == 0:
            item_ids = ['']

        logger.info(f"批量添加图片关键词: cid={cid}, keywords={keywords}, item_ids={item_ids}, image_url={image_url}")

        # 检查重复并批量添加
        success_count = 0
        fail_count = 0
        duplicates = []

        for keyword in keywords:
            keyword = keyword.strip()
            if not keyword:
                continue

            for item_id in item_ids:
                normalized_item_id = item_id if item_id and item_id.strip() else None

                # 检查是否重复
                if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
                    item_id_text = f"（商品ID: {normalized_item_id}）" if normalized_item_id else "（通用关键词）"
                    duplicates.append(f'"{keyword}" {item_id_text}')
                    fail_count += 1
                    continue

                # 保存图片关键词
                success = db_manager.save_image_keyword(cid, keyword, image_url, normalized_item_id)
                if success:
                    success_count += 1
                else:
                    fail_count += 1

        if duplicates:
            log_with_user('warning', f"批量添加图片关键词有重复: {cid}, duplicates={duplicates}", current_user)

        log_with_user('info', f"批量添加图片关键词完成: {cid}, success={success_count}, fail={fail_count}", current_user)

        return {
            "msg": "批量添加完成",
            "success_count": success_count,
            "fail_count": fail_count,
            "duplicates": duplicates,
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"批量添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"批量添加图片关键词失败: {str(e)}")


@app.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """上传图片（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")

        return {
            "message": "图片上传成功",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")


@app.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")


@app.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")


@app.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")


# 卡券管理API
@app.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的卡券列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新卡券"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_name = card_data.get('name', '未命名卡券')

        log_with_user('info', f"创建卡券: {card_name}", current_user)

        # 调试日志：记录接收到的多规格数据
        is_multi_spec = card_data.get('is_multi_spec', False)
        logger.info(f"[DEBUG] 创建卡券 - is_multi_spec: {is_multi_spec}")
        logger.info(f"[DEBUG] 创建卡券 - spec_name: {card_data.get('spec_name')}")
        logger.info(f"[DEBUG] 创建卡券 - spec_value: {card_data.get('spec_value')}")
        logger.info(f"[DEBUG] 创建卡券 - spec_name_2: {card_data.get('spec_name_2')}")
        logger.info(f"[DEBUG] 创建卡券 - spec_value_2: {card_data.get('spec_value_2')}")

        # 验证多规格字段
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_id = db_manager.create_card(
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            spec_name_2=card_data.get('spec_name_2') if is_multi_spec else None,
            spec_value_2=card_data.get('spec_value_2') if is_multi_spec else None,
            user_id=user_id
        )

        # 检查是否需要生成对应发货规则
        generate_delivery_rule = card_data.get('generate_delivery_rule', False)
        if generate_delivery_rule:
            try:
                # 生成发货规则
                rule_id = db_manager.create_delivery_rule(
                    keyword=card_data.get('name'),  # 商品关键字设置为卡券名称
                    card_id=card_id,  # 匹配卡券设置为当前新添加的卡券ID
                    delivery_count=1,  # 默认发货数量为1
                    enabled=True,  # 默认启用
                    description=f"自动生成的发货规则 - 对应卡券: {card_data.get('name')}",
                    user_id=user_id
                )
                log_with_user('info', f"自动生成发货规则成功: 卡券ID={card_id}, 规则ID={rule_id}", current_user)
            except Exception as e:
                log_with_user('error', f"生成发货规则失败: {str(e)}", current_user)
                # 不影响卡券创建，仅记录错误

        log_with_user('info', f"卡券创建成功: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "卡券创建成功"}
    except Exception as e:
        log_with_user('error', f"创建卡券失败: {card_data.get('name', '未知')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个卡券详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新卡券"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']

        # 调试日志：记录接收到的多规格数据
        is_multi_spec = card_data.get('is_multi_spec')
        logger.info(f"[DEBUG] 更新卡券 {card_id} - is_multi_spec: {is_multi_spec}")
        logger.info(f"[DEBUG] 更新卡券 {card_id} - spec_name: {card_data.get('spec_name')}")
        logger.info(f"[DEBUG] 更新卡券 {card_id} - spec_value: {card_data.get('spec_value')}")
        logger.info(f"[DEBUG] 更新卡券 {card_id} - spec_name_2: {card_data.get('spec_name_2')}")
        logger.info(f"[DEBUG] 更新卡券 {card_id} - spec_value_2: {card_data.get('spec_value_2')}")

        # 验证多规格字段
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value'),
            spec_name_2=card_data.get('spec_name_2'),
            spec_value_2=card_data.get('spec_value_2'),
            user_id=user_id
        )
        if success:
            return {"message": "卡券更新成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    spec_name_2: str = Form(default=""),
    spec_value_2: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新带图片的卡券"""
    try:
        logger.info(f"接收到带图片的卡券更新请求: card_id={card_id}, name={name}, type={type}")
        user_id = current_user['user_id']

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 验证多规格字段
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 更新卡券
        from db_manager import db_manager
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None,
            spec_name_2=spec_name_2 if is_multi_spec else None,
            spec_value_2=spec_value_2 if is_multi_spec else None,
            user_id=user_id
        )

        if success:
            logger.info(f"卡券更新成功: {name} (ID: {card_id})")
            return {"message": "卡券更新成功", "image_url": image_url}
        else:
            # 如果数据库更新失败，删除已保存的图片
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="卡券不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新带图片的卡券失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 自动发货规则API
@app.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货规则列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-rules/stats")
def get_delivery_stats(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货统计信息"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        today_count = db_manager.get_today_delivery_count(user_id)
        return {"today_delivery_count": today_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-logs/recent")
def get_recent_delivery_logs(limit: int = 20, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取最近发货日志（真实发货事件，含失败原因）"""
    try:
        from db_manager import db_manager

        def extract_spec_mode_context(reason: str):
            reason_text = (reason or '').strip()
            context = {
                'order_spec_mode': None,
                'rule_spec_mode': None,
                'item_config_mode': None
            }

            pattern = re.compile(r'\[(?:[^\]]*?)(order_spec_mode=[^\],]+|rule_spec_mode=[^\],]+|item_config_mode=[^\],]+)(?:[^\]]*?)\]$')
            if not reason_text or '[' not in reason_text or ']' not in reason_text:
                return reason_text, context

            bracket_start = reason_text.rfind('[')
            bracket_end = reason_text.rfind(']')
            if bracket_start == -1 or bracket_end == -1 or bracket_end < bracket_start:
                return reason_text, context

            suffix = reason_text[bracket_start:bracket_end + 1]
            if not pattern.search(suffix):
                return reason_text, context

            body = suffix[1:-1]
            for part in body.split(','):
                key, _, value = part.strip().partition('=')
                if key in context and value:
                    context[key] = value.strip()

            cleaned_reason = reason_text[:bracket_start].rstrip()
            return cleaned_reason or reason_text, context

        def is_redundant_skip_log(log: Dict[str, Any], successful_orders: set):
            if str(log.get('status') or '').lower() != 'skipped':
                return False

            reason_text = str(log.get('reason') or '').strip()
            order_id = str(log.get('order_id') or '').strip()
            if not order_id or order_id not in successful_orders:
                return False

            redundant_reasons = {
                '获取锁后发现订单已处理，跳过发货',
                '订单延迟锁持有中，跳过发货',
                '订单在冷却期内，跳过发货',
            }
            return reason_text in redundant_reasons

        user_id = current_user['user_id']
        safe_limit = max(1, min(int(limit), 200))
        raw_logs = db_manager.get_recent_delivery_logs(user_id=user_id, limit=min(safe_limit * 3, 600))
        successful_orders = {
            str(log.get('order_id') or '').strip()
            for log in raw_logs
            if str(log.get('status') or '').lower() == 'success' and str(log.get('order_id') or '').strip()
        }

        logs = []
        for log in raw_logs:
            cleaned_reason, context = extract_spec_mode_context(log.get('reason'))
            log['reason'] = cleaned_reason
            log.update(context)
            if is_redundant_skip_log(log, successful_orders):
                continue
            logs.append(log)
            if len(logs) >= safe_limit:
                break
        return {"logs": logs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_id = rule_data.get('card_id')

        if card_id is not None:
            card = db_manager.get_card_by_id(card_id, user_id)
            if not card:
                raise HTTPException(status_code=404, detail="卡券不存在")

        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=card_id,
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "发货规则创建成功"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个发货规则详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_id = rule_data.get('card_id')

        if card_id is not None:
            card = db_manager.get_card_by_id(card_id, user_id)
            if not card:
                raise HTTPException(status_code=404, detail="卡券不存在")

        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=card_id,
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "发货规则更新成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/cards/{card_id}")
def delete_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除卡券"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_card(card_id, user_id)
        if success:
            return {"message": "卡券删除成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "发货规则删除成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 备份和恢复 API ====================

@app.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出用户备份"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # 导出当前用户的数据
        backup_data = db_manager.export_backup(user_id)

        # 生成文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # 返回JSON响应，设置下载头
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")


@app.post("/backup/import")
def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入用户备份"""
    try:
        # 验证文件类型
        if not file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

        # 读取文件内容
        content = file.file.read()
        backup_data = json.loads(content.decode('utf-8'))

        # 导入备份到当前用户
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            # 备份导入成功后，刷新 CookieManager 的内存缓存
            import cookie_manager
            if cookie_manager.manager:
                try:
                    cookie_manager.manager.reload_from_db()
                    logger.info("备份导入后已刷新 CookieManager 缓存")
                except Exception as e:
                    logger.error(f"刷新 CookieManager 缓存失败: {e}")

            return {"message": "备份导入成功"}
        else:
            raise HTTPException(status_code=400, detail="备份导入失败")

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="备份文件格式无效")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入备份失败: {str(e)}")


@app.post("/system/reload-cache")
def reload_cache(current_user: Dict[str, Any] = Depends(get_current_user)):
    """重新加载系统缓存（用于手动刷新数据）"""
    try:
        import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "系统缓存已刷新", "success": True}
            else:
                raise HTTPException(status_code=500, detail="缓存刷新失败")
        else:
            raise HTTPException(status_code=500, detail="CookieManager 未初始化")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")


# ==================== 商品管理 API ====================

@app.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


# ==================== 商品搜索 API ====================

class ItemSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 20

class ItemSearchMultipleRequest(BaseModel):
    keyword: str
    total_pages: int = 1


def _parse_optional_non_negative_float(value: Any, field_label: str) -> Optional[float]:
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    try:
        parsed = float(raw_value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"{field_label}必须是数字")

    if parsed < 0:
        raise HTTPException(status_code=400, detail=f"{field_label}必须大于等于 0")

    return parsed


def _parse_form_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    normalized = str(value or "").strip().lower()
    return normalized in {"1", "true", "yes", "on", "y"}


def _persist_cookie_value_for_account(
    cookie_id: str,
    current_user: Dict[str, Any],
    original_cookie_value: str,
    latest_cookie_value: str,
):
    cleaned_latest = str(latest_cookie_value or "").strip()
    if not cleaned_latest or cleaned_latest == str(original_cookie_value or "").strip():
        return

    db_manager.update_cookie_account_info(
        cookie_id,
        cookie_value=cleaned_latest,
        user_id=current_user["user_id"],
    )
    if cookie_manager.manager is not None:
        cookie_manager.manager.update_cookie(cookie_id, cleaned_latest, save_to_db=False)


async def _sync_items_after_publish(
    cookie_id: str,
    cookies_str: str,
    published_item_id: Optional[str] = None,
) -> Dict[str, Any]:
    from XianyuAutoAsync import XianyuLive

    xianyu_instance = XianyuLive(cookies_str, cookie_id, register_instance=False)
    fallback_result = None
    page_sync_result = None
    item_synced = None

    try:
        page_sync_result = await xianyu_instance.get_item_list_info(
            page_number=1,
            page_size=100,
            sync_item_details=True,
        )

        if published_item_id:
            item_synced = bool(db_manager.get_item_info(cookie_id, published_item_id))

        if published_item_id and not item_synced:
            fallback_result = await xianyu_instance.get_all_items(
                page_size=100,
                max_pages=3,
                sync_item_details=True,
            )
            item_synced = bool(db_manager.get_item_info(cookie_id, published_item_id))

        sync_success = bool(page_sync_result and page_sync_result.get("success"))
        fallback_success = bool(fallback_result and fallback_result.get("success"))

        summary_message = "已同步最新商品列表"
        if published_item_id:
            if item_synced:
                summary_message = f"已同步发布商品 {published_item_id}"
            else:
                summary_message = f"已执行同步，但暂未在本地列表确认商品 {published_item_id}"
        elif not sync_success and not fallback_success:
            summary_message = "发布成功，但同步最新商品列表失败"

        return {
            "success": sync_success or fallback_success,
            "message": summary_message,
            "published_item_id": published_item_id,
            "item_synced": item_synced,
            "page_sync": {
                "success": bool(page_sync_result and page_sync_result.get("success")),
                "current_count": int((page_sync_result or {}).get("current_count", 0) or 0),
                "saved_count": int((page_sync_result or {}).get("saved_count", 0) or 0),
                "error": (page_sync_result or {}).get("error"),
            },
            "full_sync": {
                "used": fallback_result is not None,
                "success": bool(fallback_result and fallback_result.get("success")),
                "total_count": int((fallback_result or {}).get("total_count", 0) or 0),
                "total_saved": int((fallback_result or {}).get("total_saved", 0) or 0),
                "error": (fallback_result or {}).get("error"),
            },
        }
    finally:
        await xianyu_instance.close_session()


@app.post("/items/search")
async def search_items(
    search_request: ItemSearchRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始单页搜索: 关键词='{search_request.keyword}', 页码={search_request.page}, 每页={search_request.page_size}")

        from utils.item_search import search_xianyu_items

        # 执行搜索
        result = await search_xianyu_items(
            keyword=search_request.keyword,
            page=search_request.page,
            page_size=search_request.page_size
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 单页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "page": search_request.page,
            "page_size": search_request.page_size,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"商品搜索失败: {error_msg}")


@app.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """检查是否有有效的cookies账户（必须是启用状态）"""
    try:
        if cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        from db_manager import db_manager

        # 获取所有cookies
        all_cookies = db_manager.get_all_cookies()

        # 检查启用状态和有效性
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # 检查是否启用
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # 检查是否有效（长度大于50）
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"检查cookies失败: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }

@app.post("/items/search_multiple")
async def search_multiple_pages(
    search_request: ItemSearchMultipleRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索多页闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始多页搜索: 关键词='{search_request.keyword}', 页数={search_request.total_pages}")

        from utils.item_search import search_multiple_pages_xianyu

        # 执行多页搜索
        result = await search_multiple_pages_xianyu(
            keyword=search_request.keyword,
            total_pages=search_request.total_pages
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 多页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "total_pages": search_request.total_pages,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "is_fallback": result.get("is_fallback", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 多页商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"多页商品搜索失败: {error_msg}")


@app.post("/item-publish")
async def publish_item(
    cookie_id: str = Form(...),
    title: str = Form(...),
    description: str = Form(default=""),
    current_price: str = Form(default=""),
    original_price: str = Form(default=""),
    delivery_choice: str = Form(...),
    post_price: str = Form(default=""),
    can_self_pickup: str = Form(default="false"),
    images: List[UploadFile] = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """发布单个商品，并在成功后同步到本地商品列表。"""
    user_prefix = get_user_log_prefix(current_user)

    cleaned_cookie_id = _ensure_cookie_access(cookie_id, current_user)
    cookies_map = _get_user_cookies_map(current_user)
    cookies_str = str(cookies_map.get(cleaned_cookie_id) or "").strip()
    if not cookies_str:
        raise HTTPException(status_code=400, detail="账号 Cookie 为空，无法发布商品")

    cleaned_title = str(title or "").strip()
    cleaned_description = str(description or "").strip()
    if not cleaned_title:
        raise HTTPException(status_code=400, detail="商品标题不能为空")
    if not cleaned_description:
        raise HTTPException(status_code=400, detail="商品描述不能为空")

    if not images:
        raise HTTPException(status_code=400, detail="请至少上传 1 张商品图片")
    if len(images) > 9:
        raise HTTPException(status_code=400, detail="单次最多上传 9 张商品图片")

    current_price_value = _parse_optional_non_negative_float(current_price, "现价")
    original_price_value = _parse_optional_non_negative_float(original_price, "原价")
    post_price_value = _parse_optional_non_negative_float(post_price, "邮费")
    can_self_pickup_value = _parse_form_bool(can_self_pickup)

    if original_price_value is not None and current_price_value is None:
        raise HTTPException(status_code=400, detail="填写原价时必须同时填写现价")
    if delivery_choice == "一口价" and post_price_value is None:
        raise HTTPException(status_code=400, detail="运费方式为一口价时必须填写邮费")

    image_payloads = []
    for index, image in enumerate(images, start=1):
        if image.content_type and not image.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail=f"第 {index} 张文件不是图片")

        image_content = await image.read()
        if not image_content:
            raise HTTPException(status_code=400, detail=f"第 {index} 张图片为空")

        image_payloads.append(
            {
                "filename": image.filename or f"publish-image-{index}.jpg",
                "content": image_content,
            }
        )

    try:
        from utils.item_publisher import ItemPublisher

        logger.info(
            f"{user_prefix} 开始发布商品: cookie_id={cleaned_cookie_id}, "
            f"title={cleaned_title}, images={len(image_payloads)}, delivery_choice={delivery_choice}"
        )

        async with ItemPublisher(cookies_str, cleaned_cookie_id) as publisher:
            publish_result = await publisher.publish_item(
                title=cleaned_title,
                description=cleaned_description,
                images=image_payloads,
                current_price=current_price_value,
                original_price=original_price_value,
                delivery_choice=delivery_choice,
                post_price=post_price_value,
                can_self_pickup=can_self_pickup_value,
            )
            latest_cookies_str = publisher.cookies_str
            published_item_id = publisher.extract_published_item_id(publish_result)

            if not publisher.is_success_response(publish_result):
                raise HTTPException(
                    status_code=400,
                    detail=f"商品发布失败: {publisher.extract_error_message(publish_result)}",
                )

        _persist_cookie_value_for_account(
            cleaned_cookie_id,
            current_user,
            cookies_str,
            latest_cookies_str,
        )

        sync_result = await _sync_items_after_publish(
            cleaned_cookie_id,
            latest_cookies_str or cookies_str,
            published_item_id=published_item_id,
        )

        sync_success = bool(sync_result.get("success"))
        success_message = "商品发布成功"
        if sync_success:
            success_message = "商品发布成功，已同步到商品管理"
        elif sync_result.get("message"):
            success_message = f"商品发布成功，{sync_result['message']}"

        logger.info(
            f"{user_prefix} 商品发布完成: cookie_id={cleaned_cookie_id}, "
            f"published_item_id={published_item_id or 'unknown'}, sync_success={sync_success}"
        )

        return {
            "success": True,
            "message": success_message,
            "published_item_id": published_item_id,
            "publish_result": publish_result,
            "sync_result": sync_result,
        }

    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except RuntimeError as exc:
        logger.error(f"{user_prefix} 商品发布运行失败: {mask_sensitive_text(exc)}")
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        logger.error(f"{user_prefix} 商品发布异常: {mask_sensitive_text(exc)}")
        raise HTTPException(status_code=500, detail=f"商品发布异常: {str(exc)}")



@app.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_items_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@app.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="商品不存在")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_detail: str


@app.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_item_detail(cookie_id, item_id, update_data.item_detail)
        if success:
            return {"message": "商品详情更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品详情失败: {str(e)}")


@app.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "商品信息删除成功"}
        else:
            raise HTTPException(status_code=404, detail="商品信息不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    model_name: str = "qwen-plus"
    api_key: str = ""
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    api_type: str = ""
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""


class AIConfigPreset(BaseModel):
    preset_name: str
    model_name: str
    api_key: str = ""
    base_url: str = ""
    api_type: str = ""


@app.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """批量删除商品信息"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="删除列表不能为空")

        success_count = db_manager.batch_delete_item_info(request.items)
        total_count = len(request.items)

        return {
            "message": f"批量删除完成",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except Exception as e:
        logger.error(f"批量删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== AI回复管理API ====================

@app.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        settings = db_manager.get_ai_reply_settings(cookie_id)
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        # 保存设置
        settings_dict = settings.dict()
        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:

            # 如果启用了AI回复，记录日志
            if settings.ai_enabled:
                logger.info(f"账号 {cookie_id} 启用AI回复")
            else:
                logger.info(f"账号 {cookie_id} 禁用AI回复")

            return {"message": "AI回复设置更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的AI回复设置"""
    try:
        # 只返回当前用户的AI回复设置
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_settings = db_manager.get_all_ai_reply_settings()
        # 过滤只属于当前用户的设置
        user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
        return user_settings
    except Exception as e:
        logger.error(f"获取所有AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.get("/ai-config-presets")
def list_ai_config_presets(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的AI配置预设列表"""
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager
        presets = db_manager.get_ai_config_presets(user_id)
        return presets
    except Exception as e:
        logger.error(f"获取AI配置预设列表异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.post("/ai-config-presets")
def save_ai_config_preset(
    preset: AIConfigPreset,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """创建或更新AI配置预设"""
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 检查预设数量上限
        existing = db_manager.get_ai_config_presets(user_id)
        existing_names = [p['preset_name'] for p in existing]
        if preset.preset_name not in existing_names and len(existing) >= 20:
            raise HTTPException(status_code=400, detail="预设数量已达上限（最多20个）")

        preset_id = db_manager.save_ai_config_preset(
            user_id=user_id,
            preset_name=preset.preset_name,
            model_name=preset.model_name,
            api_key=preset.api_key,
            base_url=preset.base_url,
            api_type=preset.api_type
        )
        return {"message": "预设保存成功", "preset_id": preset_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存AI配置预设异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.delete("/ai-config-presets/{preset_id}")
def delete_ai_config_preset(
    preset_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除AI配置预设"""
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager
        deleted = db_manager.delete_ai_config_preset(user_id, preset_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="预设不存在或无权删除")
        return {"message": "预设删除成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除AI配置预设异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(cookie_id: str, test_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """测试AI回复功能"""
    try:
        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='账号不存在')

        # 检查是否启用AI回复
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')

        # 构造测试数据
        test_message = test_data.get('message', '你好')
        test_item_info = {
            'title': test_data.get('item_title', '测试商品'),
            'price': test_data.get('item_price', 100),
            'desc': test_data.get('item_desc', '这是一个测试商品')
        }

        # 生成测试回复（跳过去抖等待）
        reply = ai_reply_engine.generate_reply(
            message=test_message,
            item_info=test_item_info,
            chat_id=f"test_{int(time.time())}",
            cookie_id=cookie_id,
            user_id="test_user",
            item_id="test_item",
            skip_wait=True
        )

        if reply:
            return {"message": "测试成功", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AI回复生成失败")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"测试AI回复异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== 日志管理API ====================

@app.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取实时系统日志"""
    try:
        # 获取文件日志收集器
        collector = get_file_log_collector()

        # 获取日志
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}


@app.get("/risk-control-logs")
async def get_risk_control_logs(
    cookie_id: str = None,
    processing_status: str = None,
    event_type: str = None,
    trigger_scene: str = None,
    session_id: str = None,
    result_code: str = None,
    date_from: str = None,
    date_to: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user(
            'info',
            f"查询风控日志: cookie_id={cookie_id}, processing_status={processing_status}, event_type={event_type}, trigger_scene={trigger_scene}, session_id={session_id}, result_code={result_code}, date_from={date_from}, date_to={date_to}, limit={limit}, offset={offset}",
            admin_user,
        )

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(
            cookie_id=cookie_id,
            processing_status=processing_status,
            event_type=event_type,
            trigger_scene=trigger_scene,
            session_id=session_id,
            result_code=result_code,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset
        )
        total_count = db_manager.get_risk_control_logs_count(
            cookie_id=cookie_id,
            processing_status=processing_status,
            event_type=event_type,
            trigger_scene=trigger_scene,
            session_id=session_id,
            result_code=result_code,
            date_from=date_from,
            date_to=date_to,
        )

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"获取风控日志失败: {str(e)}", admin_user)
        return {
            "success": False,
            "message": f"获取风控日志失败: {str(e)}",
            "data": [],
            "total": 0
        }


@app.get("/admin/slider-verification-stats")
async def get_slider_verification_stats(
    cookie_id: str = None,
    range_key: str = 'all',
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取当前系统用户下的滑块验证统计。"""
    try:
        user_id = admin_user['user_id']
        user_cookie_ids = sorted(db_manager.get_all_cookies(user_id).keys())
        normalized_range = str(range_key or '').strip().lower()
        if normalized_range not in {'today', '7d', 'all'}:
            normalized_range = 'all'
        range_label = {
            'today': '当日',
            '7d': '近 7 天',
            'all': '所有',
        }[normalized_range]

        if cookie_id:
            if cookie_id not in user_cookie_ids:
                return {
                    'success': True,
                    'data': {
                        **_empty_slider_session_stats(),
                        'scope_label': cookie_id,
                        'selected_cookie_id': cookie_id,
                        'selected_range': normalized_range,
                        'range_label': range_label,
                        'summary_text': '暂无滑块验证记录' if normalized_range == 'all' else f'{range_label}暂无滑块验证记录',
                    }
                }
            target_cookie_ids = [cookie_id]
            scope_label = cookie_id
        else:
            target_cookie_ids = user_cookie_ids
            scope_label = '全部账号'

        stats = db_manager.get_slider_verification_session_stats(target_cookie_ids, range_key=normalized_range)
        stats.update({
            'scope_label': scope_label,
            'selected_cookie_id': cookie_id or '',
        })

        log_with_user(
            'info',
            f"获取滑块验证统计成功: scope={scope_label}, range={range_label}, sessions={stats['total_sessions']}, success={stats['success_count']}, failure={stats['failure_count']}",
            admin_user,
        )

        return {
            'success': True,
            'data': stats,
        }
    except Exception as e:
        log_with_user('error', f"获取滑块验证统计失败: {str(e)}", admin_user)
        return {
            'success': False,
            'message': f'获取滑块验证统计失败: {str(e)}',
            'data': _empty_slider_session_stats(),
        }


@app.delete("/admin/risk-control-logs/{log_id}")
async def delete_risk_control_log(
    log_id: int,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """删除风控日志记录（管理员专用）"""
    try:
        log_with_user('info', f"删除风控日志记录: {log_id}", admin_user)

        success = db_manager.delete_risk_control_log(log_id)

        if success:
            log_with_user('info', f"风控日志删除成功: {log_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"风控日志删除失败: {log_id}", admin_user)
            return {"success": False, "message": "删除失败，记录可能不存在"}

    except Exception as e:
        log_with_user('error', f"删除风控日志失败: {log_id} - {str(e)}", admin_user)
        return {"success": False, "message": f"删除失败: {str(e)}"}


@app.get("/logs/stats")
async def get_log_stats(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取日志统计信息"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}


@app.post("/logs/clear")
async def clear_logs(current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空日志"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "日志已清空"}

    except Exception as e:
        return {"success": False, "message": f"清空日志失败: {str(e)}"}


# ==================== 商品管理API ====================

@app.post("/items/get-all-from-account")
async def get_all_items_from_account(request: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """从指定账号获取所有商品信息"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 获取指定账号的cookie信息
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id, register_instance=False)

        # 调用获取所有商品信息的方法（自动分页）并同步最新商品详情
        logger.info(f"开始同步账号 {cookie_id} 的所有商品信息和最新详情")
        result = await xianyu_instance.get_all_items(sync_item_details=True)

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            logger.info(f"成功同步账号 {cookie_id} 的 {total_count} 个商品（共{total_pages}页）")
            return {
                "success": True,
                "message": f"成功同步 {total_count} 个商品（共{total_pages}页），最新商品详情已更新",
                "total_count": total_count,
                "total_pages": total_pages
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


@app.post("/items/get-by-page")
async def get_items_by_page(request: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """从指定账号按页获取商品信息"""
    try:
        # 验证参数
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 验证分页参数
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "页码和每页数量必须是数字"}

        if page_number < 1:
            return {"success": False, "message": "页码必须大于0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "每页数量必须在1-100之间"}

        # 获取账号信息
        account = db_manager.get_cookie_by_id(cookie_id)
        if not account:
            return {"success": False, "message": "账号不存在"}

        cookies_str = account['cookies_str']
        if not cookies_str:
            return {"success": False, "message": "账号cookies为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id, register_instance=False)

        # 调用获取指定页商品信息的方法并同步最新商品详情
        logger.info(f"开始同步账号 {cookie_id} 第{page_number}页商品信息和最新详情（每页{page_size}条）")
        result = await xianyu_instance.get_item_list_info(page_number, page_size, sync_item_details=True)

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            current_count = result.get('current_count', 0)
            logger.info(f"成功同步账号 {cookie_id} 第{page_number}页 {current_count} 个商品")
            return {
                "success": True,
                "message": f"成功同步第{page_number}页 {current_count} 个商品，最新商品详情已更新",
                "page_number": page_number,
                "page_size": page_size,
                "current_count": current_count
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


# ------------------------- 用户设置接口 -------------------------

@app.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新用户设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"更新用户设置: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"用户设置更新成功: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"用户设置更新失败: {key}", current_user)
            raise HTTPException(status_code=400, detail='更新失败')
    except Exception as e:
        log_with_user('error', f"更新用户设置异常: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取用户特定设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='设置不存在')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 管理员专用接口 -------------------------

@app.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有用户信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询所有用户信息", admin_user)
        users = db_manager.get_all_users()

        # 为每个用户添加统计信息
        for user in users:
            user_id = user['id']
            # 统计用户的Cookie数量
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # 统计用户的卡券数量
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # 隐藏密码字段
            if 'password_hash' in user:
                del user['password_hash']

        log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/users/{user_id}')
def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除用户（管理员专用）"""
    from db_manager import db_manager
    try:
        # 不能删除管理员自己
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 获取要删除的用户信息
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="用户不存在")

        log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # 删除用户及其相关数据
        success = db_manager.delete_user_and_data(user_id)

        if success:
            log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"用户 {user_to_delete['username']} 删除成功"}
        else:
            log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="删除失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/admin/users/{user_id}/admin-status')
def update_user_admin_status(user_id: int, is_admin: bool, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新用户管理员状态（管理员专用）"""
    from db_manager import db_manager
    try:
        # 获取目标用户信息
        target_user = db_manager.get_user_by_id(user_id)
        if not target_user:
            raise HTTPException(status_code=404, detail="用户不存在")

        # 不能修改自己的管理员状态（防止误操作导致没有管理员）
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试修改自己的管理员状态", admin_user)
            raise HTTPException(status_code=400, detail="不能修改自己的管理员状态")

        log_with_user('info', f"准备{'设置' if is_admin else '取消'}{target_user['username']}的管理员权限", admin_user)

        # 更新管理员状态
        success = db_manager.update_user_admin_status(user_id, is_admin)

        if success:
            action = "设置为管理员" if is_admin else "取消管理员权限"
            log_with_user('info', f"用户 {target_user['username']} 已{action}", admin_user)
            return {
                "success": True,
                "message": f"用户 {target_user['username']} 已{action}",
                "user_id": user_id,
                "is_admin": is_admin
            }
        else:
            log_with_user('error', f"更新用户管理员状态失败: {target_user['username']}", admin_user)
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"更新用户管理员状态异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/risk-control-logs')
async def get_admin_risk_control_logs(
    cookie_id: str = None,
    processing_status: str = None,
    event_type: str = None,
    trigger_scene: str = None,
    session_id: str = None,
    result_code: str = None,
    date_from: str = None,
    date_to: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user(
            'info',
            f"查询风控日志: cookie_id={cookie_id}, processing_status={processing_status}, event_type={event_type}, trigger_scene={trigger_scene}, session_id={session_id}, result_code={result_code}, date_from={date_from}, date_to={date_to}, limit={limit}, offset={offset}",
            admin_user,
        )

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(
            cookie_id=cookie_id,
            processing_status=processing_status,
            event_type=event_type,
            trigger_scene=trigger_scene,
            session_id=session_id,
            result_code=result_code,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset
        )
        total_count = db_manager.get_risk_control_logs_count(
            cookie_id=cookie_id,
            processing_status=processing_status,
            event_type=event_type,
            trigger_scene=trigger_scene,
            session_id=session_id,
            result_code=result_code,
            date_from=date_from,
            date_to=date_to,
        )

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"查询风控日志失败: {str(e)}", admin_user)
        return {"success": False, "message": f"查询失败: {str(e)}", "data": [], "total": 0}


@app.get('/admin/cookies')
def get_admin_cookies(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有Cookie信息（管理员专用）"""
    try:
        log_with_user('info', "查询所有Cookie信息", admin_user)

        if cookie_manager.manager is None:
            return {
                "success": True,
                "cookies": [],
                "message": "CookieManager 未就绪"
            }

        # 获取所有用户的cookies
        from db_manager import db_manager
        all_users = db_manager.get_all_users()
        all_cookies = []

        for user in all_users:
            user_id = user['id']
            user_cookies = db_manager.get_all_cookies(user_id)
            for cookie_id, cookie_value in user_cookies.items():
                # 获取cookie详细信息
                cookie_details = db_manager.get_cookie_details(cookie_id)
                cookie_info = {
                    'cookie_id': cookie_id,
                    'user_id': user_id,
                    'username': user['username'],
                    'nickname': cookie_details.get('remark', '') if cookie_details else '',
                    'enabled': cookie_manager.manager.get_cookie_status(cookie_id)
                }
                all_cookies.append(cookie_info)

        log_with_user('info', f"获取到 {len(all_cookies)} 个Cookie", admin_user)
        return {
            "success": True,
            "cookies": all_cookies,
            "total": len(all_cookies)
        }

    except Exception as e:
        log_with_user('error', f"获取Cookie信息失败: {str(e)}", admin_user)
        return {
            "success": False,
            "cookies": [],
            "message": f"获取失败: {str(e)}"
        }


@app.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_admin),
                   lines: int = 100,
                   level: str = None):
    """获取系统日志（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

        # 查找日志文件
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"找到日志文件: {log_files}")

        if not log_files:
            logger.warning("未找到日志文件")
            return {"logs": [], "message": "未找到日志文件", "success": False}

        # 获取最新的日志文件
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"使用最新日志文件: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"读取到 {len(all_lines)} 行日志")

                # 如果指定了日志级别，进行过滤
                if level:
                    filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                    logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                else:
                    filtered_lines = all_lines

                # 获取最后N行
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"取最后 {len(recent_lines)} 行日志")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"读取日志文件失败: {str(e)}")
            log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

        log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
        logger.info(f"成功返回 {len(logs)} 条日志记录")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"获取系统日志失败: {str(e)}")
        log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
        return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

@app.get('/admin/log-files')
def list_log_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出所有可用的系统日志文件"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询日志文件列表", admin_user)

        log_dir = "logs"
        if not os.path.exists(log_dir):
            logger.warning("日志目录不存在")
            return {"success": True, "files": []}

        log_pattern = os.path.join(log_dir, "xianyu_*.log")
        log_files = glob.glob(log_pattern)

        files_info = []
        for file_path in log_files:
            try:
                stat_info = os.stat(file_path)
                files_info.append({
                    "name": os.path.basename(file_path),
                    "size": stat_info.st_size,
                    "modified_at": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                    "modified_ts": stat_info.st_mtime
                })
            except OSError as e:
                logger.warning(f"读取日志文件信息失败 {file_path}: {e}")

        # 按修改时间倒序排序
        files_info.sort(key=lambda item: item.get("modified_ts", 0), reverse=True)

        logger.info(f"返回日志文件列表，共 {len(files_info)} 个文件")
        return {"success": True, "files": files_info}

    except Exception as e:
        logger.error(f"获取日志文件列表失败: {str(e)}")
        log_with_user('error', f"获取日志文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/logs/export')
def export_log_file(file: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """导出指定的日志文件"""
    import os
    from fastapi.responses import StreamingResponse

    try:
        if not file:
            raise HTTPException(status_code=400, detail="缺少文件参数")

        safe_name = os.path.basename(file)
        log_dir = os.path.abspath("logs")
        target_path = os.path.abspath(os.path.join(log_dir, safe_name))

        # 防止目录遍历
        if not target_path.startswith(log_dir):
            log_with_user('warning', f"尝试访问非法日志文件: {file}", admin_user)
            raise HTTPException(status_code=400, detail="非法的日志文件路径")

        if not os.path.exists(target_path):
            log_with_user('warning', f"日志文件不存在: {file}", admin_user)
            raise HTTPException(status_code=404, detail="日志文件不存在")

        log_with_user('info', f"导出日志文件: {safe_name}", admin_user)
        def iter_file(path: str):
            file_handle = open(path, 'rb')
            try:
                while True:
                    chunk = file_handle.read(8192)
                    if not chunk:
                        break
                    yield chunk
            finally:
                file_handle.close()

        headers = {
            "Content-Disposition": f'attachment; filename="{safe_name}"'
        }
        return StreamingResponse(
            iter_file(target_path),
            media_type='text/plain; charset=utf-8',
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出日志文件失败: {str(e)}")
        log_with_user('error', f"导出日志文件失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取系统统计信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询系统统计信息", admin_user)

        stats = {
            "users": {
                "total": 0,
                "active_today": 0
            },
            "cookies": {
                "total": 0,
                "enabled": 0
            },
            "cards": {
                "total": 0,
                "enabled": 0
            },
            "system": {
                "uptime": "未知",
                "version": "1.0.0"
            }
        }

        # 用户统计
        all_users = db_manager.get_all_users()
        stats["users"]["total"] = len(all_users)

        # Cookie统计
        all_cookies = db_manager.get_all_cookies()
        stats["cookies"]["total"] = len(all_cookies)

        # 卡券统计
        all_cards = db_manager.get_all_cards()
        if all_cards:
            stats["cards"]["total"] = len(all_cards)
            stats["cards"]["enabled"] = len([card for card in all_cards if card.get('enabled', True)])

        log_with_user('info', "系统统计信息查询完成", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- 指定商品回复接口 -------------------------

@app.get("/itemReplays")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品回复信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复信息失败: {str(e)}")

@app.get("/itemReplays/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")

@app.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品回复失败: {str(e)}")

@app.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除商品回复失败: {str(e)}")

class ItemToDelete(BaseModel):
    cookie_id: str
    item_id: str

class BatchDeleteRequest(BaseModel):
    items: List[ItemToDelete]

@app.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@app.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取指定商品回复
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # 找对应item_id的回复
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复失败: {str(e)}")


# ------------------------- 数据库备份和恢复接口 -------------------------

@app.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_admin)):
    """下载数据库备份文件（管理员专用）"""
    import os
    from fastapi.responses import FileResponse
    from datetime import datetime

    try:
        log_with_user('info', "请求下载数据库备份", admin_user)

        # 使用db_manager的实际数据库路径
        from db_manager import db_manager
        db_file_path = db_manager.db_path

        # 检查数据库文件是否存在
        if not os.path.exists(db_file_path):
            log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="数据库文件不存在")

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/admin/backup/upload')
async def upload_database_backup(admin_user: Dict[str, Any] = Depends(require_admin),
                                backup_file: UploadFile = File(...)):
    """上传并恢复数据库备份文件（管理员专用）"""
    import os
    import shutil
    import sqlite3
    from datetime import datetime

    try:
        log_with_user('info', f"开始上传数据库备份: {backup_file.filename}", admin_user)

        # 验证文件类型
        if not backup_file.filename.endswith('.db'):
            log_with_user('warning', f"无效的备份文件类型: {backup_file.filename}", admin_user)
            raise HTTPException(status_code=400, detail="只支持.db格式的数据库文件")

        # 验证文件大小（限制100MB）
        content = await backup_file.read()
        if len(content) > 100 * 1024 * 1024:  # 100MB
            log_with_user('warning', f"备份文件过大: {len(content)} bytes", admin_user)
            raise HTTPException(status_code=400, detail="备份文件大小不能超过100MB")

        # 验证是否为有效的SQLite数据库文件
        temp_file_path = f"temp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        try:
            # 保存临时文件
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(content)

            # 验证数据库文件完整性
            conn = sqlite3.connect(temp_file_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()

            # 检查是否包含必要的表
            table_names = [table[0] for table in tables]
            required_tables = ['users', 'cookies']  # 最基本的表

            missing_tables = [table for table in required_tables if table not in table_names]
            if missing_tables:
                log_with_user('warning', f"备份文件缺少必要的表: {missing_tables}", admin_user)
                raise HTTPException(status_code=400, detail=f"备份文件不完整，缺少表: {', '.join(missing_tables)}")

            log_with_user('info', f"备份文件验证通过，包含 {len(table_names)} 个表", admin_user)

        except sqlite3.Error as e:
            log_with_user('error', f"备份文件验证失败: {str(e)}", admin_user)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            raise HTTPException(status_code=400, detail="无效的数据库文件")

        # 备份当前数据库
        from db_manager import db_manager
        current_db_path = db_manager.db_path

        # 生成备份文件路径（与原数据库在同一目录）
        db_dir = os.path.dirname(current_db_path)
        backup_filename = f"xianyu_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_current_path = os.path.join(db_dir, backup_filename)

        if os.path.exists(current_db_path):
            shutil.copy2(current_db_path, backup_current_path)
            log_with_user('info', f"当前数据库已备份为: {backup_current_path}", admin_user)

        # 关闭当前数据库连接
        if hasattr(db_manager, 'conn') and db_manager.conn:
            db_manager.conn.close()
            log_with_user('info', "已关闭当前数据库连接", admin_user)

        # 替换数据库文件
        shutil.move(temp_file_path, current_db_path)
        log_with_user('info', f"数据库文件已替换: {current_db_path}", admin_user)

        # 重新初始化数据库连接（使用原有的db_path）
        db_manager.__init__(db_manager.db_path)
        log_with_user('info', "数据库连接已重新初始化", admin_user)

        # 验证新数据库
        try:
            test_users = db_manager.get_all_users()
            log_with_user('info', f"数据库恢复成功，包含 {len(test_users)} 个用户", admin_user)
        except Exception as e:
            log_with_user('error', f"数据库恢复后验证失败: {str(e)}", admin_user)
            # 如果验证失败，尝试恢复原数据库
            if os.path.exists(backup_current_path):
                shutil.copy2(backup_current_path, current_db_path)
                db_manager.__init__()
                log_with_user('info', "已恢复原数据库", admin_user)
            raise HTTPException(status_code=500, detail="数据库恢复失败，已回滚到原数据库")

        return {
            "success": True,
            "message": "数据库恢复成功",
            "backup_file": backup_current_path,
            "user_count": len(test_users)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"上传数据库备份失败: {str(e)}", admin_user)
        # 清理临时文件
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出服务器上的备份文件（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询备份文件列表", admin_user)

        # 查找备份文件（在data目录中）
        backup_files = glob.glob("data/xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

        # 按修改时间倒序排列
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 数据管理接口 -------------------------

@app.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许访问该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/data/{table_name}/export')
def export_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """导出指定表的数据为Excel文件（管理员专用）"""
    from db_manager import db_manager
    import io
    try:
        log_with_user('info', f"导出表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试导出不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许导出该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        if not data:
            raise HTTPException(status_code=400, detail="表中没有数据")

        # 创建Excel文件
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_with_user('info', f"表 {table_name} 导出成功，共 {len(data)} 条记录", admin_user)

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={table_name}_export.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"导出表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除指定表的指定记录（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许操作该表")

        # 特殊保护：不能删除管理员用户
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 删除记录
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """清空指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"清空表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        # 不允许清空用户表
        if table_name == 'users':
            log_with_user('warning', "尝试清空用户表", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空用户表")

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空该表")

        # 清空表数据
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
            return {"success": True, "message": "清空成功"}
        else:
            log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="清空失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 商品多规格管理API
@app.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(cookie_id: str, item_id: str, spec_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新商品的多规格状态"""
    try:
        from db_manager import db_manager

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"商品多规格状态已{'开启' if is_multi_spec else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 商品多数量发货管理API
@app.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(cookie_id: str, item_id: str, delivery_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新商品的多数量发货状态"""
    try:
        from db_manager import db_manager

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"商品多数量发货状态已{'开启' if multi_quantity_delivery else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





# ==================== 订单管理接口 ====================

class OrderHistorySyncRequest(BaseModel):
    cookie_id: Optional[str] = None
    start_date: str
    end_date: str
    max_orders: int = 120
    fetch_details: bool = True


def _normalize_history_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_history_amount_text(value: Any) -> Optional[str]:
    text = _normalize_history_optional_text(value)
    if not text:
        return None
    return text if parse_order_amount_value(text) is not None else None


def _create_order_history_sync_job_snapshot(job: Dict[str, Any]) -> Dict[str, Any]:
    return {
        'job_id': job.get('job_id'),
        'status': job.get('status'),
        'message': job.get('message'),
        'error': job.get('error'),
        'created_at': job.get('created_at'),
        'started_at': job.get('started_at'),
        'finished_at': job.get('finished_at'),
        'request': job.get('request'),
        'current_account': job.get('current_account'),
        'current_order_id': job.get('current_order_id'),
        'accounts_total': job.get('accounts_total', 0),
        'accounts_completed': job.get('accounts_completed', 0),
        'orders_discovered': job.get('orders_discovered', 0),
        'orders_processed': job.get('orders_processed', 0),
        'orders_saved': job.get('orders_saved', 0),
        'orders_skipped': job.get('orders_skipped', 0),
        'orders_failed': job.get('orders_failed', 0),
        'matched_orders': job.get('matched_orders', 0),
        'warnings': list(job.get('warnings') or []),
    }


def _append_order_history_sync_warning(job: Dict[str, Any], message: str) -> None:
    warnings = job.setdefault('warnings', [])
    if len(warnings) >= 20:
        return
    warnings.append(str(message))


def _cleanup_order_history_sync_jobs() -> None:
    now_ts = time.time()
    expired_job_ids = []
    for job_id, job in order_history_sync_jobs.items():
        status_value = str(job.get('status') or '')
        finished_ts = job.get('finished_ts') or 0
        if status_value in {'completed', 'failed', 'cancelled'} and finished_ts and (now_ts - finished_ts) > ORDER_HISTORY_SYNC_JOB_RETENTION_SECONDS:
            expired_job_ids.append(job_id)

    for job_id in expired_job_ids:
        order_history_sync_jobs.pop(job_id, None)
        order_history_sync_tasks.pop(job_id, None)


def _save_history_order_candidate(cookie_id: str, candidate: Dict[str, Any]) -> bool:
    order_status = _normalize_history_optional_text(candidate.get('order_status'))
    normalized_status = normalize_order_status_value(order_status) if order_status else None

    return db_manager.insert_or_update_order(
        order_id=str(candidate.get('order_id') or '').strip(),
        item_id=_normalize_history_optional_text(candidate.get('item_id')),
        buyer_id=_normalize_history_optional_text(candidate.get('buyer_id')),
        buyer_nick=_normalize_history_optional_text(candidate.get('buyer_nick')),
        sid=_normalize_history_optional_text(candidate.get('sid')),
        amount=_normalize_history_amount_text(candidate.get('amount')),
        order_status=normalized_status,
        cookie_id=cookie_id,
        platform_created_at=_normalize_history_optional_text(candidate.get('platform_created_at')),
        platform_paid_at=_normalize_history_optional_text(candidate.get('platform_paid_at')),
        platform_completed_at=_normalize_history_optional_text(candidate.get('platform_completed_at')),
    )


def _save_history_order_detail_result(cookie_id: str, candidate: Dict[str, Any], result: Dict[str, Any]) -> bool:
    order_id = _normalize_history_optional_text(result.get('order_id')) or _normalize_history_optional_text(candidate.get('order_id'))
    if not order_id:
        return False

    raw_status = _normalize_history_optional_text(result.get('order_status'))
    normalized_status = normalize_order_status_value(raw_status) if raw_status and raw_status.lower() != 'unknown' else None

    return db_manager.insert_or_update_order(
        order_id=order_id,
        item_id=_normalize_history_optional_text(result.get('item_id')) or _normalize_history_optional_text(candidate.get('item_id')),
        buyer_id=_normalize_history_optional_text(candidate.get('buyer_id')),
        buyer_nick=_normalize_history_optional_text(candidate.get('buyer_nick')),
        sid=_normalize_history_optional_text(candidate.get('sid')),
        spec_name=_normalize_history_optional_text(result.get('spec_name')),
        spec_value=_normalize_history_optional_text(result.get('spec_value')),
        spec_name_2=_normalize_history_optional_text(result.get('spec_name_2')),
        spec_value_2=_normalize_history_optional_text(result.get('spec_value_2')),
        quantity=_normalize_history_optional_text(result.get('quantity')),
        amount=_normalize_history_amount_text(result.get('amount')) or _normalize_history_amount_text(candidate.get('amount')),
        order_status=normalized_status,
        cookie_id=cookie_id,
        platform_created_at=_normalize_history_optional_text(result.get('platform_created_at')) or _normalize_history_optional_text(candidate.get('platform_created_at')),
        platform_paid_at=_normalize_history_optional_text(result.get('platform_paid_at')) or _normalize_history_optional_text(candidate.get('platform_paid_at')),
        platform_completed_at=_normalize_history_optional_text(result.get('platform_completed_at')) or _normalize_history_optional_text(candidate.get('platform_completed_at')),
    )


async def _run_order_history_sync_job(job_id: str) -> None:
    job = order_history_sync_jobs.get(job_id)
    if not job:
        return

    request_data = dict(job.get('request') or {})
    user_info = dict(job.get('user_info') or {})
    current_user_id = user_info.get('user_id')

    from utils.order_history_sync import OrderHistoryPageFetcher, OrderHistorySyncError

    try:
        utc_start = local_date_to_utc_start(request_data.get('start_date'))
        utc_end_exclusive = local_date_to_utc_end_exclusive(request_data.get('end_date'))
        if not utc_start or not utc_end_exclusive:
            raise ValueError('日期格式错误，应为 YYYY-MM-DD')
        if utc_start >= utc_end_exclusive:
            raise ValueError('开始日期必须早于结束日期')

        max_orders = int(request_data.get('max_orders') or 120)
        max_orders = min(max(max_orders, 1), 500)
        fetch_details = bool(request_data.get('fetch_details', True))

        user_cookies = db_manager.get_all_cookies(current_user_id)
        selected_cookie_id = _normalize_history_optional_text(request_data.get('cookie_id'))
        if selected_cookie_id:
            if selected_cookie_id not in user_cookies:
                raise ValueError('指定账号不存在或无权限访问')
            target_cookie_ids = [selected_cookie_id]
        else:
            target_cookie_ids = list(user_cookies.keys())

        if not target_cookie_ids:
            raise ValueError('当前没有可同步的账号')

        _cleanup_order_history_sync_jobs()

        job.update({
            'status': 'running',
            'message': '开始同步历史订单',
            'error': None,
            'started_at': get_local_now().strftime('%Y-%m-%d %H:%M:%S'),
            'accounts_total': len(target_cookie_ids),
            'accounts_completed': 0,
            'orders_discovered': 0,
            'orders_processed': 0,
            'orders_saved': 0,
            'orders_skipped': 0,
            'orders_failed': 0,
            'matched_orders': 0,
            'warnings': [],
        })

        for account_index, cookie_id in enumerate(target_cookie_ids, start=1):
            if job.get('status') == 'cancelled':
                return

            remaining_limit = max_orders - int(job.get('matched_orders') or 0)
            if remaining_limit <= 0:
                break

            cookie_string = user_cookies.get(cookie_id)
            if not cookie_string:
                _append_order_history_sync_warning(job, f'账号 {cookie_id} 缺少 Cookie，已跳过')
                job['accounts_completed'] = account_index
                continue

            job['current_account'] = cookie_id
            job['current_order_id'] = None
            job['message'] = f'正在抓取账号 {cookie_id} 的历史订单列表'

            history_fetcher = OrderHistoryPageFetcher(cookie_string, cookie_id_for_log=cookie_id, headless=True)
            live_instance = cookie_manager.manager.get_xianyu_instance(cookie_id) if cookie_manager.manager else None

            try:
                try:
                    fetch_result = await history_fetcher.fetch_recent_orders(
                        max_orders=remaining_limit,
                        utc_start=utc_start,
                        utc_end_exclusive=utc_end_exclusive,
                    )
                except OrderHistorySyncError as history_exc:
                    logger.warning(
                        f"历史订单列表同步跳过账号: cookie_id={cookie_id}, "
                        f"kind={history_exc.kind}, error={history_exc}"
                    )
                    warning_message = str(history_exc)
                    if history_exc.guidance:
                        warning_message = f'{warning_message}；处理建议：{history_exc.guidance}'
                    _append_order_history_sync_warning(job, warning_message)
                    job['orders_failed'] += 1
                    job['accounts_completed'] = account_index
                    continue

                candidates = list(fetch_result.get('orders') or [])
                scanned_count = int(fetch_result.get('scanned_count') or 0)
                matched_count = int(fetch_result.get('matched_count') or 0)
                out_of_range_count = int(fetch_result.get('out_of_range_count') or 0)

                job['orders_discovered'] += scanned_count
                job['matched_orders'] += matched_count
                job['orders_skipped'] += out_of_range_count

                if live_instance is not None:
                    await history_fetcher.close()

                if job.get('status') == 'cancelled':
                    return

                if not candidates:
                    if scanned_count > 0 and out_of_range_count > 0:
                        _append_order_history_sync_warning(job, f'账号 {cookie_id} 未命中时间范围内的历史订单')
                    else:
                        _append_order_history_sync_warning(job, f'账号 {cookie_id} 未抓到历史订单候选')
                    job['accounts_completed'] = account_index
                    continue

                for candidate in candidates:
                    if job.get('status') == 'cancelled':
                        return

                    order_id = _normalize_history_optional_text(candidate.get('order_id'))
                    if not order_id:
                        continue

                    job['current_order_id'] = order_id
                    job['orders_processed'] += 1
                    job['message'] = f'正在同步账号 {cookie_id} 的订单 {order_id}'

                    detail_saved = False
                    detail_result = None

                    if fetch_details:
                        try:
                            if live_instance is not None:
                                detail_result = await live_instance.fetch_order_detail_info(
                                    order_id=order_id,
                                    item_id=_normalize_history_optional_text(candidate.get('item_id')),
                                    buyer_id=_normalize_history_optional_text(candidate.get('buyer_id')),
                                    sid=_normalize_history_optional_text(candidate.get('sid')),
                                    force_refresh=True,
                                    buyer_nick=_normalize_history_optional_text(candidate.get('buyer_nick')),
                                    buyer_id_source='history_sync',
                                )
                                detail_saved = bool(detail_result)
                            else:
                                detail_result = await history_fetcher.fetch_order_detail(order_id, force_refresh=True)
                                if detail_result:
                                    detail_saved = _save_history_order_detail_result(cookie_id, candidate, detail_result)
                        except Exception as sync_exc:
                            logger.warning(f"历史订单详情同步失败: cookie_id={cookie_id}, order_id={order_id}, error={sync_exc}")
                            _append_order_history_sync_warning(job, f'订单 {order_id} 详情刷新失败: {sync_exc}')

                    if not fetch_details or not detail_saved:
                        if _save_history_order_candidate(cookie_id, candidate):
                            detail_saved = True
                        else:
                            _append_order_history_sync_warning(job, f'订单 {order_id} 基础信息写库失败')

                    if detail_saved:
                        job['orders_saved'] += 1
                    else:
                        job['orders_skipped'] += 1
                        job['orders_failed'] += 1

                job['accounts_completed'] = account_index
            finally:
                await history_fetcher.close()

        job['status'] = 'completed'
        job['message'] = (
            f"历史订单同步完成，共扫描 {job.get('orders_discovered', 0)} 单，"
            f"命中时间范围 {job.get('matched_orders', 0)} 单，入库/更新 {job.get('orders_saved', 0)} 单"
        )
    except asyncio.CancelledError:
        logger.info(f"历史订单同步任务已取消: {job_id}")
        job['status'] = 'cancelled'
        job['error'] = None
        job['message'] = job.get('message') or '历史订单同步已取消'
    except Exception as exc:
        logger.error(f"历史订单同步任务失败: {exc}")
        job['status'] = 'failed'
        job['error'] = str(exc)
        job['message'] = f'历史订单同步失败: {exc}'
    finally:
        job['current_order_id'] = None
        job['current_account'] = None
        job['finished_at'] = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
        job['finished_ts'] = time.time()


@app.post('/api/orders/history-sync')
async def start_order_history_sync(request: OrderHistorySyncRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    """按时间范围同步历史订单。"""
    try:
        request_data = request.dict()
        start_date = str(request_data.get('start_date') or '').strip()
        end_date = str(request_data.get('end_date') or '').strip()
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail='开始日期和结束日期不能为空')

        cookie_id = _normalize_history_optional_text(request_data.get('cookie_id'))
        max_orders = min(max(int(request_data.get('max_orders') or 120), 1), 500)
        fetch_details = bool(request_data.get('fetch_details', True))

        _cleanup_order_history_sync_jobs()

        job_id = f"history_sync_{secrets.token_hex(8)}"
        created_at = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
        job = {
            'job_id': job_id,
            'status': 'pending',
            'message': '历史订单同步任务已创建，等待执行',
            'error': None,
            'created_at': created_at,
            'started_at': None,
            'finished_at': None,
            'finished_ts': None,
            'request': {
                'cookie_id': cookie_id,
                'start_date': start_date,
                'end_date': end_date,
                'max_orders': max_orders,
                'fetch_details': fetch_details,
            },
            'user_id': current_user['user_id'],
            'user_info': {
                'user_id': current_user['user_id'],
                'username': current_user.get('username'),
            },
            'current_account': None,
            'current_order_id': None,
            'accounts_total': 0,
            'accounts_completed': 0,
            'orders_discovered': 0,
            'orders_processed': 0,
            'orders_saved': 0,
            'orders_skipped': 0,
            'orders_failed': 0,
            'matched_orders': 0,
            'warnings': [],
        }
        order_history_sync_jobs[job_id] = job

        task = asyncio.create_task(_run_order_history_sync_job(job_id))
        order_history_sync_tasks[job_id] = task

        def _on_task_done(done_task: asyncio.Task) -> None:
            order_history_sync_tasks.pop(job_id, None)
            try:
                done_task.result()
            except asyncio.CancelledError:
                pass
            except Exception as task_exc:
                logger.error(f"历史订单同步后台任务异常: job_id={job_id}, error={task_exc}")

        task.add_done_callback(_on_task_done)

        log_with_user(
            'info',
            f"创建历史订单同步任务: job_id={job_id}, cookie_id={cookie_id or 'ALL'}, range={start_date}~{end_date}, max_orders={max_orders}, fetch_details={fetch_details}",
            current_user
        )
        return {"success": True, "data": _create_order_history_sync_job_snapshot(job)}
    except HTTPException:
        raise
    except Exception as exc:
        log_with_user('error', f"创建历史订单同步任务失败: {exc}", current_user)
        raise HTTPException(status_code=500, detail=f"创建历史订单同步任务失败: {exc}")


@app.get('/api/orders/history-sync/{job_id}')
def get_order_history_sync_status(job_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """查询历史订单同步任务状态。"""
    _cleanup_order_history_sync_jobs()

    job = order_history_sync_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail='历史订单同步任务不存在或已过期')
    if job.get('user_id') != current_user['user_id']:
        raise HTTPException(status_code=403, detail='无权访问该历史订单同步任务')

    return {"success": True, "data": _create_order_history_sync_job_snapshot(job)}


@app.post('/api/orders/history-sync/{job_id}/cancel')
def cancel_order_history_sync(job_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """取消历史订单同步任务。"""
    job = order_history_sync_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail='历史订单同步任务不存在或已过期')
    if job.get('user_id') != current_user['user_id']:
        raise HTTPException(status_code=403, detail='无权取消该历史订单同步任务')

    if str(job.get('status') or '') in {'completed', 'failed', 'cancelled'}:
        return {"success": True, "data": _create_order_history_sync_job_snapshot(job)}

    job['status'] = 'cancelled'
    job['error'] = None
    job['message'] = '历史订单同步已取消'
    job['finished_at'] = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
    job['finished_ts'] = time.time()

    task = order_history_sync_tasks.get(job_id)
    if task and not task.done():
        task.cancel()

    return {"success": True, "data": _create_order_history_sync_job_snapshot(job)}


@app.get('/api/orders')
def get_user_orders(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的订单信息"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', "查询用户订单信息", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 获取所有订单数据
        all_orders = []
        for cookie_id in user_cookies.keys():
            orders = db_manager.get_orders_by_cookie(cookie_id, limit=1000)  # 增加限制数量
            # 为每个订单添加cookie_id信息
            for order in orders:
                order['cookie_id'] = cookie_id
                all_orders.append(order)

        # 历史订单补录后优先按平台下单时间展示，回退到本地入库时间
        all_orders.sort(
            key=lambda x: x.get('platform_created_at') or x.get('created_at') or '',
            reverse=True
        )

        log_with_user('info', f"用户订单查询成功，共 {len(all_orders)} 条记录", current_user)
        return {"success": True, "data": all_orders}

    except Exception as e:
        log_with_user('error', f"查询用户订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"查询订单失败: {str(e)}")


@app.get('/api/orders/stream')
def stream_user_orders(current_user: Dict[str, Any] = Depends(get_current_user)):
    """订单实时事件流，仅在订单页激活时使用。"""
    user_id = current_user['user_id']
    subscriber = order_event_hub.subscribe(user_id)

    def event_generator():
        try:
            yield format_sse_event('stream.ready', {'type': 'stream.ready', 'timestamp': int(time.time() * 1000)})
            while True:
                try:
                    event = subscriber.get(timeout=25)
                    yield format_sse_event(event.get('type', 'message'), event)
                except queue.Empty:
                    yield format_sse_event('ping', {'type': 'ping', 'timestamp': int(time.time() * 1000)})
        finally:
            order_event_hub.unsubscribe(user_id, subscriber)

    return StreamingResponse(
        event_generator(),
        media_type='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        }
    )


@app.get('/api/chat/sessions')
async def get_chat_sessions(
    cookie_id: str = None,
    include_order_fallback: bool = True,
    limit: int = 100,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取指定账号的会话列表"""
    try:
        if not cookie_id:
            raise HTTPException(status_code=400, detail="缺少 cookie_id 参数")
        cookie_id = _ensure_cookie_access(cookie_id, current_user)
        sessions = db_manager.get_chat_sessions(cookie_id, limit=min(limit, 200))
        logger.info(
            f"获取聊天会话列表: cookie_id={cookie_id}, local_sessions={len(sessions)}, include_order_fallback={include_order_fallback}, limit={limit}"
        )
        if include_order_fallback:
            fallback_sessions = _build_chat_sessions_from_recent_orders(cookie_id, limit=min(max(limit, 50), 300))
            logger.info(f"聊天会话列表订单兜底结果: cookie_id={cookie_id}, fallback_sessions={len(fallback_sessions)}")
            sessions = _merge_chat_sessions_with_order_fallback(sessions, fallback_sessions, limit=min(max(limit, 50), 300))
            logger.info(f"聊天会话列表合并结果: cookie_id={cookie_id}, merged_sessions={len(sessions)}")
        sessions = _annotate_chat_sessions(cookie_id, sessions)
        sessions = await _enrich_chat_sessions(cookie_id, sessions, limit=min(max(limit, 20), 30))
        return {'success': True, 'sessions': sessions}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取会话列表失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="获取会话列表失败")


@app.get('/api/chat/messages')
async def get_chat_messages(
    cookie_id: str = None,
    chat_id: str = None,
    limit: int = 50,
    before_id: int = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取指定会话的消息列表（仅读本地 DB，新消息走 /api/chat/stream 实时推送）"""
    try:
        if not cookie_id or not chat_id:
            raise HTTPException(status_code=400, detail="缺少 cookie_id 或 chat_id 参数")
        cookie_id = _ensure_cookie_access(cookie_id, current_user)
        messages = db_manager.get_chat_messages(cookie_id, chat_id, limit=min(limit, 100), before_id=before_id)
        return {'success': True, 'messages': messages}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取聊天消息失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="获取聊天消息失败")


@app.post('/api/chat/send')
async def chat_send_message(
    req: ChatSendRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """在线客服发送消息"""
    try:
        cookie_id = _ensure_cookie_access(req.cookie_id, current_user)

        from XianyuAutoAsync import XianyuLive, ConnectionState
        live_instance = XianyuLive.get_instance(cookie_id)
        if not live_instance:
            raise HTTPException(status_code=400, detail="账号未启动")
        if live_instance.connection_state != ConnectionState.CONNECTED:
            raise HTTPException(status_code=400, detail="账号WebSocket未连接")
        if not live_instance.ws:
            raise HTTPException(status_code=400, detail="WebSocket连接未就绪")

        await _run_live_instance_on_manager_loop(
            cookie_id,
            lambda: live_instance.send_msg(
                live_instance.ws, req.chat_id, req.to_user_id, req.message
            ),
            timeout=15,
        )

        return {'success': True, 'message': '发送成功'}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"客服发送消息失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="发送消息失败")


@app.get('/api/chat/stream')
def stream_chat_messages(current_user: Dict[str, Any] = Depends(get_current_user)):
    """聊天消息实时事件流"""
    user_id = current_user['user_id']
    subscriber = chat_event_hub.subscribe(user_id)

    def event_generator():
        try:
            yield format_sse_event('stream.ready', {'type': 'stream.ready', 'timestamp': int(time.time() * 1000)})
            while True:
                try:
                    event = subscriber.get(timeout=25)
                    yield format_sse_event(event.get('type', 'chat.message'), event)
                except queue.Empty:
                    yield format_sse_event('ping', {'type': 'ping', 'timestamp': int(time.time() * 1000)})
        finally:
            chat_event_hub.unsubscribe(user_id, subscriber)

    return StreamingResponse(
        event_generator(),
        media_type='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        }
    )


@app.get('/api/chat/accounts')
def get_chat_accounts(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有账号列表（在线客服三栏布局用）"""
    try:
        user_cookies = _get_user_cookies_map(current_user)
        accounts = []
        for cid in user_cookies.keys():
            status = _build_live_runtime_status(cid)
            detail = db_manager.get_cookie_details(cid) or {}
            display_name = detail.get('remark') or detail.get('username') or cid
            accounts.append({
                'id': cid,
                'name': display_name,
                'enabled': db_manager.get_cookie_status(cid),
                'connected': status.get('connection_state') == 'connected' if status else False,
            })
        return {'success': True, 'accounts': accounts}
    except Exception as e:
        logger.error(f"获取聊天账号列表失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="获取账号列表失败")


@app.get('/api/chat/keywords/{cid}/item/{item_id}')
def get_item_keywords(
    cid: str, item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取指定商品的关键词列表"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        keywords = db_manager.get_keywords_by_item_id(cid, item_id)
        item_reply_data = db_manager.get_item_reply(cid, item_id)
        item_reply = item_reply_data.get('reply_content') if item_reply_data else None
        return {'success': True, 'keywords': keywords, 'item_reply': item_reply}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取商品关键词失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="获取商品关键词失败")


@app.post('/api/chat/keywords/{cid}/item/{item_id}')
def save_item_keywords(
    cid: str, item_id: str,
    req: SaveItemKeywordsRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """保存指定商品的关键词和指定商品回复"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        success = db_manager.save_keywords_for_item(cid, item_id, req.keywords)
        if req.item_reply is not None:
            reply_content = str(req.item_reply or '').strip()
            if reply_content:
                db_manager.update_item_reply(cid, item_id, reply_content)
            else:
                db_manager.delete_item_reply(cid, item_id)
        return {'success': success, 'count': len(req.keywords)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存商品关键词失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="保存商品关键词失败")


@app.post('/api/chat/keywords/{cid}/copy')
def copy_item_keywords(
    cid: str,
    req: CopyKeywordsRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """复制商品关键词和指定商品回复到其他商品"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        results = {}
        source_reply = db_manager.get_item_reply(cid, req.source_item_id)
        source_reply_content = source_reply.get('reply_content', '') if source_reply else ''

        for target in req.target_item_ids:
            if target == req.source_item_id:
                continue
            count = db_manager.copy_keywords_to_item(cid, req.source_item_id, target)
            results[target] = count
            if source_reply_content:
                db_manager.update_item_reply(cid, target, source_reply_content)

        return {'success': True, 'results': results, 'total': sum(results.values())}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"复制商品关键词失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="复制商品关键词失败")


@app.get('/api/chat/items/{cid}')
def get_account_items(
    cid: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取账号下的商品列表（用于复制回复的目标选择）"""
    try:
        cid = _ensure_cookie_access(cid, current_user)
        cursor = db_manager.conn.cursor()
        db_manager._execute_sql(cursor, """
            SELECT item_id, item_title FROM item_info
            WHERE cookie_id = ? ORDER BY item_id
        """, (cid,))
        rows = cursor.fetchall()
        items = [{'item_id': r[0], 'item_title': r[1]} for r in rows]
        return {'success': True, 'items': items}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取商品列表失败: {mask_sensitive_text(e)}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@app.delete('/api/orders/{order_id}')
def delete_user_order(order_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除当前用户自己的订单"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        order = db_manager.get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在")

        cookie_id = order.get('cookie_id')
        cookie_info = db_manager.get_cookie_details(cookie_id) if cookie_id else None
        if not cookie_info or cookie_info.get('user_id') != user_id:
            raise HTTPException(status_code=403, detail="无权删除此订单")

        success = db_manager.delete_order(order_id, cookie_id=cookie_id)
        if not success:
            raise HTTPException(status_code=400, detail="删除订单失败")

        log_with_user('info', f"删除订单成功: {order_id}", current_user)
        return {"success": True, "message": "订单删除成功"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除订单失败: {order_id} - {mask_sensitive_text(e)}", current_user)
        raise HTTPException(status_code=500, detail="删除订单失败，请稍后重试")


@app.post('/api/orders/{order_id}/deliver')
async def manual_deliver_order(order_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """手动发货 - 根据订单信息匹配发货规则并发送卡券"""
    try:
        from db_manager import db_manager
        import cookie_manager

        user_id = current_user['user_id']
        log_with_user('info', f"手动发货请求: 订单 {order_id}", current_user)

        # 获取订单信息
        order = db_manager.get_order_by_id(order_id)
        if not order:
            return {"success": False, "delivered": False, "message": "订单不存在"}

        # 验证订单属于当前用户
        cookie_id = order.get('cookie_id')
        if not cookie_id:
            return {"success": False, "delivered": False, "message": "订单缺少账号信息"}

        cookie_info = db_manager.get_cookie_details(cookie_id)
        if not cookie_info or cookie_info.get('user_id') != user_id:
            return {"success": False, "delivered": False, "message": "无权操作此订单"}

        # 获取 XianyuLive 实例
        xianyu_instance = cookie_manager.manager.get_xianyu_instance(cookie_id) if cookie_manager.manager else None
        if not xianyu_instance:
            return {"success": False, "delivered": False, "message": f"账号 {cookie_id} 未运行，请先启动账号"}

        # 获取订单详情
        item_id = order.get('item_id')
        buyer_id = order.get('buyer_id')

        if not item_id:
            return {"success": False, "delivered": False, "message": "订单缺少商品信息"}

        if not buyer_id:
            return {"success": False, "delivered": False, "message": "订单缺少买家信息，无法发送消息"}

        # 获取商品标题
        item_info = db_manager.get_item_info(cookie_id, item_id)
        item_title = item_info.get('item_title', '') if item_info else ''

        try:
            expected_quantity = max(1, int(order.get('quantity') or 1))
        except (TypeError, ValueError):
            expected_quantity = 1

        progress_summary_before = xianyu_instance._summarize_delivery_progress(order_id, expected_quantity)
        pending_finalize_units = list(progress_summary_before.get('pending_finalize_unit_indexes') or [])
        finalize_completed_units = 0
        for unit_index in pending_finalize_units:
            pending_finalize_meta = xianyu_instance._get_pending_delivery_finalization_meta(order_id, unit_index)
            if not pending_finalize_meta:
                continue

            finalize_result = await xianyu_instance._finalize_delivery_after_send(
                delivery_meta=pending_finalize_meta,
                order_id=order_id,
                item_id=item_id
            )
            if not finalize_result.get('success'):
                xianyu_instance._persist_delivery_finalization_state(
                    order_id=order_id,
                    item_id=item_id,
                    buyer_id=buyer_id,
                    delivery_meta=pending_finalize_meta,
                    channel='manual',
                    status='sent',
                    last_error=finalize_result.get('error') or f'检测到第 {unit_index} 个发货单元已发送记录，但补完成收尾失败'
                )
                return {"success": False, "delivered": False, "message": finalize_result.get('error') or f'检测到第 {unit_index} 个发货单元已发送记录，但补完成收尾失败'}

            xianyu_instance._persist_delivery_finalization_state(
                order_id=order_id,
                item_id=item_id,
                buyer_id=buyer_id,
                delivery_meta=pending_finalize_meta,
                channel='manual',
                status='finalized'
            )
            finalize_completed_units += 1

        if finalize_completed_units > 0:
            progress_after_finalize = xianyu_instance._sync_order_delivery_progress(
                order_id=order_id,
                cookie_id=cookie_id,
                expected_quantity=expected_quantity,
                context="手动发货补完成收尾成功"
            )
            publish_order_update_event(order_id, source='manual_delivery_finalize')
            log_with_user('info', f"检测到订单 {order_id} 存在待完成收尾记录，已先补完成 {finalize_completed_units} 个单元，继续执行补发", current_user)
        else:
            progress_after_finalize = progress_summary_before

        remaining_unit_indexes = list(progress_after_finalize.get('remaining_unit_indexes') or [])
        if not remaining_unit_indexes:
            aggregate_status = progress_after_finalize.get('aggregate_status')
            if aggregate_status == 'shipped':
                return {"success": True, "delivered": True, "message": "订单所有发货单元都已完成，本次仅补完成未收尾记录"}
            return {"success": True, "delivered": True, "message": "订单当前没有可补发的未完成单元"}

        unit_results = []
        prepared_units = []

        def format_delivery_reason(reason: str, order_spec_mode: str = None, rule_spec_mode: str = None, item_config_mode: str = None) -> str:
            context_parts = []
            if order_spec_mode:
                context_parts.append(f"order_spec_mode={order_spec_mode}")
            if rule_spec_mode:
                context_parts.append(f"rule_spec_mode={rule_spec_mode}")
            if item_config_mode:
                context_parts.append(f"item_config_mode={item_config_mode}")

            if not context_parts:
                return reason

            reason_text = (reason or '').strip() or '未提供发货日志原因'
            if any(part.split('=')[0] + '=' in reason_text for part in context_parts):
                return reason_text
            return f"{reason_text} [{', '.join(context_parts)}]"

        for unit_index in remaining_unit_indexes:
            delivery_result = await xianyu_instance._auto_delivery(
                item_id=item_id,
                item_title=item_title,
                order_id=order_id,
                send_user_id=buyer_id,
                include_meta=True,
                delivery_unit_index=unit_index
            )

            if isinstance(delivery_result, dict):
                delivery_content = delivery_result.get('content')
                delivery_steps = delivery_result.get('delivery_steps') or []
                delivery_success = bool(delivery_result.get('success') and delivery_content)
                rule_id = delivery_result.get('rule_id')
                rule_keyword = delivery_result.get('rule_keyword')
                card_type = delivery_result.get('card_type')
                card_id = delivery_result.get('card_id')
                match_mode = delivery_result.get('match_mode')
                order_spec_mode = delivery_result.get('order_spec_mode')
                rule_spec_mode = delivery_result.get('rule_spec_mode')
                item_config_mode = delivery_result.get('item_config_mode')
                data_card_pending_consume = delivery_result.get('data_card_pending_consume')
                data_line = delivery_result.get('data_line')
                data_reservation_id = delivery_result.get('data_reservation_id')
                data_reservation_status = delivery_result.get('data_reservation_status')
                failure_reason = delivery_result.get('error')
            else:
                delivery_content = delivery_result
                delivery_steps = []
                delivery_success = bool(delivery_content)
                rule_id = None
                rule_keyword = None
                card_type = None
                card_id = None
                match_mode = None
                order_spec_mode = None
                rule_spec_mode = None
                item_config_mode = None
                data_card_pending_consume = None
                data_line = None
                data_reservation_id = None
                data_reservation_status = None
                failure_reason = None

            if delivery_success:
                if not delivery_steps:
                    delivery_steps = xianyu_instance._build_delivery_steps(delivery_content, '')
                if not delivery_steps:
                    fail_reason = f"第 {unit_index} 个发货单元发货步骤构建失败"
                    xianyu_instance._release_data_reservation_if_needed(
                        {'data_reservation_id': data_reservation_id},
                        error=fail_reason
                    )
                    db_manager.create_delivery_log(
                        user_id=user_id,
                        cookie_id=cookie_id,
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        buyer_nick=order.get('buyer_nick'),
                        rule_id=rule_id,
                        rule_keyword=rule_keyword,
                        card_type=card_type,
                        match_mode=match_mode,
                        channel='manual',
                        status='failed',
                        reason=format_delivery_reason(fail_reason, order_spec_mode, rule_spec_mode, item_config_mode)
                    )
                    unit_results.append({'unit_index': unit_index, 'status': 'failed', 'error': fail_reason})
                    continue

                prepared_units.append({
                    'unit_index': unit_index,
                    'delivery_steps': delivery_steps,
                    'card_type': card_type,
                    'rule_meta': {
                        'success': True,
                        'rule_id': rule_id,
                        'rule_keyword': rule_keyword,
                        'card_id': card_id,
                        'card_type': card_type,
                        'match_mode': match_mode,
                        'order_spec_mode': order_spec_mode,
                        'rule_spec_mode': rule_spec_mode,
                        'item_config_mode': item_config_mode,
                        'data_card_pending_consume': data_card_pending_consume,
                        'data_line': data_line,
                        'data_reservation_id': data_reservation_id,
                        'data_reservation_status': data_reservation_status,
                        'delivery_unit_index': unit_index,
                    }
                })
            else:
                fail_reason = failure_reason or f"第 {unit_index} 个发货单元未匹配到发货规则，请检查卡券和发货规则配置"
                db_manager.create_delivery_log(
                    user_id=user_id,
                    cookie_id=cookie_id,
                    order_id=order_id,
                    item_id=item_id,
                    buyer_id=buyer_id,
                    buyer_nick=order.get('buyer_nick'),
                    rule_id=rule_id,
                    rule_keyword=rule_keyword,
                    card_type=card_type,
                    match_mode=match_mode,
                    channel='manual',
                    status='failed',
                    reason=format_delivery_reason(fail_reason, order_spec_mode, rule_spec_mode, item_config_mode)
                )
                unit_results.append({'unit_index': unit_index, 'status': 'failed', 'error': fail_reason})

        ws = getattr(xianyu_instance, 'ws', None)
        manual_chat_id = buyer_id
        if ws:
            sid = order.get('sid', '')
            if sid:
                manual_chat_id = sid.replace('@goofish', '')
                log_with_user('info', f"手动发货: 使用现有WebSocket连接发送, cid={manual_chat_id}, buyer_id={buyer_id}", current_user)
            else:
                log_with_user('warning', f"手动发货: 订单无sid，尝试使用buyer_id作为cid, buyer_id={buyer_id}", current_user)
        else:
            log_with_user('warning', f"手动发货: 无现有WebSocket连接，使用send_delivery_steps_once, buyer_id={buyer_id}", current_user)

        send_groups = xianyu_instance._build_delivery_send_groups(prepared_units, expected_quantity)
        total_send_groups = len(send_groups)

        for group_index, send_group in enumerate(send_groups, start=1):
            group_units = send_group.get('units') or []
            if not group_units:
                continue

            first_unit = group_units[0]
            first_unit_index = first_unit.get('unit_index') or 1
            is_batched_text_group = send_group.get('mode') == 'batched_text'

            try:
                if ws:
                    await xianyu_instance._send_delivery_steps(
                        ws,
                        manual_chat_id,
                        buyer_id,
                        send_group.get('delivery_steps') or [],
                        log_prefix=(
                            f"手动发货 order_id={order_id} batch={group_index}/{total_send_groups}"
                            if is_batched_text_group else
                            f"手动发货 order_id={order_id} unit={first_unit_index}"
                        )
                    )
                else:
                    await xianyu_instance.send_delivery_steps_once(buyer_id, item_id, send_group.get('delivery_steps') or [])
            except Exception as send_error:
                send_error_text = str(send_error)
                for prepared_unit in group_units:
                    unit_index = prepared_unit.get('unit_index') or 1
                    rule_meta = prepared_unit.get('rule_meta') or {}
                    xianyu_instance._release_data_reservation_if_needed(
                        rule_meta,
                        error=f"手动发货发送失败(unit={unit_index}): {send_error_text}"
                    )
                    db_manager.create_delivery_log(
                        user_id=user_id,
                        cookie_id=cookie_id,
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        buyer_nick=order.get('buyer_nick'),
                        rule_id=rule_meta.get('rule_id'),
                        rule_keyword=rule_meta.get('rule_keyword'),
                        card_type=rule_meta.get('card_type'),
                        match_mode=rule_meta.get('match_mode'),
                        channel='manual',
                        status='failed',
                        reason=format_delivery_reason(f"第 {unit_index} 个发货单元消息发送失败: {send_error_text}", rule_meta.get('order_spec_mode'), rule_meta.get('rule_spec_mode'), rule_meta.get('item_config_mode'))
                    )
                    unit_results.append({'unit_index': unit_index, 'status': 'failed', 'error': send_error_text})
                continue

            for prepared_unit in group_units:
                unit_index = prepared_unit.get('unit_index') or 1
                rule_meta = prepared_unit.get('rule_meta') or {}

                try:
                    if not xianyu_instance._mark_data_reservation_sent_if_needed(rule_meta):
                        xianyu_instance._release_data_reservation_if_needed(
                            rule_meta,
                            error=f'手动发货发送成功后标记预占已发送失败(unit={unit_index})'
                        )
                        db_manager.create_delivery_log(
                            user_id=user_id,
                            cookie_id=cookie_id,
                            order_id=order_id,
                            item_id=item_id,
                            buyer_id=buyer_id,
                            buyer_nick=order.get('buyer_nick'),
                            rule_id=rule_meta.get('rule_id'),
                            rule_keyword=rule_meta.get('rule_keyword'),
                            card_type=rule_meta.get('card_type'),
                            match_mode=rule_meta.get('match_mode'),
                            channel='manual',
                            status='failed',
                            reason=format_delivery_reason('批量数据预占标记已发送失败', rule_meta.get('order_spec_mode'), rule_meta.get('rule_spec_mode'), rule_meta.get('item_config_mode'))
                        )
                        unit_results.append({'unit_index': unit_index, 'status': 'failed', 'error': '批量数据预占标记已发送失败'})
                        continue

                    xianyu_instance._persist_delivery_finalization_state(
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        delivery_meta=rule_meta,
                        channel='manual',
                        status='sent'
                    )

                    finalize_result = await xianyu_instance._finalize_delivery_after_send(
                        delivery_meta=rule_meta,
                        order_id=order_id,
                        item_id=item_id
                    )
                    if not finalize_result.get('success'):
                        xianyu_instance._persist_delivery_finalization_state(
                            order_id=order_id,
                            item_id=item_id,
                            buyer_id=buyer_id,
                            delivery_meta=rule_meta,
                            channel='manual',
                            status='sent',
                            last_error=finalize_result.get('error') or f'第 {unit_index} 个发货单元发送成功但提交发货副作用失败'
                        )
                        db_manager.create_delivery_log(
                            user_id=user_id,
                            cookie_id=cookie_id,
                            order_id=order_id,
                            item_id=item_id,
                            buyer_id=buyer_id,
                            buyer_nick=order.get('buyer_nick'),
                            rule_id=rule_meta.get('rule_id'),
                            rule_keyword=rule_meta.get('rule_keyword'),
                            card_type=rule_meta.get('card_type'),
                            match_mode=rule_meta.get('match_mode'),
                            channel='manual',
                            status='failed',
                            reason=format_delivery_reason(finalize_result.get('error') or f'第 {unit_index} 个发货单元发送成功但提交发货副作用失败', rule_meta.get('order_spec_mode'), rule_meta.get('rule_spec_mode'), rule_meta.get('item_config_mode'))
                        )
                        unit_results.append({'unit_index': unit_index, 'status': 'pending_finalize', 'error': finalize_result.get('error') or '发送成功但提交发货副作用失败'})
                        continue

                    xianyu_instance._persist_delivery_finalization_state(
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        delivery_meta=rule_meta,
                        channel='manual',
                        status='finalized'
                    )
                    success_reason = f'手动发货第 {unit_index} 个单元发送成功'
                    if is_batched_text_group and len(group_units) > 1:
                        success_reason += '（批量合并发送）'
                    db_manager.create_delivery_log(
                        user_id=user_id,
                        cookie_id=cookie_id,
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        buyer_nick=order.get('buyer_nick'),
                        rule_id=rule_meta.get('rule_id'),
                        rule_keyword=rule_meta.get('rule_keyword'),
                        card_type=rule_meta.get('card_type'),
                        match_mode=rule_meta.get('match_mode'),
                        channel='manual',
                        status='success',
                        reason=format_delivery_reason(success_reason, rule_meta.get('order_spec_mode'), rule_meta.get('rule_spec_mode'), rule_meta.get('item_config_mode'))
                    )
                    unit_results.append({'unit_index': unit_index, 'status': 'finalized'})

                except Exception as unit_post_error:
                    unit_error_text = str(unit_post_error)
                    xianyu_instance._persist_delivery_finalization_state(
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        delivery_meta=rule_meta,
                        channel='manual',
                        status='sent',
                        last_error=f'第 {unit_index} 个发货单元消息已发送，但发送后处理异常: {unit_error_text}'
                    )
                    db_manager.create_delivery_log(
                        user_id=user_id,
                        cookie_id=cookie_id,
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        buyer_nick=order.get('buyer_nick'),
                        rule_id=rule_meta.get('rule_id'),
                        rule_keyword=rule_meta.get('rule_keyword'),
                        card_type=rule_meta.get('card_type'),
                        match_mode=rule_meta.get('match_mode'),
                        channel='manual',
                        status='failed',
                        reason=format_delivery_reason(f"第 {unit_index} 个发货单元消息已发送，但发送后处理异常: {unit_error_text}", rule_meta.get('order_spec_mode'), rule_meta.get('rule_spec_mode'), rule_meta.get('item_config_mode'))
                    )
                    unit_results.append({'unit_index': unit_index, 'status': 'pending_finalize', 'error': unit_error_text})

        progress_summary_after = xianyu_instance._sync_order_delivery_progress(
            order_id=order_id,
            cookie_id=cookie_id,
            expected_quantity=expected_quantity,
            context="手动发货发送成功"
        )
        publish_order_update_event(order_id, source='manual_delivery')

        finalized_now = [r for r in unit_results if r.get('status') == 'finalized']
        pending_finalize_now = [r for r in unit_results if r.get('status') == 'pending_finalize']
        failed_now = [r for r in unit_results if r.get('status') == 'failed']

        message_parts = []
        if finalize_completed_units > 0:
            message_parts.append(f"已补完成 {finalize_completed_units} 个未收尾单元")
        if finalized_now:
            message_parts.append(f"本次补发成功 {len(finalized_now)} 个单元")
        if pending_finalize_now:
            message_parts.append(f"仍有 {len(pending_finalize_now)} 个单元待收尾")
        if failed_now:
            message_parts.append(f"仍有 {len(failed_now)} 个单元补发失败")

        aggregate_status = progress_summary_after.get('aggregate_status')
        if aggregate_status == 'shipped':
            message_parts.append(f"订单已全部完成（{progress_summary_after.get('finalized_count', 0)}/{expected_quantity}）")
        elif aggregate_status == 'partial_pending_finalize':
            message_parts.append(
                f"订单当前为部分待收尾（已完成 {progress_summary_after.get('finalized_count', 0)}/{expected_quantity}，待收尾 {progress_summary_after.get('pending_finalize_count', 0)}）"
            )
        elif aggregate_status == 'partial_success':
            message_parts.append(
                f"订单当前为部分发货（已完成 {progress_summary_after.get('finalized_count', 0)}/{expected_quantity}，待补发 {progress_summary_after.get('remaining_count', 0)}）"
            )

        delivered = bool(finalized_now or finalize_completed_units > 0)
        if not message_parts:
            message_parts.append("订单当前没有可推进的发货单元")

        return {"success": True, "delivered": delivered, "message": '，'.join(message_parts)}

    except Exception as e:
        log_with_user('error', f"手动发货异常: 订单 {order_id} - {str(e)}", current_user)
        import traceback
        logger.error(f"手动发货异常堆栈: {traceback.format_exc()}")
        return {"success": False, "delivered": False, "message": f"发货失败: {str(e)}"}


@app.post('/api/orders/{order_id}/refresh')
async def refresh_order_status(order_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """刷新订单状态 - 从闲鱼平台获取最新订单状态"""
    try:
        from db_manager import db_manager
        import cookie_manager

        user_id = current_user['user_id']
        log_with_user('info', f"刷新订单状态请求: 订单 {order_id}", current_user)

        # 获取订单信息
        order = db_manager.get_order_by_id(order_id)
        if not order:
            return {"success": False, "updated": False, "message": "订单不存在"}

        old_status = order.get('order_status', '')

        # 验证订单属于当前用户
        cookie_id = order.get('cookie_id')
        if not cookie_id:
            return {"success": False, "updated": False, "message": "订单缺少账号信息"}

        cookie_info = db_manager.get_cookie_details(cookie_id)
        if not cookie_info or cookie_info.get('user_id') != user_id:
            return {"success": False, "updated": False, "message": "无权操作此订单"}

        # 获取 XianyuLive 实例
        xianyu_instance = cookie_manager.manager.get_xianyu_instance(cookie_id) if cookie_manager.manager else None
        if not xianyu_instance:
            return {"success": False, "updated": False, "message": f"账号 {cookie_id} 未运行，请先启动账号"}

        # 获取订单详情（强制从闲鱼平台获取最新信息，跳过缓存）
        item_id = order.get('item_id')
        buyer_id = order.get('buyer_id')
        sid = order.get('sid')

        result = await xianyu_instance.fetch_order_detail_info(
            order_id=order_id,
            item_id=item_id,
            buyer_id=buyer_id,
            sid=sid,
            force_refresh=True  # 强制刷新，跳过缓存
        )

        if result:
            # 获取更新后的订单信息
            updated_order = db_manager.get_order_by_id(order_id)
            new_status = updated_order.get('order_status', '') if updated_order else ''
            status_changed = old_status != new_status
            log_with_user('info', f"刷新订单状态成功: 订单 {order_id}, 状态: {old_status} -> {new_status}", current_user)
            return {
                "success": True,
                "updated": status_changed,
                "new_status": new_status,
                "message": f"状态已更新: {new_status}" if status_changed else "订单状态无变化"
            }
        else:
            log_with_user('warning', f"刷新订单状态失败: 订单 {order_id}", current_user)
            return {"success": False, "updated": False, "message": "获取订单详情失败，请稍后重试"}

    except Exception as e:
        log_with_user('error', f"刷新订单状态异常: 订单 {order_id} - {str(e)}", current_user)
        import traceback
        logger.error(f"刷新订单状态异常堆栈: {traceback.format_exc()}")
        return {"success": False, "updated": False, "message": f"刷新失败: {str(e)}"}


# ==================== 自动更新接口 ====================

from auto_updater import get_updater, UpdateStatus, init_updater
from pydantic import BaseModel as PydanticBaseModel

class UpdateCheckResponse(PydanticBaseModel):
    """更新检查响应"""
    has_update: bool
    current_version: str
    new_version: str = ""
    description: str = ""
    changelog: list = []
    files_count: int = 0
    total_size: int = 0
    release_date: str = ""


class UpdateProgressResponse(PydanticBaseModel):
    """更新进度响应"""
    status: str
    current_file: str = ""
    current_index: int = 0
    total_files: int = 0
    downloaded_bytes: int = 0
    total_bytes: int = 0
    message: str = ""
    error: str = ""


class UpdateResultResponse(PydanticBaseModel):
    """更新结果响应"""
    success: bool
    message: str
    updated_files: list = []
    deleted_files: list = []
    needs_restart: bool = False
    new_version: str = ""


@app.get('/api/update/check')
async def check_for_updates(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    检查是否有可用更新
    
    返回更新信息，包括新版本号、更新内容等
    """
    try:
        updater = get_updater()
        manifest = await updater.check_for_updates()
        
        if manifest is None:
            return {
                "success": True,
                "data": {
                    "has_update": False,
                    "current_version": updater.current_version,
                    "message": "已是最新版本"
                }
            }
        
        # 获取需要更新的文件
        files_to_update = await updater.get_files_to_update(manifest)
        files_to_delete = await updater.get_files_to_delete(manifest)
        total_size = sum(f.size for f in files_to_update)

        if not files_to_update and not files_to_delete:
            return {
                "success": True,
                "data": {
                    "has_update": False,
                    "current_version": updater.current_version,
                    "message": "已是最新版本"
                }
            }
        
        return {
            "success": True,
            "data": {
                "has_update": True,
                "current_version": updater.current_version,
                "new_version": manifest.version,
                "description": manifest.description,
                "changelog": manifest.changelog or [],
                "files_count": len(files_to_update),
                "deleted_files_count": len(files_to_delete),
                "total_size": total_size,
                "release_date": manifest.release_date,
                "files": [
                    {
                        "path": f.path,
                        "size": f.size,
                        "requires_restart": f.requires_restart,
                        "description": f.description
                    }
                    for f in files_to_update
                ],
                "deleted_files": [
                    {
                        "path": f.path,
                        "requires_restart": f.requires_restart,
                        "description": f.description
                    }
                    for f in files_to_delete
                ]
            }
        }
        
    except Exception as e:
        logger.error(f"检查更新失败: {e}")
        return {
            "success": False,
            "message": f"检查更新失败: {str(e)}"
        }


@app.post('/api/update/apply')
async def apply_updates(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    应用更新
    
    下载并安装所有可用更新
    """
    try:
        # 只允许管理员执行更新，兼容历史 admin 用户名判断
        if not current_user.get('is_admin') and current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以执行更新")
        
        updater = get_updater()
        
        log_with_user('info', "开始执行自动更新", current_user)
        
        result = await updater.perform_update()
        
        if result["success"]:
            log_with_user('info', f"更新完成: {result['message']}", current_user)
        else:
            log_with_user('error', f"更新失败: {result['message']}", current_user)
        
        return {
            "success": result["success"],
            "data": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"应用更新失败: {e}")
        return {
            "success": False,
            "message": f"应用更新失败: {str(e)}"
        }


@app.get('/api/update/progress')
async def get_update_progress(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取更新进度
    
    返回当前更新状态和进度信息
    """
    try:
        updater = get_updater()
        progress = updater.progress
        
        return {
            "success": True,
            "data": {
                "status": progress.status.value,
                "current_file": progress.current_file,
                "current_index": progress.current_index,
                "total_files": progress.total_files,
                "downloaded_bytes": progress.downloaded_bytes,
                "total_bytes": progress.total_bytes,
                "message": progress.message,
                "error": progress.error
            }
        }
        
    except Exception as e:
        logger.error(f"获取更新进度失败: {e}")
        return {
            "success": False,
            "message": f"获取更新进度失败: {str(e)}"
        }


@app.get('/api/update/local-hashes')
async def get_local_file_hashes(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取本地文件哈希值
    
    用于服务端比对哪些文件需要更新
    """
    try:
        # 只允许管理员查看（检查username是否为admin）
        if current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以查看文件哈希")
        
        updater = get_updater()
        hashes = updater.get_local_file_hashes()
        
        return {
            "success": True,
            "data": {
                "version": updater.current_version,
                "files": hashes,
                "count": len(hashes)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取文件哈希失败: {e}")
        return {
            "success": False,
            "message": f"获取文件哈希失败: {str(e)}"
        }


@app.post('/api/update/cleanup-backups')
async def cleanup_old_backups(days: int = 7, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    清理旧的备份文件
    
    Args:
        days: 保留天数，默认7天
    """
    try:
        # 只允许管理员执行（检查username是否为admin）
        if current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以清理备份")
        
        updater = get_updater()
        updater.cleanup_old_backups(keep_days=days)
        
        log_with_user('info', f"清理了 {days} 天前的备份文件", current_user)
        
        return {
            "success": True,
            "message": f"已清理 {days} 天前的备份文件"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"清理备份失败: {e}")
        return {
            "success": False,
            "message": f"清理备份失败: {str(e)}"
        }


@app.get('/api/update/file-changes')
async def get_file_changes(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    比较当前文件与上次更新后的哈希清单
    
    用于检测哪些文件在更新后被本地修改过
    """
    try:
        # 只允许管理员查看
        if current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以查看文件变化")
        
        updater = get_updater()
        result = updater.compare_file_hashes()
        
        return {
            "success": True,
            "data": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"比较文件变化失败: {e}")
        return {
            "success": False,
            "message": f"比较文件变化失败: {str(e)}"
        }


@app.post('/api/update/save-hashes')
async def save_current_hashes(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    手动保存当前文件的哈希清单
    
    用于记录当前状态，以便以后比较
    """
    try:
        # 只允许管理员执行
        if current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以保存哈希清单")
        
        updater = get_updater()
        updater.save_file_hashes(updater.current_version)
        
        log_with_user('info', "手动保存文件哈希清单", current_user)
        
        return {
            "success": True,
            "message": "文件哈希清单已保存"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存哈希清单失败: {e}")
        return {
            "success": False,
            "message": f"保存哈希清单失败: {str(e)}"
        }


@app.get('/api/update/saved-hashes')
async def get_saved_hashes(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取上次保存的文件哈希清单
    """
    try:
        # 只允许管理员查看
        if current_user.get('username') != 'admin':
            raise HTTPException(status_code=403, detail="只有管理员可以查看哈希清单")
        
        updater = get_updater()
        saved_hashes = updater.load_file_hashes()
        
        if saved_hashes is None:
            return {
                "success": True,
                "data": None,
                "message": "没有保存的哈希清单"
            }
        
        return {
            "success": True,
            "data": {
                "version": saved_hashes.get("version"),
                "updated_at": saved_hashes.get("updated_at"),
                "total_files": saved_hashes.get("total_files"),
                "last_updated_files": saved_hashes.get("last_updated_files", []),
                "last_updated_count": saved_hashes.get("last_updated_count", 0)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取哈希清单失败: {e}")
        return {
            "success": False,
            "message": f"获取哈希清单失败: {str(e)}"
        }


@app.post('/api/update/restart')
async def restart_application(current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    重启应用（用于更新后重启）
    
    注意：此操作会重启整个应用
    """
    try:
        # 只允许管理员执行
        if not current_user.get('is_admin'):
            raise HTTPException(status_code=403, detail="只有管理员可以重启应用")
        
        log_with_user('info', "用户请求重启应用", current_user)
        
        import subprocess
        import sys
        
        # 返回响应后异步重启
        async def delayed_restart():
            await asyncio.sleep(2)  # 等待2秒让响应返回
            logger.info("正在重启应用...")
            
            # 获取当前Python解释器和脚本路径
            python = sys.executable
            script = sys.argv[0]
            
            # 在Windows上使用start命令启动新进程
            if sys.platform == 'win32':
                subprocess.Popen(
                    [python, script],
                    creationflags=subprocess.CREATE_NEW_CONSOLE
                )
            else:
                # Linux/Mac
                subprocess.Popen([python, script])
            
            # 退出当前进程
            os._exit(0)
        
        # 创建后台任务
        asyncio.create_task(delayed_restart())
        
        return {
            "success": True,
            "message": "应用将在2秒后重启"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"重启应用失败: {e}")
        return {
            "success": False,
            "message": f"重启应用失败: {str(e)}"
        }


# ==================== 一键擦亮API ====================

@app.post("/accounts/{cid}/polish-items")
async def polish_account_items(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """擦亮指定账号的所有在售商品"""
    try:
        cookie_info = db_manager.get_cookie_by_id(cid)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cid, register_instance=False)

        logger.info(f"开始擦亮账号 {cid} 的所有商品")
        result = await xianyu_instance.polish_all_items()

        await xianyu_instance.close_session()

        return result

    except Exception as e:
        logger.error(f"擦亮账号商品异常: {str(e)}")
        return {"success": False, "message": f"擦亮异常: {str(e)}"}


# ==================== 定时任务管理API ====================

def _parse_enabled_flag(value):
    """将不同类型的 enabled 入参统一转换为 0/1"""
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if int(value) else 0
    if isinstance(value, str):
        return 1 if value.strip().lower() in {'1', 'true', 'yes', 'on'} else 0
    return 1 if value else 0


def _parse_run_hour(value, default=8):
    run_hour = default if value is None else int(value)
    if run_hour < 0 or run_hour > 23:
        raise ValueError("运行时间必须在 0-23 之间")
    return run_hour


def _parse_random_delay(value, default=10):
    random_delay_max = default if value is None else int(value)
    if random_delay_max < 0:
        raise ValueError("随机分钟不能小于 0")
    return random_delay_max

@app.post("/scheduled-tasks")
async def create_scheduled_task(request: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建定时任务"""
    try:
        account_id = request.get('account_id', '').strip()
        run_hour = _parse_run_hour(request.get('run_hour', request.get('delay_minutes', 8)))
        random_delay_max = _parse_random_delay(request.get('random_delay_max', 10), 10)
        enabled = _parse_enabled_flag(request.get('enabled', True))

        if not account_id:
            return {"success": False, "message": "账号ID不能为空"}

        name = f"每日擦亮-{account_id}"
        next_run_at = db_manager.calculate_next_daily_run(run_hour, random_delay_max, include_today=True)

        existing_task = db_manager.get_scheduled_task_by_account(
            account_id,
            user_id=current_user['user_id'],
            task_type='item_polish'
        )

        if existing_task:
            updated = db_manager.update_scheduled_task(
                existing_task['id'],
                name=name,
                interval_hours=24,
                delay_minutes=run_hour,
                random_delay_max=random_delay_max,
                enabled=enabled,
                next_run_at=next_run_at
            )
            if updated:
                task = db_manager.get_scheduled_task(existing_task['id'])
                return {
                    "success": True,
                    "message": "定时擦亮任务更新成功",
                    "task_id": existing_task['id'],
                    "task": task
                }
            return {"success": False, "message": "更新定时任务失败"}

        task_id = db_manager.create_scheduled_task(
            name=name, task_type='item_polish', account_id=account_id,
            user_id=current_user['user_id'],
            interval_hours=24, delay_minutes=run_hour,
            random_delay_max=random_delay_max,
            next_run_at=next_run_at,
            enabled=enabled
        )

        if task_id:
            task = db_manager.get_scheduled_task(task_id)
            return {"success": True, "message": "定时擦亮任务创建成功", "task_id": task_id, "task": task}
        else:
            return {"success": False, "message": "创建定时任务失败"}
    except Exception as e:
        logger.error(f"创建定时任务异常: {str(e)}")
        return {"success": False, "message": f"创建定时任务异常: {str(e)}"}


@app.get("/scheduled-tasks")
async def list_scheduled_tasks(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取定时任务列表"""
    try:
        tasks = db_manager.get_scheduled_tasks(user_id=current_user['user_id'])
        return {"success": True, "tasks": tasks}
    except Exception as e:
        logger.error(f"获取定时任务列表异常: {str(e)}")
        return {"success": False, "message": f"获取定时任务列表异常: {str(e)}"}


@app.put("/scheduled-tasks/{task_id}")
async def update_scheduled_task(task_id: int, request: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新定时任务"""
    try:
        task = db_manager.get_scheduled_task(task_id)
        if not task:
            return {"success": False, "message": "任务不存在"}
        if task['user_id'] != current_user['user_id']:
            return {"success": False, "message": "无权修改此任务"}

        kwargs = {}

        if 'name' in request:
            name = str(request.get('name') or '').strip()
            if name:
                kwargs['name'] = name

        if 'interval_hours' in request:
            kwargs['interval_hours'] = int(request.get('interval_hours', task.get('interval_hours', 24)))

        if 'run_hour' in request or 'delay_minutes' in request:
            kwargs['delay_minutes'] = _parse_run_hour(request.get('run_hour', request.get('delay_minutes')))

        if 'random_delay_max' in request:
            kwargs['random_delay_max'] = _parse_random_delay(
                request.get('random_delay_max'),
                task.get('random_delay_max', 10)
            )

        if 'enabled' in request:
            kwargs['enabled'] = _parse_enabled_flag(request.get('enabled'))

        effective_enabled = kwargs.get('enabled', 1 if task['enabled'] else 0)
        effective_run_hour = kwargs.get('delay_minutes', task.get('delay_minutes', 8))
        effective_random_delay = kwargs.get('random_delay_max', task.get('random_delay_max', 10))

        if task['task_type'] == 'item_polish' and effective_enabled:
            should_reschedule = (
                'delay_minutes' in kwargs or
                'random_delay_max' in kwargs or
                ('enabled' in kwargs and not task['enabled'])
            )
            if should_reschedule:
                kwargs['next_run_at'] = db_manager.calculate_next_daily_run(
                    effective_run_hour,
                    effective_random_delay,
                    include_today=True
                )

        if not kwargs:
            return {"success": False, "message": "没有可更新的字段"}

        if db_manager.update_scheduled_task(task_id, **kwargs):
            updated_task = db_manager.get_scheduled_task(task_id)
            return {"success": True, "message": "定时任务更新成功", "task": updated_task}
        else:
            return {"success": False, "message": "更新失败"}
    except Exception as e:
        logger.error(f"更新定时任务异常: {str(e)}")
        return {"success": False, "message": f"更新定时任务异常: {str(e)}"}


@app.delete("/scheduled-tasks/{task_id}")
async def delete_scheduled_task(task_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除定时任务"""
    try:
        task = db_manager.get_scheduled_task(task_id)
        if not task:
            return {"success": False, "message": "任务不存在"}
        if task['user_id'] != current_user['user_id']:
            return {"success": False, "message": "无权删除此任务"}

        if db_manager.delete_scheduled_task(task_id):
            return {"success": True, "message": "定时任务已删除"}
        else:
            return {"success": False, "message": "删除失败"}
    except Exception as e:
        logger.error(f"删除定时任务异常: {str(e)}")
        return {"success": False, "message": f"删除定时任务异常: {str(e)}"}


@app.put("/scheduled-tasks/{task_id}/toggle")
async def toggle_scheduled_task(task_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """启用/禁用定时任务"""
    try:
        task = db_manager.get_scheduled_task(task_id)
        if not task:
            return {"success": False, "message": "任务不存在"}
        if task['user_id'] != current_user['user_id']:
            return {"success": False, "message": "无权操作此任务"}

        new_enabled = 0 if task['enabled'] else 1
        update_kwargs = {'enabled': new_enabled}
        if new_enabled:
            update_kwargs['next_run_at'] = db_manager.calculate_next_daily_run(
                task.get('delay_minutes', 8),
                task.get('random_delay_max', 10),
                include_today=True
            )

        if db_manager.update_scheduled_task(task_id, **update_kwargs):
            status = "启用" if new_enabled else "禁用"
            updated_task = db_manager.get_scheduled_task(task_id)
            return {
                "success": True,
                "message": f"定时任务已{status}",
                "enabled": bool(new_enabled),
                "task": updated_task
            }
        else:
            return {"success": False, "message": "操作失败"}
    except Exception as e:
        logger.error(f"切换定时任务状态异常: {str(e)}")
        return {"success": False, "message": f"操作异常: {str(e)}"}


# ==================== 定时任务调度器 ====================

async def scheduled_task_checker():
    """每60秒检查并执行到期的定时任务"""
    while True:
        try:
            due_tasks = db_manager.get_due_tasks()
            for task in due_tasks:
                try:
                    account_id = task['account_id']
                    task_id = task['id']
                    task_type = task['task_type']

                    logger.info(f"执行定时任务: {task['name']} (ID: {task_id}, 账号: {account_id})")

                    if task_type == 'item_polish':
                        cookie_info = db_manager.get_cookie_by_id(account_id)
                        if not cookie_info:
                            logger.warning(f"定时任务 {task_id} 账号 {account_id} 不存在，跳过")
                            result = {"success": False, "message": "账号不存在"}
                        else:
                            cookies_str = cookie_info.get('cookies_str', '')
                            if not cookies_str:
                                result = {"success": False, "message": "账号cookie为空"}
                            else:
                                from XianyuAutoAsync import XianyuLive
                                xianyu_instance = XianyuLive(cookies_str, account_id, register_instance=False)
                                result = await xianyu_instance.polish_all_items()
                                await xianyu_instance.close_session()
                    else:
                        result = {"success": False, "message": f"未知任务类型: {task_type}"}

                    run_hour = task.get('delay_minutes', 8)  # delay_minutes 复用为每日运行小时
                    random_max = task.get('random_delay_max', 10)
                    next_run_str = db_manager.calculate_next_daily_run(
                        run_hour,
                        random_max,
                        include_today=False
                    )

                    db_manager.update_task_run_result(task_id, result, next_run_str)
                    logger.info(f"定时任务 {task_id} 执行完毕，下次运行: {next_run_str}")

                except Exception as e:
                    logger.error(f"执行定时任务 {task.get('id')} 异常: {str(e)}")
        except Exception as e:
            logger.error(f"定时任务检查异常: {str(e)}")
        await asyncio.sleep(60)


# 移除自动启动，由Start.py或手动启动
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)
